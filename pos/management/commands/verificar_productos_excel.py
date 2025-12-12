# -*- coding: utf-8 -*-
"""
Comando para verificar que todos los productos del Excel estén creados con sus atributos
"""
from django.core.management.base import BaseCommand
from pos.models import Producto
from django.db.models import Q
import openpyxl
import os


class Command(BaseCommand):
    help = 'Verifica que todos los productos del Excel estén creados con sus atributos correctos'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel a verificar'
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return
        
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('VERIFICACIÓN DE PRODUCTOS DESDE EXCEL'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'Archivo: {archivo}')
        self.stdout.write('')
        
        try:
            # Cargar el archivo Excel
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Leer la primera fila como encabezados
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.stdout.write(f'Columnas encontradas: {headers}')
            self.stdout.write('')
            
            # Mapear columnas
            col_indices = {}
            for idx, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'codigo' in header_lower and 'barra' not in header_lower:
                        col_indices['codigo'] = idx
                    elif 'nombre' in header_lower and 'atributo' not in header_lower:
                        col_indices['nombre'] = idx
                    elif 'atributo' in header_lower or 'nombreatributo' in header_lower:
                        col_indices['atributo'] = idx
                    elif 'existencia' in header_lower or 'stock' in header_lower or 'cantidad' in header_lower:
                        col_indices['cantidad'] = idx
                    elif 'precio' in header_lower and 'compra' not in header_lower:
                        col_indices['precio'] = idx
            
            self.stdout.write(f'Columnas mapeadas: {col_indices}')
            self.stdout.write('')
            
            if 'codigo' not in col_indices:
                self.stdout.write(self.style.ERROR('ERROR: No se encontró la columna "codigo"'))
                return
            
            # Leer todas las filas de datos
            productos_excel = []
            productos_faltantes = []
            productos_atributo_incorrecto = []
            productos_ok = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    codigo = self._get_cell_value(row, col_indices.get('codigo'))
                    nombre = self._get_cell_value(row, col_indices.get('nombre'))
                    atributo = self._get_cell_value(row, col_indices.get('atributo'))
                    
                    if not codigo:
                        continue  # Saltar filas vacías
                    
                    codigo = str(codigo).strip()
                    nombre = str(nombre).strip() if nombre else ''
                    
                    # Atributo - mantener tal cual está en el Excel
                    if atributo:
                        atributo = str(atributo).strip()
                        if atributo.upper() == 'SIN ATRIBUTO' or atributo == '':
                            atributo = None
                    else:
                        atributo = None
                    
                    productos_excel.append({
                        'codigo': codigo,
                        'nombre': nombre,
                        'atributo': atributo,
                        'fila': row_num
                    })
                    
                    # Buscar producto en la base de datos
                    if atributo:
                        producto = Producto.objects.filter(
                            codigo=codigo,
                            atributo=atributo
                        ).first()
                    else:
                        producto = Producto.objects.filter(
                            codigo=codigo
                        ).filter(
                            Q(atributo__isnull=True) | Q(atributo='')
                        ).first()
                    
                    if producto:
                        # Verificar que el atributo coincida exactamente
                        if producto.atributo != atributo:
                            productos_atributo_incorrecto.append({
                                'codigo': codigo,
                                'nombre': nombre,
                                'atributo_excel': atributo,
                                'atributo_bd': producto.atributo,
                                'fila': row_num
                            })
                        else:
                            productos_ok.append({
                                'codigo': codigo,
                                'nombre': nombre,
                                'atributo': atributo,
                                'fila': row_num
                            })
                    else:
                        productos_faltantes.append({
                            'codigo': codigo,
                            'nombre': nombre,
                            'atributo': atributo,
                            'fila': row_num
                        })
                        
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'Error en fila {row_num}: {str(e)}'))
            
            # Mostrar resultados
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(self.style.SUCCESS('RESULTADOS DE LA VERIFICACIÓN'))
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(f'Total de productos en Excel: {len(productos_excel)}')
            self.stdout.write(f'Productos OK: {len(productos_ok)}')
            self.stdout.write(f'Productos faltantes: {len(productos_faltantes)}')
            self.stdout.write(f'Productos con atributo incorrecto: {len(productos_atributo_incorrecto)}')
            self.stdout.write('')
            
            # Mostrar productos faltantes
            if productos_faltantes:
                self.stdout.write(self.style.ERROR('=' * 70))
                self.stdout.write(self.style.ERROR('PRODUCTOS FALTANTES'))
                self.stdout.write(self.style.ERROR('=' * 70))
                for prod in productos_faltantes:
                    atributo_str = f" - Atributo: {prod['atributo']}" if prod['atributo'] else " - Sin atributo"
                    self.stdout.write(
                        f'  Fila {prod["fila"]}: {prod["codigo"]} - {prod["nombre"]}{atributo_str}'
                    )
                self.stdout.write('')
            
            # Mostrar productos con atributo incorrecto
            if productos_atributo_incorrecto:
                self.stdout.write(self.style.WARNING('=' * 70))
                self.stdout.write(self.style.WARNING('PRODUCTOS CON ATRIBUTO INCORRECTO'))
                self.stdout.write(self.style.WARNING('=' * 70))
                for prod in productos_atributo_incorrecto:
                    self.stdout.write(
                        f'  Fila {prod["fila"]}: {prod["codigo"]} - {prod["nombre"]}'
                    )
                    self.stdout.write(
                        f'    Excel: {prod["atributo_excel"] or "Sin atributo"}'
                    )
                    self.stdout.write(
                        f'    BD: {prod["atributo_bd"] or "Sin atributo"}'
                    )
                self.stdout.write('')
            
            # Mostrar muestra de productos OK
            if productos_ok:
                self.stdout.write(self.style.SUCCESS('Muestra de productos OK:'))
                for prod in productos_ok[:5]:
                    atributo_str = f" - Atributo: {prod['atributo']}" if prod['atributo'] else " - Sin atributo"
                    self.stdout.write(
                        f'  {prod["codigo"]} - {prod["nombre"]}{atributo_str}'
                    )
                if len(productos_ok) > 5:
                    self.stdout.write(f'  ... y {len(productos_ok) - 5} más')
                self.stdout.write('')
            
            # Resumen final
            if productos_faltantes or productos_atributo_incorrecto:
                self.stdout.write('')
                self.stdout.write(self.style.WARNING('RECOMENDACIÓN:'))
                self.stdout.write('Ejecuta el comando de importación con --actualizar para crear/actualizar productos:')
                self.stdout.write(f'  python manage.py importar_productos_excel "{archivo}" --actualizar --confirm')
            else:
                self.stdout.write('')
                self.stdout.write(self.style.SUCCESS('[OK] Todos los productos del Excel estan creados correctamente con sus atributos'))
            
        except Exception as e:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('=' * 70))
            self.stdout.write(self.style.ERROR('ERROR AL VERIFICAR'))
            self.stdout.write(self.style.ERROR('=' * 70))
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
            import traceback
            self.stdout.write(self.style.ERROR(traceback.format_exc()))
    
    def _get_cell_value(self, row, col_index):
        """Obtener valor de una celda de forma segura"""
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        if cell is None:
            return None
        return cell.value


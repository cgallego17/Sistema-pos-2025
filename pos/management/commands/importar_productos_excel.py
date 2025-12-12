# -*- coding: utf-8 -*-
"""
Comando para importar productos desde un archivo Excel con imágenes desde API
"""
from django.core.management.base import BaseCommand
from django.core.files.base import ContentFile
from django.db import transaction, models
from django.db.models import Q
from pos.models import Producto
import openpyxl
import os
import requests
from urllib.parse import urlparse


class Command(BaseCommand):
    help = 'Importa productos desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel a importar'
        )
        parser.add_argument(
            '--actualizar',
            action='store_true',
            help='Actualizar productos existentes si el código ya existe',
        )
        parser.add_argument(
            '--confirm',
            action='store_true',
            help='Confirmar sin preguntar',
        )
        parser.add_argument(
            '--codigos',
            type=str,
            nargs='+',
            help='Lista de códigos de productos a importar (solo estos)',
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return
        
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('IMPORTACIÓN DE PRODUCTOS DESDE EXCEL'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'Archivo: {archivo}')
        self.stdout.write('')
        
        try:
            # Cargar el archivo Excel con data_only=True para obtener valores calculados
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Leer la primera fila como encabezados
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.stdout.write(f'Columnas encontradas: {headers}')
            self.stdout.write('')
            
            # Mapear columnas (buscar por nombre, case-insensitive)
            col_indices = {}
            for idx, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if header_lower == 'id':
                        col_indices['id'] = idx
                    elif 'codigo' in header_lower and 'barra' not in header_lower:
                        col_indices['codigo'] = idx
                    elif 'nombre' in header_lower and 'atributo' not in header_lower:
                        col_indices['nombre'] = idx
                    elif 'nombreatributo' in header_lower:
                        col_indices['atributo'] = idx
                    elif 'atributo' in header_lower and 'id' not in header_lower and 'nombre' not in header_lower:
                        # Solo usar 'atributo' si no hay 'nombreAtributo'
                        if 'atributo' not in col_indices:
                            col_indices['atributo'] = idx
                    elif 'existencia' in header_lower or 'stock' in header_lower:
                        col_indices['stock'] = idx
                    elif 'precio' in header_lower:
                        col_indices['precio'] = idx
                    elif 'barra' in header_lower or 'codigo_barra' in header_lower:
                        col_indices['codigo_barras'] = idx
            
            self.stdout.write(f'Columnas mapeadas: {col_indices}')
            self.stdout.write('')
            
            if 'id' not in col_indices:
                self.stdout.write(self.style.ERROR('ERROR: No se encontró la columna requerida (id)'))
                return
            
            if 'codigo' not in col_indices or 'nombre' not in col_indices:
                self.stdout.write(self.style.ERROR('ERROR: No se encontraron las columnas requeridas (codigo, nombre)'))
                return
            
            # Obtener lista de códigos a filtrar (si se especifica)
            codigos_filtro = None
            if options.get('codigos'):
                codigos_filtro = [c.strip().upper() for c in options['codigos']]
                self.stdout.write(f'Filtrando solo estos códigos: {codigos_filtro}')
                self.stdout.write('')
            
            # Leer todas las filas de datos
            productos_a_importar = []
            filas_con_error = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    # Obtener valores de las celdas
                    producto_id = self._get_cell_value(row, col_indices.get('id'))
                    codigo = self._get_cell_value(row, col_indices.get('codigo'))
                    nombre = self._get_cell_value(row, col_indices.get('nombre'))
                    atributo = self._get_cell_value(row, col_indices.get('atributo'))
                    precio = self._get_cell_value(row, col_indices.get('precio'))
                    codigo_barras = self._get_cell_value(row, col_indices.get('codigo_barras'))
                    
                    # Validar campos requeridos
                    if not producto_id or not codigo or not nombre:
                        continue  # Saltar filas vacías
                    
                    # Filtrar por códigos si se especificó
                    if codigos_filtro and codigo.strip().upper() not in codigos_filtro:
                        continue  # Saltar productos que no están en la lista
                    
                    # Limpiar y convertir valores
                    producto_id = str(producto_id).strip()
                    codigo = str(codigo).strip()
                    nombre = str(nombre).strip()
                    
                    # Atributo
                    if atributo:
                        atributo = str(atributo).strip()
                        if atributo.upper() == 'SIN ATRIBUTO' or atributo == '':
                            atributo = None
                    else:
                        atributo = None
                    
                    # Stock siempre será 0 (no montar stock)
                    stock = 0
                    
                    # Precio
                    if precio is None:
                        precio = 0
                    else:
                        try:
                            precio = int(float(precio))
                            if precio < 0:
                                precio = 0
                        except (ValueError, TypeError):
                            precio = 0
                    
                    # Código de barras
                    if codigo_barras:
                        try:
                            # Convertir a string y limpiar
                            codigo_barras = str(int(float(codigo_barras))).strip()
                            if codigo_barras == '0' or codigo_barras == '':
                                codigo_barras = None
                        except (ValueError, TypeError):
                            codigo_barras = None
                    else:
                        codigo_barras = None
                    
                    productos_a_importar.append({
                        'id': producto_id,
                        'codigo': codigo,
                        'nombre': nombre,
                        'atributo': atributo,
                        'stock': stock,
                        'precio': precio,
                        'codigo_barras': codigo_barras,
                        'fila': row_num
                    })
                    
                except Exception as e:
                    filas_con_error.append({
                        'fila': row_num,
                        'error': str(e)
                    })
            
            self.stdout.write(f'Productos encontrados en el archivo: {len(productos_a_importar)}')
            if filas_con_error:
                self.stdout.write(self.style.WARNING(f'Filas con errores: {len(filas_con_error)}'))
            self.stdout.write('')
            
            if not productos_a_importar:
                self.stdout.write(self.style.WARNING('No se encontraron productos para importar'))
                return
            
            # Mostrar muestra de productos
            self.stdout.write('Muestra de productos a importar:')
            for i, prod in enumerate(productos_a_importar[:5], 1):
                self.stdout.write(f'  {i}. ID: {prod["id"]} - {prod["codigo"]} - {prod["nombre"]} (Atributo: {prod["atributo"] or "N/A"}, Precio: ${prod["precio"]:,})')
            if len(productos_a_importar) > 5:
                self.stdout.write(f'  ... y {len(productos_a_importar) - 5} más')
            self.stdout.write('')
            
            # Confirmación
            if not options['confirm']:
                respuesta = input('¿Deseas continuar con la importación? (si/no): ')
                if respuesta.lower() != 'si':
                    self.stdout.write(self.style.WARNING('Importación cancelada'))
                    return
            
            # Descargar productos del API
            self.stdout.write('Descargando productos del API...', ending=' ')
            api_url = 'https://catalogo.tersacosmeticos.com/prod/api/productos-publicos/?format=json'
            base_url_imagenes = 'https://catalogo.tersacosmeticos.com'
            productos_api_dict = {}
            
            try:
                response = requests.get(api_url, timeout=30)
                if response.status_code == 200:
                    productos_api = response.json()
                    if isinstance(productos_api, list):
                        for prod_api in productos_api:
                            if isinstance(prod_api, dict) and 'id' in prod_api:
                                producto_id_api = str(prod_api['id'])
                                productos_api_dict[producto_id_api] = prod_api
                        self.stdout.write(self.style.SUCCESS(f'OK ({len(productos_api_dict)} productos encontrados)'))
                    else:
                        self.stdout.write(self.style.ERROR('La API no devolvió un array'))
                        return
                else:
                    self.stdout.write(self.style.ERROR(f'Error HTTP {response.status_code}'))
                    return
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
                return
            
            self.stdout.write('')
            
            # Importar productos
            self.stdout.write(self.style.WARNING('Iniciando importación...'))
            
            productos_creados = 0
            productos_actualizados = 0
            productos_con_error = 0
            productos_sin_imagen = 0
            
            # Procesar cada producto individualmente para evitar que un error afecte a los demás
            for prod_data in productos_a_importar:
                try:
                    with transaction.atomic():
                        # Buscar producto por código Y atributo
                        if prod_data['atributo']:
                            producto = Producto.objects.filter(
                                codigo=prod_data['codigo'],
                                atributo=prod_data['atributo']
                            ).first()
                        else:
                            producto = Producto.objects.filter(
                                codigo=prod_data['codigo']
                            ).filter(
                                Q(atributo__isnull=True) | Q(atributo='')
                            ).first()
                        
                        # Buscar imagen en el API
                        imagen_url = None
                        prod_api = productos_api_dict.get(prod_data['id'])
                        if prod_api and isinstance(prod_api, dict):
                            imagen_path = prod_api.get('imagen') or prod_api.get('imagen_url') or prod_api.get('url_imagen')
                            if imagen_path:
                                if imagen_path.startswith('http://') or imagen_path.startswith('https://'):
                                    imagen_url = imagen_path
                                elif imagen_path.startswith('/'):
                                    imagen_url = f"{base_url_imagenes}{imagen_path}"
                                else:
                                    imagen_url = f"{base_url_imagenes}/{imagen_path}"
                        
                        if producto:
                            # Producto ya existe con este código y atributo exactos
                            if options['actualizar']:
                                producto.nombre = prod_data['nombre']
                                producto.precio = prod_data['precio']
                                producto.stock = 0  # No montar stock
                                # Solo actualizar código de barras si no existe en otro producto
                                if prod_data['codigo_barras']:
                                    if not Producto.objects.filter(codigo_barras=prod_data['codigo_barras']).exclude(id=producto.id).exists():
                                        producto.codigo_barras = prod_data['codigo_barras']
                                producto.activo = True
                                
                                # Actualizar imagen si se encontró en el API
                                if imagen_url and (not producto.imagen or options['actualizar']):
                                    try:
                                        img_response = requests.get(imagen_url, timeout=30, stream=True)
                                        if img_response.status_code == 200:
                                            content_type = img_response.headers.get('content-type', '')
                                            if content_type.startswith('image/'):
                                                parsed_url = urlparse(imagen_url)
                                                path = parsed_url.path
                                                ext = os.path.splitext(path)[1] or '.jpg'
                                                filename = f"{producto.codigo}{ext}"
                                                producto.imagen.save(
                                                    filename,
                                                    ContentFile(img_response.content),
                                                    save=False
                                                )
                                    except Exception:
                                        pass  # Si falla la descarga de imagen, continuar
                                
                                producto.save()
                                productos_actualizados += 1
                            # Si no es actualizar, no hacer nada (producto ya existe)
                        else:
                            # No existe producto con este código Y atributo exactos
                            # Verificar si el código de barras ya existe antes de crear
                            codigo_barras_final = prod_data['codigo_barras']
                            if codigo_barras_final and Producto.objects.filter(codigo_barras=codigo_barras_final).exists():
                                codigo_barras_final = None
                            
                            producto = Producto.objects.create(
                                codigo=prod_data['codigo'],
                                nombre=prod_data['nombre'],
                                atributo=prod_data['atributo'],
                                precio=prod_data['precio'],
                                stock=0,  # No montar stock
                                codigo_barras=codigo_barras_final,
                                activo=True
                            )
                            
                            # Descargar y guardar imagen si se encontró en el API
                            if imagen_url:
                                try:
                                    img_response = requests.get(imagen_url, timeout=30, stream=True)
                                    if img_response.status_code == 200:
                                        content_type = img_response.headers.get('content-type', '')
                                        if content_type.startswith('image/'):
                                            parsed_url = urlparse(imagen_url)
                                            path = parsed_url.path
                                            ext = os.path.splitext(path)[1] or '.jpg'
                                            filename = f"{producto.codigo}{ext}"
                                            producto.imagen.save(
                                                filename,
                                                ContentFile(img_response.content),
                                                save=True
                                            )
                                        else:
                                            productos_sin_imagen += 1
                                    else:
                                        productos_sin_imagen += 1
                                except Exception as e:
                                    productos_sin_imagen += 1
                            else:
                                productos_sin_imagen += 1
                            
                            productos_creados += 1
                                
                except Exception as e:
                    productos_con_error += 1
                    self.stdout.write(self.style.ERROR(f'Error en fila {prod_data["fila"]} (Código: {prod_data["codigo"]}): {str(e)}'))
            
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(self.style.SUCCESS('IMPORTACIÓN COMPLETADA'))
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(f'Productos creados: {productos_creados}')
            if options['actualizar']:
                self.stdout.write(f'Productos actualizados: {productos_actualizados}')
            self.stdout.write(f'Productos sin imagen: {productos_sin_imagen}')
            self.stdout.write(f'Productos con error: {productos_con_error}')
            
            if filas_con_error:
                self.stdout.write('')
                self.stdout.write(self.style.WARNING('Filas con errores:'))
                for error in filas_con_error[:10]:
                    self.stdout.write(f'  Fila {error["fila"]}: {error["error"]}')
                if len(filas_con_error) > 10:
                    self.stdout.write(f'  ... y {len(filas_con_error) - 10} más')
            
        except Exception as e:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('=' * 70))
            self.stdout.write(self.style.ERROR('ERROR AL IMPORTAR'))
            self.stdout.write(self.style.ERROR('=' * 70))
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
            import traceback
            self.stdout.write(self.style.ERROR(traceback.format_exc()))
    
    def _get_cell_value(self, row, col_index):
        """Obtener valor de una celda de forma segura"""
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        if cell is None:
            return None
        return cell.value



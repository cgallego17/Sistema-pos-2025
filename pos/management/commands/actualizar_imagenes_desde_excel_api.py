# -*- coding: utf-8 -*-
"""
Comando para actualizar imágenes de productos desde una API usando IDs del Excel
"""
from django.core.management.base import BaseCommand
from django.core.files.base import ContentFile
from pos.models import Producto
import requests
import openpyxl
import os
from urllib.parse import urlparse


class Command(BaseCommand):
    help = 'Actualiza las imágenes de productos desde una API usando IDs del Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel con los IDs de productos'
        )
        parser.add_argument(
            '--api-url',
            type=str,
            default='https://catalogo.tersacosmeticos.com/prod/api/productos-publicos/',
            help='URL de la API que devuelve todos los productos (default: https://catalogo.tersacosmeticos.com/prod/api/productos-publicos/)',
        )
        parser.add_argument(
            '--base-url-imagenes',
            type=str,
            default='https://catalogo.tersacosmeticos.com',
            help='URL base para construir URLs completas de imágenes si son relativas (default: https://catalogo.tersacosmeticos.com)',
        )
        parser.add_argument(
            '--id-field',
            type=str,
            default='id',
            help='Nombre de la columna en el Excel que contiene el ID (default: id)',
        )
        parser.add_argument(
            '--codigo-field',
            type=str,
            default='codigo',
            help='Nombre de la columna en el Excel que contiene el código (default: codigo)',
        )
        parser.add_argument(
            '--imagen-field',
            type=str,
            default='imagen',
            help='Nombre del campo en la respuesta JSON que contiene la URL de la imagen (default: imagen)',
        )
        parser.add_argument(
            '--sobrescribir',
            action='store_true',
            help='Sobrescribir imágenes existentes',
        )
        parser.add_argument(
            '--timeout',
            type=int,
            default=30,
            help='Timeout en segundos para las peticiones HTTP (default: 30)',
        )
        parser.add_argument(
            '--headers',
            type=str,
            default='',
            help='Headers adicionales en formato JSON (ej: {"Authorization": "Bearer token"})',
        )

    def _get_cell_value(self, row, col_index):
        """Obtener valor de una celda de forma segura"""
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        if cell is None:
            return None
        return cell.value

    def handle(self, *args, **options):
        archivo = options['archivo']
        api_url = options['api_url']
        base_url_imagenes = options['base_url_imagenes']
        id_field = options['id_field']
        codigo_field = options['codigo_field']
        imagen_field = options['imagen_field']
        sobrescribir = options['sobrescribir']
        timeout = options['timeout']
        headers_str = options['headers']
        
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return
        
        # Parsear headers si se proporcionan
        headers = {}
        if headers_str:
            import json
            try:
                headers = json.loads(headers_str)
            except json.JSONDecodeError:
                self.stdout.write(self.style.ERROR('Error: Los headers deben estar en formato JSON válido'))
                return
        
        # Si no se proporciona base_url_imagenes, extraerla de api_url
        if not base_url_imagenes or base_url_imagenes == '':
            parsed = urlparse(api_url)
            base_url_imagenes = f"{parsed.scheme}://{parsed.netloc}"
        
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('ACTUALIZAR IMÁGENES DESDE EXCEL Y API'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'Archivo Excel: {archivo}')
        self.stdout.write(f'URL de la API: {api_url}')
        self.stdout.write(f'URL base para imágenes: {base_url_imagenes}')
        self.stdout.write(f'Columna ID: {id_field}')
        self.stdout.write(f'Columna Código: {codigo_field}')
        self.stdout.write(f'Campo de imagen en API: {imagen_field}')
        self.stdout.write('')
        
        # Leer Excel
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            headers_excel = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.stdout.write(f'Columnas encontradas en Excel: {headers_excel}')
            self.stdout.write('')
            
            # Mapear columnas
            col_indices = {}
            for idx, header in enumerate(headers_excel):
                if header:
                    header_lower = str(header).lower().strip()
                    if header_lower == id_field.lower():
                        col_indices['id'] = idx
                    elif 'codigo' in header_lower and 'barra' not in header_lower:
                        col_indices['codigo'] = idx
            
            if 'id' not in col_indices:
                self.stdout.write(self.style.ERROR(f'ERROR: No se encontró la columna "{id_field}" en el Excel'))
                return
            
            if 'codigo' not in col_indices:
                self.stdout.write(self.style.WARNING('ADVERTENCIA: No se encontró la columna de código. Se buscará solo por ID.'))
            
            productos_excel = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    producto_id = self._get_cell_value(row, col_indices.get('id'))
                    codigo = self._get_cell_value(row, col_indices.get('codigo'))
                    
                    if not producto_id:
                        continue
                    
                    # Convertir ID a string
                    producto_id = str(producto_id).strip()
                    codigo = str(codigo).strip() if codigo else None
                    
                    productos_excel.append({
                        'id': producto_id,
                        'codigo': codigo,
                        'fila': row_num
                    })
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'Error en fila {row_num}: {str(e)}'))
            
            self.stdout.write(f'Productos encontrados en Excel: {len(productos_excel)}')
            self.stdout.write('')
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error al leer Excel: {str(e)}'))
            import traceback
            self.stdout.write(self.style.ERROR(traceback.format_exc()))
            return
        
        # Descargar todos los productos de la API una sola vez
        self.stdout.write('Descargando productos de la API...', ending=' ')
        productos_api_dict = {}
        try:
            # Agregar ?format=json si no está presente
            api_url_final = api_url
            if '?format=json' not in api_url_final:
                api_url_final = f"{api_url_final.rstrip('/')}/?format=json"
            
            response = requests.get(api_url_final, headers=headers, timeout=timeout)
            if response.status_code != 200:
                self.stdout.write(self.style.ERROR(f'Error HTTP {response.status_code}'))
                return
            
            productos_api = response.json()
            if not isinstance(productos_api, list):
                self.stdout.write(self.style.ERROR('La API no devolvió un array de productos'))
                return
            
            # Indexar por ID
            for prod_api in productos_api:
                if isinstance(prod_api, dict) and 'id' in prod_api:
                    producto_id_api = str(prod_api['id'])
                    productos_api_dict[producto_id_api] = prod_api
            
            self.stdout.write(self.style.SUCCESS(f'OK ({len(productos_api_dict)} productos encontrados)'))
            self.stdout.write('')
            
        except requests.exceptions.RequestException as e:
            self.stdout.write(self.style.ERROR(f'Error de conexión: {str(e)}'))
            return
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error al procesar respuesta: {str(e)}'))
            return
        
        productos_actualizados = 0
        productos_con_error = 0
        productos_sin_imagen_api = 0
        productos_no_encontrados = 0
        productos_sin_imagen_local = 0
        
        # Función para buscar URL de imagen en respuesta JSON
        def buscar_imagen_en_objeto(obj):
            """Buscar URL de imagen en un objeto, probando diferentes campos"""
            if not isinstance(obj, dict):
                return None
            
            # Buscar en el campo especificado
            if imagen_field in obj and obj[imagen_field]:
                return obj[imagen_field]
            
            # Buscar en campos comunes alternativos
            campos_posibles = [
                'imagen_url', 'url_imagen', 'image_url', 'url_image',
                'imagen', 'image', 'foto', 'photo', 'picture',
                'url', 'link', 'src', 'thumbnail', 'thumb'
            ]
            for campo in campos_posibles:
                if campo in obj and obj[campo]:
                    return obj[campo]
            return None
        
        # Procesar cada producto del Excel
        for prod_excel in productos_excel:
            try:
                producto_id_api = prod_excel['id']
                codigo = prod_excel['codigo']
                fila = prod_excel['fila']
                
                # Buscar producto en la base de datos
                producto = None
                if codigo:
                    # Buscar por código
                    producto = Producto.objects.filter(codigo=codigo).first()
                
                if not producto:
                    self.stdout.write(self.style.WARNING(f'Fila {fila}: Producto con ID {producto_id_api} no encontrado en BD (código: {codigo or "N/A"})'))
                    productos_no_encontrados += 1
                    continue
                
                # Verificar si ya tiene imagen
                if producto.imagen and not sobrescribir:
                    self.stdout.write(f'Fila {fila}: {producto.codigo} - {producto.nombre[:40]}... Ya tiene imagen (omitir)')
                    productos_sin_imagen_local += 1
                    continue
                
                self.stdout.write(f'Fila {fila}: {producto.codigo} - {producto.nombre[:40]}...', ending=' ')
                
                # Buscar el producto en el diccionario de la API por ID
                if producto_id_api not in productos_api_dict:
                    self.stdout.write(self.style.WARNING('No encontrado en API'))
                    productos_no_encontrados += 1
                    continue
                
                prod_api = productos_api_dict[producto_id_api]
                
                # Buscar URL de imagen
                imagen_path = buscar_imagen_en_objeto(prod_api)
                
                if not imagen_path:
                    self.stdout.write(self.style.WARNING('Sin imagen en API'))
                    productos_sin_imagen_api += 1
                    continue
                
                # Construir URL completa de la imagen
                if imagen_path.startswith('http://') or imagen_path.startswith('https://'):
                    imagen_url = imagen_path
                elif imagen_path.startswith('/'):
                    imagen_url = f"{base_url_imagenes}{imagen_path}"
                else:
                    imagen_url = f"{base_url_imagenes}/{imagen_path}"
                
                # Descargar la imagen
                try:
                    img_response = requests.get(imagen_url, headers=headers, timeout=timeout, stream=True)
                    
                    if img_response.status_code == 200:
                        # Obtener extensión del archivo
                        parsed_url = urlparse(imagen_url)
                        path = parsed_url.path
                        ext = os.path.splitext(path)[1] or '.jpg'
                        
                        # Validar que sea una imagen
                        content_type = img_response.headers.get('content-type', '')
                        if not content_type.startswith('image/'):
                            self.stdout.write(self.style.ERROR('No es una imagen válida'))
                            productos_con_error += 1
                            continue
                        
                        # Crear nombre de archivo
                        filename = f"productos/{producto.codigo}{ext}"
                        
                        # Guardar la imagen
                        producto.imagen.save(
                            filename,
                            ContentFile(img_response.content),
                            save=True
                        )
                        
                        productos_actualizados += 1
                        self.stdout.write(self.style.SUCCESS('OK'))
                    else:
                        self.stdout.write(self.style.ERROR(f'Error HTTP {img_response.status_code} al descargar imagen'))
                        productos_con_error += 1
                        
                except requests.exceptions.RequestException as e:
                    self.stdout.write(self.style.ERROR(f'Error de conexión: {str(e)[:50]}'))
                    productos_con_error += 1
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Error: {str(e)[:50]}'))
                    productos_con_error += 1
                    
            except Exception as e:
                self.stdout.write(self.style.ERROR(f'Error procesando fila {prod_excel.get("fila", "?")}: {str(e)[:50]}'))
                productos_con_error += 1
        
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('PROCESO COMPLETADO'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'Productos actualizados: {productos_actualizados}')
        self.stdout.write(f'Productos no encontrados en BD: {productos_no_encontrados}')
        self.stdout.write(f'Productos sin imagen en API: {productos_sin_imagen_api}')
        self.stdout.write(f'Productos que ya tenían imagen (omitidos): {productos_sin_imagen_local}')
        self.stdout.write(f'Productos con error: {productos_con_error}')
        self.stdout.write('')


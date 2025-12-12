# -*- coding: utf-8 -*-
"""
Comando para listar productos del Excel que no se encontraron en la base de datos
"""
from django.core.management.base import BaseCommand
from django.db.models import Q
from pos.models import Producto
import openpyxl
import os


class Command(BaseCommand):
    help = 'Lista los productos del Excel que no se encontraron en la base de datos'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel a verificar'
        )

    def _get_cell_value(self, row, col_index):
        """Obtener valor de una celda de forma segura"""
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        if cell is None:
            return None
        return cell.value

    def handle(self, *args, **options):
        archivo = options['archivo']
        
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            col_indices = {}
            for idx, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'codigo' in header_lower and 'barra' not in header_lower:
                        col_indices['codigo'] = idx
                    elif 'atributo' in header_lower or 'nombreatributo' in header_lower:
                        col_indices['atributo'] = idx
                    elif 'existencia' in header_lower or 'stock' in header_lower or 'cantidad' in header_lower:
                        col_indices['cantidad'] = idx
            
            no_encontrados = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                codigo = self._get_cell_value(row, col_indices.get('codigo'))
                atributo = self._get_cell_value(row, col_indices.get('atributo'))
                cantidad = self._get_cell_value(row, col_indices.get('cantidad'))
                
                if not codigo:
                    continue
                
                codigo = str(codigo).strip()
                
                if atributo:
                    atributo = str(atributo).strip()
                    if atributo.upper() == 'SIN ATRIBUTO' or atributo == '':
                        atributo = None
                else:
                    atributo = None
                
                # Buscar producto
                if atributo:
                    producto = Producto.objects.filter(codigo=codigo, atributo=atributo).first()
                else:
                    producto = Producto.objects.filter(codigo=codigo).filter(
                        Q(atributo__isnull=True) | Q(atributo='')
                    ).first()
                
                if not producto:
                    no_encontrados.append({
                        'fila': row_num,
                        'codigo': codigo,
                        'atributo': atributo,
                        'cantidad': cantidad
                    })
            
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(self.style.SUCCESS('PRODUCTOS NO ENCONTRADOS EN LA BASE DE DATOS'))
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(f'Total: {len(no_encontrados)} productos')
            self.stdout.write('')
            
            if no_encontrados:
                for p in no_encontrados:
                    atributo_str = p['atributo'] if p['atributo'] else 'SIN ATRIBUTO'
                    self.stdout.write(f'  Fila {p["fila"]}: {p["codigo"]} - Atributo: {atributo_str} - Cantidad: {p["cantidad"]}')
            else:
                self.stdout.write(self.style.SUCCESS('Todos los productos del Excel están en la base de datos'))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
            import traceback
            self.stdout.write(self.style.ERROR(traceback.format_exc()))



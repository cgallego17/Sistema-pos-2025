# -*- coding: utf-8 -*-
"""
Comando para verificar que las cantidades del ingreso coincidan con las del Excel
"""
from django.core.management.base import BaseCommand
from django.db.models import Q
from pos.models import IngresoMercancia, ItemIngresoMercancia, Producto
import openpyxl
import os


class Command(BaseCommand):
    help = 'Verifica que las cantidades del ingreso coincidan con las del Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            'ingreso_id',
            type=int,
            help='ID del ingreso a verificar'
        )
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel'
        )

    def _get_cell_value(self, row, col_index):
        """Obtener valor de una celda de forma segura"""
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        if cell is None:
            return None
        return cell.value

    def handle(self, *args, **options):
        ingreso_id = options['ingreso_id']
        archivo = options['archivo']
        
        try:
            ingreso = IngresoMercancia.objects.get(id=ingreso_id)
        except IngresoMercancia.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'Ingreso #{ingreso_id} no existe'))
            return
        
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return
        
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('VERIFICACIÓN DE CANTIDADES INGRESO vs EXCEL'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'Ingreso ID: #{ingreso.id}')
        self.stdout.write(f'Archivo Excel: {archivo}')
        self.stdout.write('')
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            col_indices = {}
            for idx, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'codigo' in header_lower and 'barra' not in header_lower:
                        col_indices['codigo'] = idx
                    elif 'atributo' in header_lower or 'nombreatributo' in header_lower:
                        col_indices['atributo'] = idx
                    elif 'existencia' in header_lower or 'stock' in header_lower or 'cantidad' in header_lower:
                        col_indices['cantidad'] = idx
            
            # Crear diccionario de cantidades del Excel
            cantidades_excel = {}
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                codigo = self._get_cell_value(row, col_indices.get('codigo'))
                atributo = self._get_cell_value(row, col_indices.get('atributo'))
                cantidad = self._get_cell_value(row, col_indices.get('cantidad'))
                
                if not codigo:
                    continue
                
                codigo = str(codigo).strip()
                
                if atributo:
                    atributo = str(atributo).strip()
                    if atributo.upper() == 'SIN ATRIBUTO' or atributo == '':
                        atributo = None
                else:
                    atributo = None
                
                if cantidad is None:
                    cantidad = 0
                else:
                    try:
                        cantidad = int(float(cantidad))
                    except (ValueError, TypeError):
                        cantidad = 0
                
                # Crear clave única: código + atributo
                clave = f"{codigo}|{atributo or ''}"
                if clave in cantidades_excel:
                    # Producto duplicado (mismo código y mismo atributo) - solo registrar, NO sumar
                    cantidades_excel[clave]['duplicado'] = True
                    cantidades_excel[clave]['filas'].append(row_num)
                    # Mantener la primera cantidad encontrada, no sumar
                else:
                    cantidades_excel[clave] = {
                        'codigo': codigo,
                        'atributo': atributo,
                        'cantidad': cantidad,
                        'filas': [row_num],
                        'duplicado': False
                    }
            
            # Identificar productos duplicados
            productos_duplicados = [datos for datos in cantidades_excel.values() if datos.get('duplicado', False)]
            
            self.stdout.write(f'Productos únicos en Excel: {len(cantidades_excel)}')
            self.stdout.write(f'Productos duplicados en Excel: {len(productos_duplicados)}')
            self.stdout.write(f'Items en ingreso: {ingreso.items.count()}')
            self.stdout.write('')
            
            if productos_duplicados:
                self.stdout.write(self.style.WARNING('=' * 70))
                self.stdout.write(self.style.WARNING('PRODUCTOS DUPLICADOS EN EXCEL (mismo código y mismo atributo)'))
                self.stdout.write(self.style.WARNING('=' * 70))
                for dup in productos_duplicados:
                    atributo_str = dup['atributo'] if dup['atributo'] else 'SIN ATRIBUTO'
                    filas_str = ', '.join(map(str, dup['filas']))
                    self.stdout.write(
                        f'  {dup["codigo"]} - Atributo: {atributo_str} - Filas: {filas_str} - Cantidad: {dup["cantidad"]}'
                    )
                self.stdout.write('')
            
            # Verificar items del ingreso
            items_ok = []
            items_diferentes = []
            items_no_en_excel = []
            
            for item in ingreso.items.all():
                codigo = item.producto.codigo
                atributo = item.producto.atributo
                cantidad_ingreso = item.cantidad
                
                clave = f"{codigo}|{atributo or ''}"
                
                if clave in cantidades_excel:
                    cantidad_excel = cantidades_excel[clave]['cantidad']
                    if cantidad_ingreso == cantidad_excel:
                        items_ok.append({
                            'item': item,
                            'cantidad_excel': cantidad_excel
                        })
                    else:
                        items_diferentes.append({
                            'item': item,
                            'cantidad_ingreso': cantidad_ingreso,
                            'cantidad_excel': cantidad_excel,
                            'filas': cantidades_excel[clave]['filas']
                        })
                else:
                    items_no_en_excel.append({
                        'item': item
                    })
            
            # Verificar productos del Excel que no están en el ingreso
            items_en_excel_no_en_ingreso = []
            for clave, datos in cantidades_excel.items():
                codigo = datos['codigo']
                atributo = datos['atributo']
                cantidad = datos['cantidad']
                
                # Buscar si existe en el ingreso
                existe = False
                for item in ingreso.items.all():
                    if item.producto.codigo == codigo and item.producto.atributo == atributo:
                        existe = True
                        break
                
                if not existe:
                    items_en_excel_no_en_ingreso.append(datos)
            
            # Mostrar resultados
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(self.style.SUCCESS('RESULTADOS DE LA VERIFICACIÓN'))
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(f'Items con cantidad correcta: {len(items_ok)}')
            self.stdout.write(f'Items con cantidad diferente: {len(items_diferentes)}')
            self.stdout.write(f'Items en ingreso no encontrados en Excel: {len(items_no_en_excel)}')
            self.stdout.write(f'Items en Excel no encontrados en ingreso: {len(items_en_excel_no_en_ingreso)}')
            self.stdout.write('')
            
            if items_diferentes:
                self.stdout.write(self.style.WARNING('=' * 70))
                self.stdout.write(self.style.WARNING('ITEMS CON CANTIDAD DIFERENTE'))
                self.stdout.write(self.style.WARNING('=' * 70))
                for diff in items_diferentes:
                    filas_str = ', '.join(map(str, diff["filas"]))
                    self.stdout.write(
                        f'  {diff["item"].producto.codigo} - {diff["item"].producto.nombre[:50]}'
                    )
                    self.stdout.write(
                        f'    Filas Excel: {filas_str} | Excel: {diff["cantidad_excel"]} | Ingreso: {diff["cantidad_ingreso"]} | Diferencia: {diff["cantidad_excel"] - diff["cantidad_ingreso"]}'
                    )
                self.stdout.write('')
            
            if items_no_en_excel:
                self.stdout.write(self.style.ERROR('=' * 70))
                self.stdout.write(self.style.ERROR('ITEMS EN INGRESO NO ENCONTRADOS EN EXCEL'))
                self.stdout.write(self.style.ERROR('=' * 70))
                for item_data in items_no_en_excel:
                    item = item_data['item']
                    self.stdout.write(
                        f'  {item.producto.codigo} - {item.producto.nombre[:50]} - Cantidad: {item.cantidad}'
                    )
                self.stdout.write('')
            
            if items_en_excel_no_en_ingreso:
                self.stdout.write(self.style.WARNING('=' * 70))
                self.stdout.write(self.style.WARNING('ITEMS EN EXCEL NO ENCONTRADOS EN INGRESO'))
                self.stdout.write(self.style.WARNING('=' * 70))
                for datos in items_en_excel_no_en_ingreso:
                    atributo_str = datos['atributo'] if datos['atributo'] else 'SIN ATRIBUTO'
                    filas_str = ', '.join(map(str, datos['filas']))
                    self.stdout.write(
                        f'  Filas {filas_str}: {datos["codigo"]} - Atributo: {atributo_str} - Cantidad: {datos["cantidad"]}'
                    )
                self.stdout.write('')
            
            if len(items_ok) == ingreso.items.count() and len(items_diferentes) == 0:
                self.stdout.write(self.style.SUCCESS('[OK] Todas las cantidades coinciden correctamente'))
            else:
                self.stdout.write(self.style.WARNING('[ADVERTENCIA] Hay diferencias en las cantidades'))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
            import traceback
            self.stdout.write(self.style.ERROR(traceback.format_exc()))


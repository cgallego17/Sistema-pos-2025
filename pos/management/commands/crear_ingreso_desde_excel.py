# -*- coding: utf-8 -*-
"""
Comando para crear un ingreso de mercancía desde un archivo Excel
"""
from django.core.management.base import BaseCommand
from django.db import transaction
from django.db.models import Q
from django.contrib.auth.models import User
from django.utils import timezone
from pos.models import (
    Producto, IngresoMercancia, ItemIngresoMercancia, MovimientoStock
)
import openpyxl
import os


class Command(BaseCommand):
    help = 'Crea un ingreso de mercancía desde un archivo Excel y lo completa automáticamente'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta del archivo Excel a importar'
        )
        parser.add_argument(
            '--proveedor',
            type=str,
            default='Proveedor Excel',
            help='Nombre del proveedor (default: "Proveedor Excel")'
        )
        parser.add_argument(
            '--numero-factura',
            type=str,
            default=None,
            help='Número de factura (opcional)'
        )
        parser.add_argument(
            '--usuario',
            type=str,
            default=None,
            help='Username del usuario que crea el ingreso (default: primer superusuario)'
        )
        parser.add_argument(
            '--confirmar',
            action='store_true',
            help='Confirmar sin preguntar',
        )

    def handle(self, *args, **options):
        archivo = options['archivo']
        
        if not os.path.exists(archivo):
            self.stdout.write(self.style.ERROR(f'El archivo {archivo} no existe'))
            return
        
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('CREACIÓN DE INGRESO DE MERCADERÍA DESDE EXCEL'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'Archivo: {archivo}')
        self.stdout.write('')
        
        try:
            # Obtener usuario
            if options['usuario']:
                try:
                    usuario = User.objects.get(username=options['usuario'])
                except User.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f'Usuario {options["usuario"]} no existe'))
                    return
            else:
                # Usar el primer superusuario o el primer usuario activo
                usuario = User.objects.filter(is_superuser=True).first()
                if not usuario:
                    usuario = User.objects.filter(is_active=True).first()
                if not usuario:
                    self.stdout.write(self.style.ERROR('No se encontró ningún usuario para crear el ingreso'))
                    return
            
            self.stdout.write(f'Usuario: {usuario.username} ({usuario.get_full_name() or usuario.email})')
            self.stdout.write('')
            
            # Cargar el archivo Excel con data_only=True para obtener valores calculados
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Leer la primera fila como encabezados
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.stdout.write(f'Columnas encontradas: {headers}')
            self.stdout.write('')
            
            # Mapear columnas (buscar por nombre, case-insensitive)
            col_indices = {}
            for idx, header in enumerate(headers):
                if header:
                    header_lower = str(header).lower().strip()
                    if 'codigo' in header_lower and 'barra' not in header_lower:
                        col_indices['codigo'] = idx
                    elif 'nombre' in header_lower and 'atributo' not in header_lower:
                        col_indices['nombre'] = idx
                    elif 'existencia' in header_lower or 'stock' in header_lower or 'cantidad' in header_lower:
                        col_indices['cantidad'] = idx
                    elif 'precio' in header_lower:
                        if 'compra' in header_lower:
                            col_indices['precio_compra'] = idx
                        elif 'precio_compra' not in col_indices:
                            # Si no hay precio_compra específico, usar precio como precio_compra
                            col_indices['precio_compra'] = idx
                        # También mapear precio para crear productos
                        if 'precio' not in col_indices or col_indices.get('precio') != col_indices.get('precio_compra'):
                            col_indices['precio'] = idx
                    elif 'atributo' in header_lower or 'nombreatributo' in header_lower:
                        col_indices['atributo'] = idx
                    elif 'barra' in header_lower or 'codigo_barra' in header_lower:
                        col_indices['codigo_barras'] = idx
            
            self.stdout.write(f'Columnas mapeadas: {col_indices}')
            self.stdout.write('')
            
            if 'codigo' not in col_indices:
                self.stdout.write(self.style.ERROR('ERROR: No se encontró la columna "codigo"'))
                return
            
            if 'cantidad' not in col_indices:
                self.stdout.write(self.style.ERROR('ERROR: No se encontró la columna de cantidad/stock/existencia'))
                return
            
            # Leer todas las filas de datos
            # Usar un diccionario para consolidar productos duplicados (mismo código y atributo)
            items_dict = {}  # clave: (codigo, atributo) -> {producto, cantidad_total, precio_compra, filas}
            productos_no_encontrados = []
            filas_con_error = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    # Obtener valores de las celdas
                    codigo = self._get_cell_value(row, col_indices.get('codigo'))
                    cantidad = self._get_cell_value(row, col_indices.get('cantidad'))
                    precio_compra = self._get_cell_value(row, col_indices.get('precio_compra'))
                    atributo = self._get_cell_value(row, col_indices.get('atributo'))
                    
                    # Validar campos requeridos
                    if not codigo:
                        continue  # Saltar filas vacías
                    
                    # Limpiar y convertir valores
                    codigo = str(codigo).strip()
                    
                    # Atributo - mantener tal cual está en el Excel
                    if atributo:
                        atributo = str(atributo).strip()
                        if atributo.upper() == 'SIN ATRIBUTO' or atributo == '':
                            atributo = None
                    else:
                        atributo = None
                    
                    # Cantidad
                    if cantidad is None:
                        cantidad = 0
                    else:
                        try:
                            cantidad = int(float(cantidad))
                            if cantidad <= 0:
                                continue  # Saltar si la cantidad es 0 o negativa
                        except (ValueError, TypeError):
                            continue  # Saltar si no es un número válido
                    
                    # Precio de compra
                    if precio_compra is None:
                        precio_compra = 0
                    else:
                        try:
                            precio_compra = int(float(precio_compra))
                            if precio_compra < 0:
                                precio_compra = 0
                        except (ValueError, TypeError):
                            precio_compra = 0
                    
                    # Buscar producto por código Y atributo
                    producto = None
                    if atributo:
                        producto = Producto.objects.filter(codigo=codigo, atributo=atributo).first()
                    else:
                        # Si no hay atributo, buscar por código sin atributo (atributo=None o atributo='')
                        producto = Producto.objects.filter(codigo=codigo).filter(
                            Q(atributo__isnull=True) | Q(atributo='')
                        ).first()
                    
                    # Si no existe, crear el producto
                    if not producto:
                        # Obtener nombre del producto desde el Excel si está disponible
                        nombre = self._get_cell_value(row, col_indices.get('nombre'))
                        if nombre:
                            nombre = str(nombre).strip()
                        else:
                            nombre = f'Producto {codigo}'  # Nombre por defecto
                        
                        # Obtener precio de venta desde el Excel si está disponible
                        precio_venta = self._get_cell_value(row, col_indices.get('precio'))
                        if precio_venta:
                            try:
                                precio_venta = int(float(precio_venta))
                                if precio_venta < 0:
                                    precio_venta = 0
                            except (ValueError, TypeError):
                                precio_venta = 0
                        else:
                            precio_venta = 0
                        
                        # Obtener código de barras si está disponible
                        codigo_barras = self._get_cell_value(row, col_indices.get('codigo_barras'))
                        if codigo_barras:
                            try:
                                codigo_barras = str(int(float(codigo_barras))).strip()
                                if codigo_barras == '0' or codigo_barras == '':
                                    codigo_barras = None
                                # Verificar si el código de barras ya existe
                                if codigo_barras and Producto.objects.filter(codigo_barras=codigo_barras).exists():
                                    codigo_barras = None  # No asignar si ya existe
                            except (ValueError, TypeError):
                                codigo_barras = None
                        else:
                            codigo_barras = None
                        
                        # Crear el producto
                        producto = Producto.objects.create(
                            codigo=codigo,
                            nombre=nombre,
                            atributo=atributo,
                            precio=precio_venta,
                            stock=0,  # Stock inicial en 0
                            codigo_barras=codigo_barras,
                            activo=True
                        )
                        self.stdout.write(f'  [CREADO] Producto: {codigo} - {nombre}')
                    
                    # Consolidar productos duplicados (mismo código y atributo) - usar solo la primera ocurrencia
                    clave = (codigo, atributo)
                    if clave in items_dict:
                        # Si ya existe, NO sumar - solo registrar que hay duplicado
                        items_dict[clave]['filas'].append(row_num)
                        items_dict[clave]['duplicado'] = True
                        self.stdout.write(
                            self.style.WARNING(
                                f'  [DUPLICADO OMITIDO] {codigo} - Atributo: {atributo or "SIN ATRIBUTO"} - '
                                f'Fila {row_num} (cantidad: {cantidad}) omitida. '
                                f'Ya existe en fila {items_dict[clave]["filas"][0]} (cantidad: {items_dict[clave]["cantidad"]})'
                            )
                        )
                    else:
                        items_dict[clave] = {
                            'producto': producto,
                            'cantidad': cantidad,
                            'precio_compra': precio_compra,
                            'filas': [row_num],
                            'atributo_excel': atributo,
                            'duplicado': False
                        }
                    
                except Exception as e:
                    filas_con_error.append({
                        'fila': row_num,
                        'error': str(e)
                    })
            
            # Identificar productos duplicados (solo para mostrar información)
            productos_duplicados_info = []
            for clave, datos in items_dict.items():
                if datos.get('duplicado', False) or len(datos['filas']) > 1:
                    productos_duplicados_info.append({
                        'codigo': datos['producto'].codigo,
                        'atributo': datos['atributo_excel'],
                        'filas': datos['filas'],
                        'cantidad_usada': datos['cantidad']  # Solo la primera cantidad
                    })
            
            # Convertir diccionario a lista
            items_a_procesar = list(items_dict.values())
            
            self.stdout.write(f'Items únicos encontrados en el archivo: {len(items_a_procesar)}')
            if productos_duplicados_info:
                self.stdout.write(self.style.WARNING(f'Productos duplicados consolidados: {len(productos_duplicados_info)}'))
            if productos_no_encontrados:
                self.stdout.write(self.style.WARNING(f'Productos no encontrados: {len(productos_no_encontrados)}'))
            if filas_con_error:
                self.stdout.write(self.style.WARNING(f'Filas con errores: {len(filas_con_error)}'))
            self.stdout.write('')
            
            if not items_a_procesar:
                self.stdout.write(self.style.WARNING('No se encontraron items válidos para procesar'))
                if productos_no_encontrados:
                    self.stdout.write('')
                    self.stdout.write('Productos no encontrados:')
                    for prod in productos_no_encontrados[:10]:
                        atributo_str = f" - Atributo: {prod.get('atributo', 'N/A')}" if prod.get('atributo') else ""
                        self.stdout.write(f'  Fila {prod["fila"]}: Código {prod["codigo"]}{atributo_str} (Cantidad: {prod["cantidad"]})')
                    if len(productos_no_encontrados) > 10:
                        self.stdout.write(f'  ... y {len(productos_no_encontrados) - 10} más')
                return
            
            # Mostrar muestra de items
            self.stdout.write('Muestra de items a procesar:')
            for i, item in enumerate(items_a_procesar[:10], 1):
                atributo_str = f" - Atributo: {item['producto'].atributo}" if item['producto'].atributo else ""
                filas_str = ', '.join(map(str, item['filas']))
                self.stdout.write(
                    f'  {i}. {item["producto"].codigo} - {item["producto"].nombre}{atributo_str} '
                    f'(Cantidad: {item["cantidad"]}, Precio Compra: ${item["precio_compra"]:,}, Filas: {filas_str})'
                )
            if len(items_a_procesar) > 10:
                self.stdout.write(f'  ... y {len(items_a_procesar) - 10} más')
            self.stdout.write('')
            
            # Mostrar productos duplicados (solo información)
            if productos_duplicados_info:
                self.stdout.write('')
                self.stdout.write(self.style.WARNING('Productos duplicados en Excel (solo se usó la primera ocurrencia):'))
                for dup in productos_duplicados_info:
                    atributo_str = dup['atributo'] if dup['atributo'] else 'SIN ATRIBUTO'
                    filas_str = ', '.join(map(str, dup['filas']))
                    self.stdout.write(
                        f'  {dup["codigo"]} - Atributo: {atributo_str} - Filas: {filas_str} - Cantidad usada (primera fila): {dup["cantidad_usada"]}'
                    )
                self.stdout.write('')
            
            # Mostrar productos no encontrados
            if productos_no_encontrados:
                self.stdout.write('')
                self.stdout.write(self.style.WARNING('Productos no encontrados (serán omitidos):'))
                for prod in productos_no_encontrados[:10]:
                    atributo_str = f" - Atributo: {prod.get('atributo', 'N/A')}" if prod.get('atributo') else ""
                    self.stdout.write(f'  Fila {prod["fila"]}: Código {prod["codigo"]}{atributo_str} (Cantidad: {prod["cantidad"]})')
                if len(productos_no_encontrados) > 10:
                    self.stdout.write(f'  ... y {len(productos_no_encontrados) - 10} más')
                self.stdout.write('')
            
            # Confirmación
            if not options['confirmar']:
                respuesta = input('¿Deseas continuar con la creación del ingreso? (si/no): ')
                if respuesta.lower() != 'si':
                    self.stdout.write(self.style.WARNING('Proceso cancelado'))
                    return
            
            # Crear ingreso y procesar items
            self.stdout.write(self.style.WARNING('Creando ingreso de mercancía...'))
            
            proveedor = options['proveedor']
            numero_factura = options['numero_factura']
            
            with transaction.atomic():
                # Crear ingreso
                ingreso = IngresoMercancia.objects.create(
                    proveedor=proveedor,
                    numero_factura=numero_factura,
                    observaciones=f'Ingreso creado desde Excel: {os.path.basename(archivo)}',
                    usuario=usuario,
                    completado=False
                )
                
                total = 0
                items_creados = 0
                
                # Crear items
                for item_data in items_a_procesar:
                    producto = item_data['producto']
                    cantidad = item_data['cantidad']
                    precio_compra = item_data['precio_compra']
                    subtotal = cantidad * precio_compra
                    
                    ItemIngresoMercancia.objects.create(
                        ingreso=ingreso,
                        producto=producto,
                        cantidad=cantidad,
                        precio_compra=precio_compra,
                        subtotal=subtotal
                    )
                    
                    total += subtotal
                    items_creados += 1
                
                # Actualizar total del ingreso
                ingreso.total = total
                ingreso.save()
                
                # NO completar automáticamente - el usuario debe verificar producto por producto
                # El ingreso queda pendiente para que el usuario verifique cada item
                
            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(self.style.SUCCESS('INGRESO CREADO EXITOSAMENTE'))
            self.stdout.write(self.style.SUCCESS('=' * 70))
            self.stdout.write(f'Ingreso ID: #{ingreso.id}')
            self.stdout.write(f'Proveedor: {ingreso.proveedor}')
            if ingreso.numero_factura:
                self.stdout.write(f'Número de Factura: {ingreso.numero_factura}')
            self.stdout.write(f'Total: ${ingreso.total:,}')
            self.stdout.write(f'Items creados: {items_creados}')
            self.stdout.write(f'Fecha: {ingreso.fecha.strftime("%d/%m/%Y %H:%M:%S")}')
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('IMPORTANTE: El ingreso está pendiente de verificación.'))
            self.stdout.write(self.style.WARNING('Debes ir al detalle del ingreso y verificar cada producto antes de completarlo.'))
            self.stdout.write('')
            
            if productos_no_encontrados:
                self.stdout.write(self.style.WARNING(f'Nota: {len(productos_no_encontrados)} productos no fueron encontrados y fueron omitidos'))
            
        except Exception as e:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('=' * 70))
            self.stdout.write(self.style.ERROR('ERROR AL CREAR INGRESO'))
            self.stdout.write(self.style.ERROR('=' * 70))
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
            import traceback
            self.stdout.write(self.style.ERROR(traceback.format_exc()))
    
    def _get_cell_value(self, row, col_index):
        """Obtener valor de una celda de forma segura"""
        if col_index is None or col_index >= len(row):
            return None
        cell = row[col_index]
        if cell is None:
            return None
        return cell.value


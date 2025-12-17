from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import datetime
from functools import wraps
import json
import logging

# Configurar logger para trazabilidad
logger = logging.getLogger(__name__)

from .models import (
    Producto, Venta, ItemVenta, Caja, CajaUsuario,
    MovimientoStock, PerfilUsuario, GastoCaja,
    IngresoMercancia, ItemIngresoMercancia,
    SalidaMercancia, ItemSalidaMercancia,
    CampanaMarketing, ClientePotencial
)


# ============================================
# FUNCIONES AUXILIARES DE TRAZABILIDAD
# ============================================

def obtener_caja_mostrar(usuario=None, fecha=None):
    """
    Función auxiliar para obtener la caja única global que se debe mostrar/usar.
    Esta función asegura consistencia en toda la aplicación.
    
    Lógica: Solo existe UNA caja en el sistema que se reutiliza.
    1. Obtiene o crea la Caja Principal (número 1)
    2. Busca la única caja del sistema (solo hay una)
    3. Retorna la caja encontrada o None
    
    Args:
        usuario: Parámetro ignorado (mantenido por compatibilidad)
        fecha: Parámetro ignorado (mantenido por compatibilidad)
    
    Returns:
        CajaUsuario o None
    """
    # Obtener o crear la Caja Principal
    caja_principal = Caja.objects.filter(numero=1).first()
    if not caja_principal:
        caja_principal = Caja.objects.create(
            numero=1,
            nombre='Caja Principal',
            activa=True
        )
    
    # Buscar la única caja del sistema (solo hay una)
    caja_unica = CajaUsuario.objects.filter(
        caja=caja_principal
    ).order_by('-fecha_apertura').first()
    
    return caja_unica


# ============================================
# SISTEMA DE ROLES Y PERMISOS
# ============================================

def tiene_rol(user, *nombres_roles):
    """Verifica si el usuario pertenece a alguno de los roles especificados"""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    grupos_usuario = user.groups.values_list('name', flat=True)
    return any(rol in grupos_usuario for rol in nombres_roles)


def requiere_rol(*nombres_roles):
    """Decorador para requerir que el usuario tenga uno de los roles especificados"""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, 'Debes iniciar sesión para acceder')
                return redirect('pos:login')
            
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            
            if not tiene_rol(request.user, *nombres_roles):
                messages.error(request, 'No tienes permisos para acceder a esta sección')
                return redirect('pos:home')
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def puede_anular_ventas(user):
    """Verifica si el usuario puede anular ventas (Administradores)"""
    return tiene_rol(user, 'Administradores') or user.is_superuser or user.is_staff


def puede_gestionar_productos(user):
    """Verifica si el usuario puede gestionar productos (Administradores, Inventario)"""
    return tiene_rol(user, 'Administradores', 'Inventario') or user.is_superuser


def puede_gestionar_caja(user):
    """Verifica si el usuario puede gestionar cajas (Administradores, Cajeros)"""
    return tiene_rol(user, 'Administradores', 'Cajeros') or user.is_superuser


def puede_ver_reportes(user):
    """Verifica si el usuario puede ver reportes (Administradores, Staff)"""
    return tiene_rol(user, 'Administradores') or user.is_superuser or user.is_staff


def puede_realizar_ventas(user):
    """Verifica si el usuario puede realizar ventas (todos los roles excepto algunos)"""
    if user.is_superuser:
        return True
    grupos_usuario = user.groups.values_list('name', flat=True)
    return len(grupos_usuario) > 0 or user.is_staff


def login_view(request):
    """Vista de login con usuario y PIN"""
    if request.user.is_authenticated:
        return redirect('pos:home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        pin = request.POST.get('pin')
        
        if not username or not pin:
            messages.error(request, 'Debes ingresar usuario y PIN')
            return render(request, 'pos/login.html')
        
        # Buscar usuario por username
        try:
            from django.contrib.auth.models import User
            user = User.objects.get(username=username, is_active=True)
            
            # Verificar PIN
            try:
                perfil = PerfilUsuario.objects.get(usuario=user, pin=pin, pin_establecido=True)
                login(request, user)
                return redirect('pos:home')
            except PerfilUsuario.DoesNotExist:
                messages.error(request, 'Usuario o PIN incorrectos')
        except User.DoesNotExist:
            messages.error(request, 'Usuario o PIN incorrectos')
    
    return render(request, 'pos/login.html')


def login_pin_view(request):
    """Vista de login con PIN"""
    if request.method == 'POST':
        pin = request.POST.get('pin')
        
        try:
            perfil = PerfilUsuario.objects.get(pin=pin, pin_establecido=True)
            user = perfil.usuario
            login(request, user)
            return redirect('pos:home')
        except PerfilUsuario.DoesNotExist:
            messages.error(request, 'PIN incorrecto')
            return redirect('pos:login')
    
    return redirect('pos:login')


@login_required
def logout_view(request):
    """Vista de logout - Cierra TODAS las registradoras del usuario antes de cerrar sesión"""
    from .models import RegistradoraActiva
    
    # IMPORTANTE: Cerrar TODAS las registradoras activas del usuario ANTES de hacer logout
    # Esto asegura que no queden registradoras abiertas cuando el usuario cierra sesión
    try:
        # Obtener todas las registradoras activas del usuario
        registradoras_activas = RegistradoraActiva.objects.filter(usuario=request.user)
        cantidad_cerradas = registradoras_activas.count()
        
        if cantidad_cerradas > 0:
            # Eliminar todas las registradoras activas del usuario de la base de datos
            registradoras_activas.delete()
        
        # Limpiar la sesión de registradora seleccionada
        if 'registradora_seleccionada' in request.session:
            del request.session['registradora_seleccionada']
            request.session.modified = True
    except Exception as e:
        # Si hay algún error, registrarlo pero continuar con el logout
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error al cerrar registradoras en logout: {str(e)}')
    
    # Hacer logout
    logout(request)
    return redirect('pos:login')


@login_required
def seleccionar_registradora_view(request):
    """Vista para seleccionar la registradora"""
    from .models import RegistradoraActiva
    
    if request.method == 'POST':
        # Verificar si hay caja abierta (caja única global)
        caja_principal = Caja.objects.filter(numero=1).first()
        if not caja_principal:
            messages.error(
                request, 
                'No existe la Caja Principal. Por favor, contacta al administrador.'
            )
            return redirect('pos:home')
        
        # Buscar la única caja del sistema (sin filtrar por usuario)
        caja_abierta = CajaUsuario.objects.filter(
            caja=caja_principal,
            fecha_cierre__isnull=True
        ).first()
        
        if not caja_abierta:
            messages.error(
                request, 
                'Debes abrir una caja antes de seleccionar una registradora. Por favor, abre la caja desde el Dashboard.'
            )
            return redirect('pos:home')
        
        registradora_id = request.POST.get('registradora_id')
        if registradora_id:
            registradora_id_int = int(registradora_id)
            registradoras = {
                1: 'Registradora 1',
                2: 'Registradora 2',
                3: 'Registradora 3',
            }
            nombre_registradora = registradoras.get(registradora_id_int, 'Registradora Desconocida')
            
            # Verificar si la registradora ya está en uso por otro usuario
            registradora_activa = RegistradoraActiva.objects.filter(
                registradora_id=registradora_id_int
            ).first()
            
            if registradora_activa and registradora_activa.usuario != request.user:
                # La registradora está en uso por otro usuario
                messages.error(
                    request, 
                    f'La {nombre_registradora} está actualmente en uso por el usuario "{registradora_activa.usuario.get_full_name() or registradora_activa.usuario.username}". '
                    f'Por favor, espera a que la cierre o contacta con ese usuario.'
                )
                return redirect('pos:home')
            
            # Si el usuario ya tiene esta registradora, no hacer nada
            if registradora_activa and registradora_activa.usuario == request.user:
                messages.info(request, f'Ya tienes la {nombre_registradora} activa')
                return redirect('pos:vender')
            
            # Cerrar cualquier otra registradora que el usuario tenga activa
            RegistradoraActiva.objects.filter(usuario=request.user).delete()
            
            # Crear o actualizar la registradora activa
            RegistradoraActiva.objects.update_or_create(
                registradora_id=registradora_id_int,
                defaults={
                    'usuario': request.user,
                    'fecha_apertura': timezone.now()
                }
            )
            
            # Guardar en sesión también
            request.session['registradora_seleccionada'] = {
                'id': registradora_id_int,
                'nombre': nombre_registradora
            }
            request.session.modified = True
            messages.success(request, f'Registradora {nombre_registradora} seleccionada')
            # Redirigir al POS después de seleccionar
            return redirect('pos:vender')
        else:
            messages.error(request, 'Debes seleccionar una registradora')
    
    return redirect('pos:home')


@login_required
def cerrar_registradora_view(request):
    """Vista para cerrar/desactivar la registradora"""
    from .models import RegistradoraActiva
    
    if request.method == 'POST':
        # Obtener registradora_id del POST (puede ser de la sesión o de un parámetro)
        registradora_id_post = request.POST.get('registradora_id')
        registradora_actual = request.session.get('registradora_seleccionada', None)
        
        # Determinar qué registradora cerrar
        if registradora_id_post:
            # Si se especifica un ID en el POST, cerrar esa registradora
            registradora_id = int(registradora_id_post)
            registradoras = {
                1: 'Registradora 1',
                2: 'Registradora 2',
                3: 'Registradora 3',
            }
            nombre_registradora = registradoras.get(registradora_id, f'Registradora {registradora_id}')
            
            # Buscar la registradora activa
            registradora_activa = RegistradoraActiva.objects.filter(
                registradora_id=registradora_id
            ).first()
            
            if registradora_activa:
                # Verificar si es del usuario actual o si el usuario es admin/staff
                if registradora_activa.usuario == request.user or request.user.is_staff or request.user.is_superuser:
                    # Eliminar de la base de datos
                    registradora_activa.delete()
                    
                    # Si es la registradora del usuario actual, eliminar de la sesión
                    if registradora_actual and registradora_actual.get('id') == registradora_id:
                        del request.session['registradora_seleccionada']
                        request.session.modified = True
                    
                    messages.success(request, f'{nombre_registradora} cerrada exitosamente')
                else:
                    messages.error(request, 'No tienes permisos para cerrar esta registradora')
            else:
                messages.warning(request, f'{nombre_registradora} no está activa')
        elif registradora_actual:
            # Si no hay ID en POST pero hay registradora en sesión, cerrar esa
            registradora_id = registradora_actual.get('id')
            nombre_registradora = registradora_actual.get('nombre', 'Registradora')
            
            # Eliminar de la base de datos
            RegistradoraActiva.objects.filter(
                registradora_id=registradora_id,
                usuario=request.user
            ).delete()
            
            # Eliminar de la sesión
            del request.session['registradora_seleccionada']
            request.session.modified = True
            messages.success(request, f'{nombre_registradora} cerrada exitosamente')
        else:
            messages.warning(request, 'No hay registradora activa para cerrar')
    
    return redirect('pos:home')


@login_required
def home_view(request):
    """Vista principal del dashboard"""
    hoy = timezone.now().date()
    
    # Ventas de hoy
    ventas_hoy = Venta.objects.filter(
        fecha__date=hoy,
        completada=True,
        anulada=False
    ).aggregate(
        total=Sum('total'),
        cantidad=Count('id')
    )
    
    # Productos
    productos_count = Producto.objects.filter(activo=True).count()
    productos_bajo_stock = Producto.objects.filter(activo=True, stock__lt=10).count()
    
    # Últimas ventas
    ultimas_ventas = Venta.objects.filter(
        completada=True
    ).order_by('-fecha')[:10]
    
    # Productos con stock bajo
    productos_stock_bajo = Producto.objects.filter(
        activo=True,
        stock__lt=10
    ).order_by('stock')[:10]
    
    # Caja abierta (caja única global)
    caja_principal = Caja.objects.filter(numero=1).first()
    caja_abierta = None
    if caja_principal:
        caja_abierta = CajaUsuario.objects.filter(
            caja=caja_principal,
            fecha_cierre__isnull=True
        ).first()
    
    # Obtener registradora seleccionada de la sesión
    registradora_seleccionada = request.session.get('registradora_seleccionada', None)
    
    # Verificar registradoras activas en la base de datos
    from .models import RegistradoraActiva
    registradoras_activas = RegistradoraActiva.objects.select_related('usuario').all()
    registradoras_activas_dict = {
        reg.registradora_id: reg.usuario 
        for reg in registradoras_activas
    }
    
    # Lista de registradoras disponibles con información de quién las tiene
    registradoras = []
    for reg_id in [1, 2, 3]:
        usuario_activo = registradoras_activas_dict.get(reg_id)
        registradoras.append({
            'id': reg_id,
            'nombre': f'Registradora {reg_id}',
            'usuario_activo': usuario_activo,
            'disponible': usuario_activo is None or usuario_activo == request.user
        })
    
    # Calcular ventas totales del mes actual para la meta
    from datetime import datetime
    inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    ventas_mes = Venta.objects.filter(
        fecha__gte=inicio_mes,
        completada=True,
        anulada=False
    ).aggregate(total=Sum('total'))
    
    ventas_totales_mes = float(ventas_mes['total'] or 0)
    meta_ventas = 100000000  # 100 millones
    porcentaje_meta = (ventas_totales_mes / meta_ventas * 100) if meta_ventas > 0 else 0
    porcentaje_meta = float(min(porcentaje_meta, 100))  # Limitar al 100% y convertir a float
    faltan_para_meta = max(0, meta_ventas - ventas_totales_mes)  # No mostrar negativo
    
    context = {
        'ventas_hoy': ventas_hoy['total'] or 0,
        'total_ventas_hoy': ventas_hoy['cantidad'] or 0,
        'productos_count': productos_count,
        'productos_bajo_stock': productos_bajo_stock,
        'ultimas_ventas': ultimas_ventas,
        'productos_stock_bajo': productos_stock_bajo,
        'caja_abierta': caja_abierta,
        'registradora_seleccionada': registradora_seleccionada,
        'registradoras': registradoras,
        'ventas_totales_mes': ventas_totales_mes,
        'meta_ventas': meta_ventas,
        'porcentaje_meta': porcentaje_meta,
        'faltan_para_meta': faltan_para_meta,
    }
    
    return render(request, 'pos/home.html', context)


@login_required
def vender_view(request):
    """Vista del punto de venta"""
    # Verificar si hay registradora seleccionada
    registradora_seleccionada = request.session.get('registradora_seleccionada', None)
    if not registradora_seleccionada:
        messages.warning(request, 'Por favor, selecciona una registradora en el Dashboard antes de vender')
        return redirect('pos:home')
    
    # Verificar si hay caja abierta (caja única global)
    caja_principal = Caja.objects.filter(numero=1).first()
    caja_abierta = None
    if caja_principal:
        caja_abierta = CajaUsuario.objects.filter(
            caja=caja_principal,
            fecha_cierre__isnull=True
        ).first()
    
    if not caja_abierta:
        messages.warning(request, 'Debes abrir una caja antes de realizar ventas. Por favor, abre la caja desde el Dashboard.')
        return redirect('pos:home')
    
    # Limpiar carrito de productos inválidos al entrar al POS
    carrito = get_carrito(request)
    if carrito:
        carrito_limpio = {}
        productos_activos_ids = set(Producto.objects.filter(activo=True).values_list('id', flat=True))
        
        for key, item in carrito.items():
            producto_id = item.get('producto_id')
            # Solo mantener productos que existan, estén activos y tengan stock suficiente
            if producto_id and producto_id in productos_activos_ids:
                try:
                    producto = Producto.objects.get(id=producto_id, activo=True)
                    cantidad_carrito = item.get('cantidad', 0)
                    if cantidad_carrito > 0 and producto.stock >= cantidad_carrito:
                        # Actualizar atributo si falta o cambió
                        item_actualizado = item.copy()
                        item_actualizado['atributo'] = producto.atributo or ''
                        carrito_limpio[key] = item_actualizado
                except Producto.DoesNotExist:
                    pass
        
        # Actualizar el carrito en la sesión si se limpió
        if len(carrito_limpio) != len(carrito):
            request.session['carrito'] = carrito_limpio
            request.session.modified = True
    
    # Si es AJAX para cargar carrito
    if request.GET.get('cargar_carrito'):
        tab_id = request.GET.get('tab_id')
        carrito = get_carrito(request, tab_id)
        # Validar y limpiar productos inválidos del carrito
        carrito_limpio = {}
        productos_activos_ids = set(Producto.objects.filter(activo=True).values_list('id', flat=True))
        
        for key, item in carrito.items():
            producto_id = item.get('producto_id')
            # Solo mantener productos que existan, estén activos y tengan stock
            if producto_id and producto_id in productos_activos_ids:
                try:
                    producto = Producto.objects.get(id=producto_id, activo=True)
                    # Verificar que el stock sea suficiente
                    cantidad_carrito = item.get('cantidad', 0)
                    if cantidad_carrito > 0 and producto.stock >= cantidad_carrito:
                        # Actualizar atributo si falta o cambió
                        item_actualizado = item.copy()
                        item_actualizado['atributo'] = producto.atributo or ''
                        carrito_limpio[key] = item_actualizado
                except Producto.DoesNotExist:
                    # Producto no existe o está inactivo, no incluirlo
                    pass
        
        # Actualizar el carrito en la sesión si se limpió
        if len(carrito_limpio) != len(carrito):
            if tab_id:
                if 'carritos' not in request.session:
                    request.session['carritos'] = {}
                request.session['carritos'][tab_id] = carrito_limpio
            else:
                request.session['carrito'] = carrito_limpio
            request.session.modified = True
        
        return JsonResponse({'carrito': carrito_limpio})
    
    productos = Producto.objects.filter(activo=True, stock__gt=0).order_by('nombre')
    
    # Obtener productos más vendidos (últimos 30 días)
    from datetime import timedelta
    hace_30_dias = timezone.now() - timedelta(days=30)
    mas_vendidos = ItemVenta.objects.filter(
        venta__fecha__gte=hace_30_dias,
        venta__completada=True,
        venta__anulada=False
    ).values('producto_id').annotate(
        total=Sum('cantidad')
    ).order_by('-total')[:10]
    
    mas_vendidos_ids = [item['producto_id'] for item in mas_vendidos]
    
    # Obtener vendedores activos
    from django.contrib.auth.models import User
    vendedores = User.objects.filter(is_active=True).order_by('username')
    
    context = {
        'productos': productos,
        'mas_vendidos_ids': mas_vendidos_ids,
        'vendedores': vendedores,
        'registradora_seleccionada': registradora_seleccionada,
    }
    
    return render(request, 'pos/vender.html', context)


@login_required
def procesar_venta(request):
    """Procesar una venta (AJAX)"""
    if request.method == 'POST':
        try:
            # Verificar si hay caja abierta (caja única global)
            caja_principal = Caja.objects.filter(numero=1).first()
            caja_abierta = None
            if caja_principal:
                caja_abierta = CajaUsuario.objects.filter(
                    caja=caja_principal,
                    fecha_cierre__isnull=True
                ).first()
            
            if not caja_abierta:
                return JsonResponse({
                    'success': False, 
                    'error': 'Debes abrir una caja antes de realizar ventas. Por favor, abre la caja desde el Dashboard.'
                })
            
            data = json.loads(request.body)
            items = data.get('items', [])
            metodo_pago = data.get('metodo_pago', 'efectivo')
            monto_recibido = data.get('monto_recibido')
            email_cliente = data.get('email_cliente', '')
            vendedor_id = data.get('vendedor_id')
            
            if not items:
                return JsonResponse({'success': False, 'error': 'No hay items en la venta'})
            
            # Obtener vendedor
            vendedor = None
            if vendedor_id:
                from django.contrib.auth.models import User
                try:
                    vendedor = User.objects.get(id=vendedor_id)
                except User.DoesNotExist:
                    pass
            
            # Obtener registradora de la sesión
            registradora_seleccionada = request.session.get('registradora_seleccionada', None)
            registradora_id = None
            if registradora_seleccionada:
                registradora_id = registradora_seleccionada.get('id')
            
            # Crear la venta
            venta = Venta.objects.create(
                usuario=request.user,
                vendedor=vendedor,
                metodo_pago=metodo_pago,
                monto_recibido=monto_recibido if monto_recibido else None,
                email_cliente=email_cliente if email_cliente else None,
                registradora_id=registradora_id,
                completada=True
            )
            
            # Agregar items y calcular total
            total = 0
            for item_data in items:
                producto = Producto.objects.get(id=item_data['id'])
                cantidad = item_data['cantidad']
                
                # Verificar stock
                if producto.stock < cantidad:
                    venta.delete()
                    return JsonResponse({
                        'success': False,
                        'error': f'Stock insuficiente para {producto.nombre}'
                    })
                
                # Crear item de venta
                subtotal = producto.precio * cantidad
                ItemVenta.objects.create(
                    venta=venta,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=producto.precio,
                    subtotal=subtotal
                )
                
                # Actualizar stock
                producto.stock -= cantidad
                producto.save()
                
                # Registrar movimiento de stock
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='salida',
                    cantidad=cantidad,
                    stock_anterior=producto.stock + cantidad,
                    stock_nuevo=producto.stock,
                    motivo=f'Venta #{venta.id}',
                    usuario=request.user
                )
                
                total += subtotal
            
            # Actualizar total de la venta
            venta.total = total
            
            # Asignar siempre a la Caja Principal (todas las ventas van a la misma caja)
            caja_principal = Caja.objects.filter(numero=1).first()
            if caja_principal:
                venta.caja = caja_principal
            
            venta.save()
            
            return JsonResponse({
                'success': True,
                'venta_id': venta.id,
                'total': float(total)
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def productos_view(request):
    """Vista de lista de productos"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q
    
    # Todos pueden ver productos, pero solo algunos pueden editarlos
    productos_list = Producto.objects.all()
    
    # Búsqueda
    busqueda = request.GET.get('buscar', '').strip()
    if busqueda:
        productos_list = productos_list.filter(
            Q(codigo__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(codigo_barras__icontains=busqueda) |
            Q(atributo__icontains=busqueda)
        )
    
    # Filtro de estado
    filtro_estado = request.GET.get('estado', '')
    if filtro_estado == 'activo':
        productos_list = productos_list.filter(activo=True)
    elif filtro_estado == 'inactivo':
        productos_list = productos_list.filter(activo=False)
    elif filtro_estado == 'bajo-stock':
        productos_list = productos_list.filter(stock__lt=10)
    
    # Ordenar
    productos_list = productos_list.order_by('nombre')
    
    # Paginación: 20 productos por página
    paginator = Paginator(productos_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        productos = paginator.page(page)
    except PageNotAnInteger:
        # Si la página no es un entero, mostrar la primera página
        productos = paginator.page(1)
    except EmptyPage:
        # Si la página está fuera de rango, mostrar la última página
        productos = paginator.page(paginator.num_pages)
    
    # Verificar si puede editar productos
    puede_editar = puede_gestionar_productos(request.user)
    
    context = {
        'productos': productos,
        'puede_editar': puede_editar,
        'busqueda_actual': busqueda,
        'filtro_estado_actual': filtro_estado,
    }
    
    return render(request, 'pos/productos.html', context)


@login_required
def lista_ventas_view(request):
    """Vista de lista de ventas"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    ventas_list = Venta.objects.filter(completada=True).order_by('-fecha')
    
    # Paginación: 20 ventas por página
    paginator = Paginator(ventas_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        ventas = paginator.page(page)
    except PageNotAnInteger:
        ventas = paginator.page(1)
    except EmptyPage:
        ventas = paginator.page(paginator.num_pages)
    
    context = {
        'ventas': ventas,
    }
    
    return render(request, 'pos/lista_ventas.html', context)


@login_required
def detalle_venta_view(request, venta_id):
    """Vista de detalle de una venta"""
    venta = get_object_or_404(Venta, id=venta_id)
    
    context = {
        'venta': venta,
    }
    
    return render(request, 'pos/detalle_venta.html', context)


@login_required
def editar_venta_view(request, venta_id):
    """Editar una venta"""
    # Solo administradores y cajeros pueden editar ventas
    if not puede_anular_ventas(request.user):
        messages.error(request, 'No tienes permisos para editar ventas')
        return redirect('pos:detalle_venta', venta_id=venta_id)
    
    venta = get_object_or_404(Venta, id=venta_id)
    
    if venta.anulada:
        messages.error(request, 'No se puede editar una venta anulada')
        return redirect('pos:detalle_venta', venta_id=venta_id)
    
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    from django.contrib.auth.models import User
    usuarios = User.objects.filter(is_active=True).order_by('username')
    
    if request.method == 'POST':
        # Validar que la venta no esté anulada
        if venta.anulada:
            error_msg = 'No se puede editar una venta anulada'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('pos:detalle_venta', venta_id=venta_id)
        
        # Actualizar método de pago
        metodo_pago = request.POST.get('metodo_pago', venta.metodo_pago)
        if metodo_pago not in ['efectivo', 'tarjeta', 'transferencia']:
            error_msg = 'Método de pago inválido'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('pos:editar_venta', venta_id=venta_id)
        venta.metodo_pago = metodo_pago
        
        # Actualizar monto recibido con validaciones
        monto_recibido = None
        if 'monto_recibido' in request.POST:
            monto = request.POST.get('monto_recibido', '').strip()
            if monto:
                try:
                    import re
                    monto_limpio = re.sub(r'[^0-9]', '', monto)
                    if monto_limpio:
                        monto_recibido = int(float(monto_limpio))
                        if monto_recibido < 0:
                            error_msg = 'El monto recibido no puede ser negativo'
                            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                                return JsonResponse({'success': False, 'error': error_msg})
                            messages.error(request, error_msg)
                            return redirect('pos:editar_venta', venta_id=venta_id)
                except (ValueError, TypeError):
                    error_msg = 'El monto recibido no es un número válido'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': error_msg})
                    messages.error(request, error_msg)
                    return redirect('pos:editar_venta', venta_id=venta_id)
        venta.monto_recibido = monto_recibido
        
        # Actualizar vendedor
        if 'vendedor_id' in request.POST:
            vendedor_id = request.POST.get('vendedor_id', '').strip()
            if vendedor_id:
                try:
                    venta.vendedor = User.objects.get(id=int(vendedor_id), is_active=True)
                except (User.DoesNotExist, ValueError):
                    venta.vendedor = None
            else:
                venta.vendedor = None
        
        # Actualizar items
        if 'items' in request.POST:
            try:
                items_data = json.loads(request.POST.get('items'))
                # Obtener items actuales
                items_actuales = {item.id: item for item in venta.items.all()}
                items_procesados = set()
                total = 0
                
                # Validar que haya al menos un item
                if not items_data:
                    error_msg = 'Debe agregar al menos un item a la venta'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': error_msg})
                    messages.error(request, error_msg)
                    return redirect('pos:editar_venta', venta_id=venta_id)
                
                # Actualizar o crear items
                for item_data in items_data:
                    # Validar datos del item
                    try:
                        producto_id = int(item_data.get('producto_id', 0))
                        cantidad = int(item_data.get('cantidad', 0))
                        precio = int(float(item_data.get('precio', 0)))
                        item_id = item_data.get('item_id')
                    except (ValueError, TypeError, KeyError):
                        error_msg = 'Datos inválidos en uno de los items'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': error_msg})
                        messages.error(request, error_msg)
                        return redirect('pos:editar_venta', venta_id=venta_id)
                    
                    # Validar que los valores sean positivos
                    if producto_id <= 0:
                        error_msg = 'Debe seleccionar un producto válido'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': error_msg})
                        messages.error(request, error_msg)
                        return redirect('pos:editar_venta', venta_id=venta_id)
                    
                    if cantidad <= 0:
                        error_msg = 'La cantidad debe ser mayor a 0'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': error_msg})
                        messages.error(request, error_msg)
                        return redirect('pos:editar_venta', venta_id=venta_id)
                    
                    if precio < 0:
                        error_msg = 'El precio no puede ser negativo'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': error_msg})
                        messages.error(request, error_msg)
                        return redirect('pos:editar_venta', venta_id=venta_id)
                    
                    if item_id and int(item_id) in items_actuales:
                        # Actualizar item existente
                        item = items_actuales[int(item_id)]
                        # Devolver stock anterior
                        item.producto.stock += item.cantidad
                        item.producto.save()
                        # Actualizar item
                        item.producto = Producto.objects.get(id=producto_id)
                        item.cantidad = cantidad
                        item.precio_unitario = precio
                        item.subtotal = precio * cantidad
                        item.save()
                        # Descontar nuevo stock
                        item.producto.stock -= cantidad
                        item.producto.save()
                        items_procesados.add(item.id)
                    else:
                        # Crear nuevo item
                        try:
                            producto = Producto.objects.get(id=producto_id, activo=True)
                        except Producto.DoesNotExist:
                            error_msg = f'El producto con ID {producto_id} no existe o está inactivo'
                            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                                return JsonResponse({'success': False, 'error': error_msg})
                            messages.error(request, error_msg)
                            return redirect('pos:editar_venta', venta_id=venta_id)
                        
                        if producto.stock < cantidad:
                            error_msg = f'Stock insuficiente para {producto.nombre}. Disponible: {producto.stock}, Solicitado: {cantidad}'
                            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                                return JsonResponse({'success': False, 'error': error_msg})
                            messages.error(request, error_msg)
                            return redirect('pos:editar_venta', venta_id=venta_id)
                        
                        ItemVenta.objects.create(
                            venta=venta,
                            producto=producto,
                            cantidad=cantidad,
                            precio_unitario=precio,
                            subtotal=precio * cantidad
                        )
                        # Descontar stock
                        producto.stock -= cantidad
                        producto.save()
                    
                    total += precio * cantidad
                
                # Eliminar items que no están en la lista
                for item_id, item in items_actuales.items():
                    if item_id not in items_procesados:
                        # Devolver stock
                        item.producto.stock += item.cantidad
                        item.producto.save()
                        item.delete()
                
                venta.total = total
                
                # Validar monto recibido si es efectivo
                if venta.metodo_pago == 'efectivo' and venta.monto_recibido is not None:
                    if venta.monto_recibido < total:
                        error_msg = f'El monto recibido (${venta.monto_recibido:,}) es menor al total de la venta (${total:,}). Faltan ${(total - venta.monto_recibido):,}'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'error': error_msg})
                        messages.error(request, error_msg)
                        return redirect('pos:editar_venta', venta_id=venta_id)
            except json.JSONDecodeError:
                error_msg = 'Error al procesar los items. Formato JSON inválido.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
                return redirect('pos:editar_venta', venta_id=venta_id)
            except Exception as e:
                error_msg = f'Error al actualizar items: {str(e)}'
                messages.error(request, error_msg)
                # Si es una petición AJAX, devolver JSON con error
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': error_msg})
                return redirect('pos:editar_venta', venta_id=venta_id)
        
        # Validar monto recibido si es efectivo (validación final)
        if venta.metodo_pago == 'efectivo' and venta.monto_recibido is not None:
            if venta.monto_recibido < venta.total:
                error_msg = f'El monto recibido (${venta.monto_recibido:,}) es menor al total de la venta (${venta.total:,}). Faltan ${(venta.total - venta.monto_recibido):,}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': error_msg})
                messages.error(request, error_msg)
                return redirect('pos:editar_venta', venta_id=venta_id)
        
        venta.save()
        messages.success(request, f'Venta #{venta.id} actualizada exitosamente')
        
        # Si es una petición AJAX, devolver JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Venta #{venta.id} actualizada exitosamente'})
        
        return redirect('pos:detalle_venta', venta_id=venta_id)
    
    context = {
        'venta': venta,
        'productos': productos,
        'usuarios': usuarios,
    }
    return render(request, 'pos/editar_venta.html', context)


@login_required
def imprimir_ticket_view(request, venta_id):
    """Vista para imprimir ticket térmico 80mm"""
    venta = get_object_or_404(Venta, id=venta_id)
    
    # IMPORTANTE: monto_recibido siempre es igual al total (no se guarda el monto pagado mayor)
    # Por lo tanto, el vuelto siempre será 0 en los tickets
    cambio = 0
    
    context = {
        'venta': venta,
        'cambio': cambio,
    }
    return render(request, 'pos/ticket_80mm.html', context)


@login_required
def enviar_ticket_email_view(request, venta_id):
    """Vista para enviar el ticket por correo electrónico como PDF adjunto"""
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string
    from django.http import JsonResponse
    import smtplib
    import io
    from xhtml2pdf import pisa
    
    venta = get_object_or_404(Venta, id=venta_id)
    
    if request.method == 'POST':
        email_destino = request.POST.get('email', '').strip()
        
        if not email_destino:
            return JsonResponse({'success': False, 'error': 'El correo electrónico es requerido'})
        
        # Validar formato de email básico
        if '@' not in email_destino or '.' not in email_destino.split('@')[1]:
            return JsonResponse({'success': False, 'error': 'El correo electrónico no es válido'})
        
        try:
            # IMPORTANTE: monto_recibido siempre es igual al total (no se guarda el monto pagado mayor)
            # Por lo tanto, el vuelto siempre será 0 en los tickets
            cambio = 0
            
            # Renderizar el ticket como HTML (usar template específico para email)
            context = {
                'venta': venta,
                'cambio': cambio,
            }
            html_content = render_to_string('pos/ticket_email.html', context)
            
            # Generar PDF desde HTML
            pdf_buffer = io.BytesIO()
            pisa_status = pisa.CreatePDF(
                html_content,
                dest=pdf_buffer,
                encoding='utf-8'
            )
            
            # Verificar si se generó el PDF correctamente
            if pisa_status.err:
                logger.error(f'Error al generar PDF: {pisa_status.err}')
                return JsonResponse({
                    'success': False,
                    'error': 'Error al generar el PDF del ticket'
                })
            
            # Obtener el contenido del PDF
            pdf_buffer.seek(0)
            pdf_content = pdf_buffer.getvalue()
            pdf_buffer.close()
            
            # Crear el email con el PDF adjunto
            subject = f'Ticket de Venta #{venta.id} - MegaPos By Megadominio.co'
            from_email = 'Ventas Bazar 2025 <noreply@tersacosmeticos.com>'
            
            # Mensaje de texto para el cuerpo del email
            mensaje_texto = f'''
Estimado cliente,

Adjunto encontrará el ticket de su compra #{venta.id}.

Fecha: {venta.fecha.strftime("%d/%m/%Y %H:%M")}
Total: ${venta.total:,}

Gracias por su compra.

Atentamente,
Ventas Bazar 2025
MegaPos By Megadominio.co
            '''.strip()
            
            email = EmailMessage(
                subject=subject,
                body=mensaje_texto,
                from_email=from_email,
                to=[email_destino],
            )
            
            # Adjuntar el PDF
            nombre_archivo = f'Ticket_{venta.id}_{venta.fecha.strftime("%Y%m%d")}.pdf'
            email.attach(nombre_archivo, pdf_content, 'application/pdf')
            
            # Enviar el email
            email.send()
            
            return JsonResponse({
                'success': True,
                'message': f'Ticket enviado exitosamente a {email_destino}'
            })
            
        except ImportError:
            logger.error('xhtml2pdf no está instalado. Instale con: pip install xhtml2pdf')
            return JsonResponse({
                'success': False,
                'error': 'Error: La librería xhtml2pdf no está instalada. Por favor, instálela con: pip install xhtml2pdf'
            })
        except smtplib.SMTPException as e:
            logger.error(f'Error SMTP al enviar ticket por email: {str(e)}', exc_info=True)
            return JsonResponse({
                'success': False,
                'error': f'Error SMTP al enviar el correo: {str(e)}'
            })
        except Exception as e:
            logger.error(f'Error al enviar ticket por email: {str(e)}', exc_info=True)
            return JsonResponse({
                'success': False,
                'error': f'Error al enviar el correo: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def anular_venta_view(request, venta_id):
    """Anular una venta"""
    if not puede_anular_ventas(request.user):
        messages.error(request, 'No tienes permisos para anular ventas. Solo administradores pueden realizar esta acción.')
        return redirect('pos:detalle_venta', venta_id=venta_id)
    
    venta = get_object_or_404(Venta, id=venta_id)
    
    if request.method == 'POST':
        if not venta.anulada:
            venta.anulada = True
            venta.fecha_anulacion = timezone.now()
            venta.usuario_anulacion = request.user
            venta.motivo_anulacion = request.POST.get('motivo', 'Sin especificar')
            venta.save()
            
            # Devolver stock
            for item in venta.items.all():
                producto = item.producto
                producto.stock += item.cantidad
                producto.save()
                
                # Registrar movimiento
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='ingreso',
                    cantidad=item.cantidad,
                    stock_anterior=producto.stock - item.cantidad,
                    stock_nuevo=producto.stock,
                    motivo=f'Anulación venta #{venta.id}',
                    usuario=request.user
                )
            
            # Manejar el dinero recibido si la venta fue en efectivo
            accion_dinero = request.POST.get('accion_dinero', 'mantener')
            if accion_dinero == 'devolver' and venta.metodo_pago == 'efectivo' and venta.total:
                # Usar la función auxiliar para obtener la caja única global
                from datetime import date
                hoy = date.today()
                caja_asociada = obtener_caja_mostrar(None, hoy)  # None porque es caja única global
                
                if caja_asociada:
                    # IMPORTANTE: Siempre devolver el total de la factura, no el monto_recibido
                    # porque monto_recibido siempre es igual al total (no se guarda el monto pagado mayor)
                    from .models import GastoCaja
                    monto_devolver = int(venta.total)
                    
                    # Log para trazabilidad
                    logger.info(f"Anulación venta #{venta.id}: Creando gasto de devolución de ${monto_devolver:,} en caja #{caja_asociada.id} (usuario: {request.user.username})")
                    
                    gasto_creado = GastoCaja.objects.create(
                        tipo='gasto',
                        monto=monto_devolver,
                        descripcion=f'Devolución por anulación de venta #{venta.id} - {venta.motivo_anulacion[:50] if venta.motivo_anulacion else "Sin motivo"}',
                        usuario=request.user,
                        caja_usuario=caja_asociada,
                        fecha=timezone.now()  # Asegurar que tenga fecha actual
                    )
                    
                    # Verificar que se creó correctamente
                    if gasto_creado.id:
                        logger.info(f"Gasto de devolución creado exitosamente: ID={gasto_creado.id}, Caja={caja_asociada.id}, Monto=${monto_devolver:,}")
                        messages.success(request, f'Venta anulada exitosamente. Se restó ${monto_devolver:,} de la caja por devolución al cliente.')
                    else:
                        logger.error(f"Error al crear gasto de devolución: Gasto creado pero sin ID")
                        messages.warning(request, 'Venta anulada exitosamente, pero hubo un problema al registrar el gasto en la caja.')
                else:
                    logger.warning(f"Anulación venta #{venta.id}: No se encontró caja para asociar el gasto de devolución (usuario: {request.user.username})")
                    messages.warning(request, 'Venta anulada exitosamente, pero no se pudo restar de la caja porque no se encontró una caja abierta o cerrada del día actual.')
            else:
                messages.success(request, 'Venta anulada exitosamente')
        else:
            messages.warning(request, 'Esta venta ya está anulada')
    
    return redirect('pos:detalle_venta', venta_id=venta_id)


@login_required
@requiere_rol('Administradores', 'Cajeros')
def caja_view(request):
    """Vista de gestión de caja - Caja única global compartida por todos"""
    from datetime import date
    
    # Obtener o crear la Caja Principal
    caja_principal = Caja.objects.filter(numero=1).first()
    if not caja_principal:
        caja_principal = Caja.objects.create(
            numero=1,
            nombre='Caja Principal',
            activa=True
        )
    
    # Obtener fecha actual
    hoy = date.today()
    inicio_dia = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    fin_dia = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Buscar la única caja del sistema (solo hay una)
    caja_abierta = None
    caja_unica = CajaUsuario.objects.filter(
        caja=caja_principal
    ).order_by('-fecha_apertura').first()
    
    if caja_unica and caja_unica.fecha_cierre is None:
        caja_abierta = caja_unica
    
    # Obtener historial de cajas con estadísticas calculadas
    # Como solo hay una caja, el historial será solo esa caja
    historial_cajas_raw = []
    if caja_unica:
        historial_cajas_raw = [caja_unica]
    
    # Enriquecer cada caja con estadísticas
    historial_cajas = []
    for caja_item in historial_cajas_raw:
        # Calcular ventas de esta caja
        if caja_item.fecha_cierre:
            ventas_caja_item = Venta.objects.filter(
                caja=caja_principal,
                fecha__gte=caja_item.fecha_apertura,
                fecha__lte=caja_item.fecha_cierre,
                completada=True,
                anulada=False
            )
        else:
            # Si está abierta, calcular hasta ahora
            ventas_caja_item = Venta.objects.filter(
                caja=caja_principal,
                fecha__gte=caja_item.fecha_apertura,
                completada=True,
                anulada=False
            )
        
        total_ventas_caja = ventas_caja_item.aggregate(total=Sum('total'))['total'] or 0
        cantidad_ventas_caja = ventas_caja_item.count()
        
        # Obtener TODAS las ventas del período (incluyendo anuladas) para calcular correctamente el saldo
        if caja_item.fecha_cierre:
            ventas_caja_todas_item = Venta.objects.filter(
                caja=caja_principal,
                fecha__gte=caja_item.fecha_apertura,
                fecha__lte=caja_item.fecha_cierre,
                completada=True
            )
        else:
            ventas_caja_todas_item = Venta.objects.filter(
                caja=caja_principal,
                fecha__gte=caja_item.fecha_apertura,
                completada=True
            )
        
        ventas_anuladas_caja_item = ventas_caja_todas_item.filter(anulada=True)
        total_anuladas_caja = ventas_anuladas_caja_item.aggregate(total=Sum('total'))['total'] or 0
        
        # Calcular gastos e ingresos
        from .models import GastoCaja
        gastos_caja_item = GastoCaja.objects.filter(caja_usuario=caja_item)
        
        # IMPORTANTE: Filtrar gastos por el período de la caja (desde fecha_apertura hasta fecha_cierre o ahora)
        # Esto asegura que solo se incluyan los gastos del período específico de esta caja
        if caja_item.fecha_cierre:
            gastos_periodo_caja = gastos_caja_item.filter(
                fecha__gte=caja_item.fecha_apertura,
                fecha__lte=caja_item.fecha_cierre
            )
        else:
            gastos_periodo_caja = gastos_caja_item.filter(
                fecha__gte=caja_item.fecha_apertura
            )
        
        # IMPORTANTE: Excluir los retiros de cierre de caja del total de gastos
        # Los retiros son salidas de dinero pero no son gastos operativos
        gastos_sin_retiros_caja = gastos_periodo_caja.filter(
            tipo='gasto'
        ).exclude(
            descripcion__icontains='Retiro de dinero al cerrar caja'
        )
        total_gastos_caja = gastos_sin_retiros_caja.aggregate(total=Sum('monto'))['total'] or 0
        total_ingresos_caja = gastos_periodo_caja.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0
        
        # Calcular saldo esperado usando la misma lógica que saldo_caja
        # IMPORTANTE: Si hay gastos de devolución, significa que las ventas anuladas ingresaron dinero
        gastos_devolucion_caja = gastos_periodo_caja.filter(
            descripcion__icontains='Devolución por anulación',
            tipo='gasto'
        )
        gastos_devolucion_caja_total = gastos_devolucion_caja.aggregate(total=Sum('monto'))['total'] or 0
        
        monto_inicial_caja = int(caja_item.monto_inicial) if caja_item.monto_inicial else 0
        
        if gastos_devolucion_caja_total > 0:
            # Si hay gastos de devolución, las ventas anuladas ingresaron dinero
            saldo_esperado = monto_inicial_caja + int(total_ventas_caja) + int(total_anuladas_caja) + int(total_ingresos_caja) - int(total_gastos_caja)
        else:
            # Si NO hay gastos de devolución, las ventas anuladas no afectan el dinero físico
            saldo_esperado = monto_inicial_caja + int(total_ventas_caja) + int(total_ingresos_caja) - int(total_gastos_caja)
        
        historial_cajas.append({
            'caja': caja_item,
            'fecha_apertura': caja_item.fecha_apertura,
            'fecha_cierre': caja_item.fecha_cierre,
            'monto_inicial': monto_inicial_caja,
            'monto_final': int(caja_item.monto_final) if caja_item.monto_final else None,
            'total_ventas': int(total_ventas_caja),
            'cantidad_ventas': cantidad_ventas_caja,
            'total_gastos': int(total_gastos_caja),
            'total_ingresos': int(total_ingresos_caja),
            'saldo_esperado': saldo_esperado,
            'diferencia': int(caja_item.monto_final) - saldo_esperado if caja_item.monto_final else None,
        })
    
    # Usar la función auxiliar para obtener la única caja del sistema
    caja_mostrar = obtener_caja_mostrar(None, hoy)  # None porque es caja única global
    
    # Inicializar variables con valores por defecto
    total_ventas = 0
    ventas_caja = []
    gastos_caja = []
    movimientos_unificados = []
    total_gastos = 0
    total_ingresos = 0
    saldo_caja = 0
    ventas_efectivo = 0
    ventas_tarjeta = 0
    ventas_transferencia = 0
    dinero_fisico_caja = 0
    porcentaje_efectivo = 0
    porcentaje_tarjeta = 0
    porcentaje_transferencia = 0
    cantidad_ventas = 0
    promedio_venta = 0
    cantidad_gastos = 0
    cantidad_ingresos = 0
    tiempo_transcurrido = None
    
    if caja_mostrar:
        # Filtrar ventas del período de la caja (desde apertura hasta cierre, o del día si está abierta)
        # IMPORTANTE: Incluir TODAS las ventas (anuladas y no anuladas) para mostrar en movimientos
        # Las ventas anuladas aparecerán como negativas
        if caja_mostrar.fecha_cierre:
            # Si la caja está cerrada, filtrar ventas entre apertura y cierre
            ventas_caja_todas = Venta.objects.filter(
                fecha__gte=caja_mostrar.fecha_apertura,
                fecha__lte=caja_mostrar.fecha_cierre,
                completada=True
            )
        else:
            # Si la caja está abierta, filtrar ventas desde la fecha de apertura de la caja
            ventas_caja_todas = Venta.objects.filter(
                caja=caja_principal,
                fecha__gte=caja_mostrar.fecha_apertura,
                completada=True
            )
        
        # Separar ventas válidas y anuladas para cálculos
        ventas_caja = ventas_caja_todas.filter(anulada=False)
        ventas_anuladas_caja = ventas_caja_todas.filter(anulada=True)
        
        # Calcular total de ventas válidas (para estadísticas)
        total_ventas_raw = ventas_caja.aggregate(total=Sum('total'))['total']
        total_ventas = int(total_ventas_raw) if total_ventas_raw else 0
        
        # Calcular total de ventas anuladas (para restar del saldo)
        total_anuladas_raw = ventas_anuladas_caja.aggregate(total=Sum('total'))['total']
        total_anuladas = int(total_anuladas_raw) if total_anuladas_raw else 0
        
        # Obtener gastos e ingresos de la caja (abierta o cerrada)
        # IMPORTANTE: Filtrar por fecha para incluir solo los del período de la caja
        # (desde fecha_apertura hasta fecha_cierre o ahora)
        from .models import GastoCaja
        gastos_todos = GastoCaja.objects.filter(
            caja_usuario=caja_mostrar
        )
        
        # Filtrar gastos por el período de la caja para cálculos de totales
        if caja_mostrar.fecha_cierre:
            gastos_periodo = gastos_todos.filter(
                fecha__gte=caja_mostrar.fecha_apertura,
                fecha__lte=caja_mostrar.fecha_cierre
            )
        else:
            gastos_periodo = gastos_todos.filter(
                fecha__gte=caja_mostrar.fecha_apertura
            )
        
        # Para el cálculo del saldo, usar solo los gastos del período de la caja
        # IMPORTANTE: Excluir los retiros de cierre de caja del total de gastos
        # Los retiros son salidas de dinero pero no son gastos operativos
        gastos_sin_retiros = gastos_periodo.filter(
            tipo='gasto'
        ).exclude(
            descripcion__icontains='Retiro de dinero al cerrar caja'
        )
        total_gastos_raw = gastos_sin_retiros.aggregate(total=Sum('monto'))['total']
        total_ingresos_raw = gastos_periodo.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total']
        
        total_gastos = int(total_gastos_raw) if total_gastos_raw else 0
        total_ingresos = int(total_ingresos_raw) if total_ingresos_raw else 0
        
        # Obtener monto inicial de la caja
        monto_inicial = int(caja_mostrar.monto_inicial) if caja_mostrar.monto_inicial else 0
        
        # Calcular ventas por método de pago (solo ventas válidas, no anuladas)
        ventas_efectivo = ventas_caja.filter(metodo_pago='efectivo').aggregate(total=Sum('total'))['total'] or 0
        ventas_tarjeta = ventas_caja.filter(metodo_pago='tarjeta').aggregate(total=Sum('total'))['total'] or 0
        ventas_transferencia = ventas_caja.filter(metodo_pago='transferencia').aggregate(total=Sum('total'))['total'] or 0
        
        # Calcular ventas anuladas en efectivo (para restar del dinero físico)
        ventas_anuladas_efectivo = ventas_anuladas_caja.filter(metodo_pago='efectivo').aggregate(total=Sum('total'))['total'] or 0
        
        # Convertir a enteros
        ventas_efectivo = int(ventas_efectivo)
        ventas_tarjeta = int(ventas_tarjeta)
        ventas_transferencia = int(ventas_transferencia)
        ventas_anuladas_efectivo = int(ventas_anuladas_efectivo)
        
        # Calcular dinero en bancos (tarjeta + transferencia)
        dinero_bancos = ventas_tarjeta + ventas_transferencia

        # Calcular dinero físico en caja (solo efectivo + monto inicial + ingresos - gastos)
        # IMPORTANTE sobre ventas anuladas en efectivo:
        # - Si una venta anulada en efectivo ingresó dinero a la caja (monto_recibido > 0), ese dinero SÍ entró físicamente
        # - Luego se devolvió cuando se anuló (gasto de devolución)
        # - Por lo tanto, debemos sumar el dinero que entró (ventas_anuladas_efectivo) y restar el que se devolvió (gastos de devolución)
        # - Si hay gastos de devolución en efectivo, significa que el dinero SÍ entró a la caja antes de anularse
        # Los otros métodos de pago van a cuentas bancarias
        # Verificar si hay gastos de devolución en efectivo para saber si el dinero de las ventas anuladas entró a la caja
        gastos_devolucion_efectivo = gastos_periodo.filter(
            descripcion__icontains='Devolución por anulación',
            tipo='gasto'
        )
        # Filtrar solo los gastos de devolución que corresponden a ventas en efectivo
        gastos_devolucion_efectivo_total = 0
        for gasto in gastos_devolucion_efectivo:
            # Extraer ID de venta de la descripción
            import re
            match = re.search(r'venta #(\d+)', gasto.descripcion, re.IGNORECASE)
            if match:
                venta_id_devolucion = int(match.group(1))
                # Verificar si la venta era en efectivo
                try:
                    venta_devolucion = Venta.objects.get(id=venta_id_devolucion)
                    if venta_devolucion.metodo_pago == 'efectivo':
                        gastos_devolucion_efectivo_total += gasto.monto
                except Venta.DoesNotExist:
                    pass
        
        if gastos_devolucion_efectivo_total > 0:
            # Si hay gastos de devolución en efectivo, significa que el dinero de las ventas anuladas SÍ entró a la caja
            # Por lo tanto, debemos sumarlo al dinero físico (y luego se resta con el gasto de devolución)
            dinero_fisico_caja = monto_inicial + ventas_efectivo + ventas_anuladas_efectivo + total_ingresos - total_gastos
        else:
            # Si NO hay gastos de devolución en efectivo, las ventas anuladas no afectan el dinero físico
            dinero_fisico_caja = monto_inicial + ventas_efectivo + total_ingresos - total_gastos
        
        # Calcular porcentajes
        porcentaje_efectivo = (ventas_efectivo * 100 / total_ventas) if total_ventas > 0 else 0
        porcentaje_tarjeta = (ventas_tarjeta * 100 / total_ventas) if total_ventas > 0 else 0
        porcentaje_transferencia = (ventas_transferencia * 100 / total_ventas) if total_ventas > 0 else 0
        
        # Calcular estadísticas adicionales
        cantidad_ventas = ventas_caja.count()
        promedio_venta = (total_ventas / cantidad_ventas) if cantidad_ventas > 0 else 0
        # Excluir retiros de cierre de caja del conteo de gastos
        cantidad_gastos = gastos_periodo.filter(
            tipo='gasto'
        ).exclude(
            descripcion__icontains='Retiro de dinero al cerrar caja'
        ).count()
        cantidad_ingresos = gastos_periodo.filter(tipo='ingreso').count()
        
        # Calcular tiempo transcurrido desde apertura
        from datetime import timedelta
        tiempo_transcurrido = None
        if caja_mostrar.fecha_apertura:
            ahora = timezone.now()
            diferencia = ahora - caja_mostrar.fecha_apertura
            horas = int(diferencia.total_seconds() / 3600)
            minutos = int((diferencia.total_seconds() % 3600) / 60)
            tiempo_transcurrido = f"{horas}h {minutos}m"
        
        # Para mostrar en la tabla, incluir solo los gastos del período de la caja
        gastos_caja = gastos_periodo.order_by('-fecha')
        
        # Crear lista unificada de movimientos (apertura, ventas, gastos, ingresos)
        movimientos_unificados = []
        
        # Agregar apertura de caja como primer movimiento
        # (monto_inicial ya está definido arriba)
        movimientos_unificados.append({
            'tipo': 'apertura',
            'fecha': caja_mostrar.fecha_apertura,
            'monto': monto_inicial,
            'descripcion': 'Apertura de Caja',
            'usuario': caja_mostrar.usuario,
            'metodo_pago': None,
            'vendedor': None,
            'venta_id': None,
        })
        
        # Verificar si hay gastos de devolución por anulación para evitar duplicación
        # Si existe un GastoCaja de devolución, no agregar el movimiento adicional de anulación
        gastos_devolucion_ids = set()
        for gasto in gastos_periodo:
            if 'Devolución por anulación de venta #' in gasto.descripcion:
                # Extraer el ID de la venta de la descripción
                import re
                match = re.search(r'venta #(\d+)', gasto.descripcion, re.IGNORECASE)
                if match:
                    venta_id_devolucion = int(match.group(1))
                    gastos_devolucion_ids.add(venta_id_devolucion)
        
        # Agregar TODAS las ventas como movimientos (incluyendo anuladas)
        # Las ventas aparecen normalmente, y si están anuladas, se agrega un movimiento adicional de anulación
        # EXCEPTO si ya existe un GastoCaja de devolución (para evitar duplicación)
        ventas_lista_todas = list(ventas_caja_todas.order_by('fecha'))
        for venta in ventas_lista_todas:
            # Obtener nombre de la registradora si existe
            registradora_nombre = None
            if venta.registradora_id:
                registradoras = {
                    1: 'Registradora 1',
                    2: 'Registradora 2',
                    3: 'Registradora 3',
                }
                registradora_nombre = registradoras.get(venta.registradora_id, f'Registradora {venta.registradora_id}')
            
            # Agregar la venta normalmente (incluso si está anulada)
            movimientos_unificados.append({
                'tipo': 'venta',
                'fecha': venta.fecha,
                'monto': venta.total,
                'descripcion': f'Venta #{venta.id}' + (' (Anulada)' if venta.anulada else ''),
                'usuario': venta.usuario,
                'metodo_pago': venta.get_metodo_pago_display(),
                'vendedor': venta.vendedor,
                'venta_id': venta.id,
                'registradora_id': venta.registradora_id,
                'registradora_nombre': registradora_nombre,
                'anulada': venta.anulada,
            })
            
            # Si la venta está anulada, agregar un movimiento adicional de anulación que reste
            # SOLO si NO existe un GastoCaja de devolución (para evitar duplicación)
            if venta.anulada and venta.id not in gastos_devolucion_ids:
                movimientos_unificados.append({
                    'tipo': 'devolucion',  # Tipo especial para devoluciones/anulaciones
                    'fecha': venta.fecha_anulacion if venta.fecha_anulacion else venta.fecha,
                    'monto': -venta.total,  # Monto negativo para restar
                    'descripcion': f'Anulación - Venta #{venta.id}',
                    'usuario': venta.usuario_anulacion if venta.usuario_anulacion else venta.usuario,
                    'metodo_pago': venta.get_metodo_pago_display(),
                    'vendedor': venta.vendedor,
                    'venta_id': venta.id,
                    'registradora_id': venta.registradora_id,
                    'registradora_nombre': registradora_nombre,
                    'anulada': True,
                })
        
        # Agregar gastos e ingresos como movimientos
        # Usar gastos_periodo que ya está filtrado por fecha
        # Convertir a lista para asegurar que se incluyan todos los gastos del período
        gastos_lista = list(gastos_periodo.order_by('fecha'))
        
        # Log para trazabilidad
        logger.debug(f"Caja #{caja_mostrar.id}: Agregando {len(gastos_lista)} gastos/ingresos a movimientos (filtrados por fecha)")
        
        for movimiento in gastos_lista:
            # Identificar si es un retiro por la descripción
            es_retiro = 'Retiro de dinero al cerrar caja' in movimiento.descripcion
            tipo_movimiento = 'retiro' if es_retiro else movimiento.tipo
            
            # Log para trazabilidad de devoluciones
            if 'Devolución por anulación' in movimiento.descripcion:
                logger.debug(f"Movimiento de devolución detectado: ID={movimiento.id}, Monto=${movimiento.monto:,}, Caja={caja_mostrar.id}")
            
            movimientos_unificados.append({
                'tipo': tipo_movimiento,
                'fecha': movimiento.fecha,
                'monto': movimiento.monto,
                'descripcion': movimiento.descripcion,
                'usuario': movimiento.usuario,
                'metodo_pago': None,
                'vendedor': None,
                'venta_id': None,
                'gasto_id': movimiento.id,  # Agregar ID del gasto para trazabilidad
            })
        
        # Ordenar todos los movimientos por fecha ascendente para calcular saldos
        movimientos_unificados.sort(key=lambda x: x['fecha'])
        
        # Calcular saldo antes y después de cada movimiento
        # IMPORTANTE: El saldo inicial es 0 antes de cualquier movimiento
        # La apertura establece el monto inicial, y luego se calculan los saldos secuencialmente
        saldo_actual = 0
        
        for movimiento in movimientos_unificados:
            saldo_antes = saldo_actual
            monto = int(movimiento['monto'])
            
            # Calcular saldo después según el tipo de movimiento
            if movimiento['tipo'] == 'apertura':
                # La apertura agrega el monto inicial
                saldo_despues = saldo_antes + monto
            elif movimiento['tipo'] == 'venta' or movimiento['tipo'] == 'ingreso':
                # Ventas e ingresos suman al saldo
                saldo_despues = saldo_antes + monto
            elif movimiento['tipo'] == 'devolucion':
                # Devoluciones restan del saldo (el monto ya es negativo)
                saldo_despues = saldo_antes + monto  # Sumar porque monto es negativo
            elif movimiento['tipo'] == 'retiro' or movimiento['tipo'] == 'gasto':
                # Retiros y gastos restan del saldo
                saldo_despues = saldo_antes - monto
            else:
                # Por defecto, no afecta el saldo
                saldo_despues = saldo_antes
            
            # Agregar saldo antes y después al movimiento
            movimiento['saldo_antes'] = saldo_antes
            movimiento['saldo_despues'] = saldo_despues
            
            # Actualizar saldo actual para el siguiente movimiento
            saldo_actual = saldo_despues
        
        # Ordenar por fecha descendente para mostrar (más recientes primero)
        movimientos_unificados.sort(key=lambda x: x['fecha'], reverse=True)
        
        # Calcular saldo en caja: Monto Inicial + Ventas Válidas + Ventas Anuladas (que ingresaron dinero) + Ingresos - Gastos
        # IMPORTANTE sobre ventas anuladas:
        # - Si una venta anulada ingresó dinero a la caja (monto_recibido > 0), ese dinero SÍ entró físicamente
        # - Luego se devolvió cuando se anuló (gasto de devolución)
        # - Por lo tanto, debemos sumar el dinero que entró (total_anuladas) y restar el que se devolvió (gastos de devolución)
        # - Si hay gastos de devolución, significa que el dinero SÍ entró a la caja antes de anularse
        # (monto_inicial ya está definido arriba)
        # IMPORTANTE: Los retiros de cierre de caja NO se incluyen en total_gastos (se excluyen del cálculo)
        # Los retiros se muestran como movimientos separados pero no afectan el total de gastos operativos
        # Este es el saldo total que debería haber en la caja (incluyendo todos los métodos de pago)
        # Verificar si hay gastos de devolución para saber si el dinero de las ventas anuladas entró a la caja
        gastos_devolucion_total = gastos_periodo.filter(descripcion__icontains='Devolución por anulación').aggregate(total=Sum('monto'))['total'] or 0
        if gastos_devolucion_total > 0:
            # Si hay gastos de devolución, significa que el dinero de las ventas anuladas SÍ entró a la caja
            # Por lo tanto, debemos sumarlo al saldo (y luego se resta con el gasto de devolución)
            saldo_caja = monto_inicial + total_ventas + total_anuladas + total_ingresos - total_gastos
        else:
            # Si NO hay gastos de devolución, las ventas anuladas no afectan el dinero físico
            saldo_caja = monto_inicial + total_ventas + total_ingresos - total_gastos
        
        # Asegurar que el saldo sea un entero
        saldo_caja = int(saldo_caja)
    
    context = {
        'caja_abierta': caja_abierta,
        'caja_mostrar': caja_mostrar,  # Puede ser abierta o cerrada del día
        'caja_principal': caja_principal,
        'historial_cajas': historial_cajas,
        'total_ventas': total_ventas,
        'ventas_caja': ventas_caja,
        'gastos_caja': gastos_caja,
        'movimientos_unificados': movimientos_unificados,
        'total_gastos': total_gastos,
        'total_ingresos': total_ingresos,
        'saldo_caja': saldo_caja,
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'dinero_bancos': dinero_bancos,
        'dinero_fisico_caja': dinero_fisico_caja,
        'porcentaje_efectivo': porcentaje_efectivo,
        'porcentaje_tarjeta': porcentaje_tarjeta,
        'porcentaje_transferencia': porcentaje_transferencia,
        'cantidad_ventas': cantidad_ventas,
        'promedio_venta': int(promedio_venta),
        'cantidad_gastos': cantidad_gastos,
        'cantidad_ingresos': cantidad_ingresos,
        'tiempo_transcurrido': tiempo_transcurrido,
        'hoy': hoy,
    }
    
    return render(request, 'pos/caja.html', context)


@login_required
@requiere_rol('Administradores', 'Cajeros')
def abrir_caja_view(request):
    """Abrir la caja única global - Sistema diario"""
    from datetime import date
    
    if request.method == 'POST':
        try:
            monto_inicial = int(float(request.POST.get('monto_inicial', 0)))
        except (ValueError, TypeError):
            monto_inicial = 0
        
        hoy = date.today()
        
        # Verificar si la única caja está abierta
        caja_principal = Caja.objects.filter(numero=1).first()
        if not caja_principal:
            caja_principal = Caja.objects.create(
                numero=1,
                nombre='Caja Principal',
                activa=True
            )
        
        caja_abierta_existente = CajaUsuario.objects.filter(
            caja=caja_principal,
            fecha_cierre__isnull=True
        ).first()
        
        if caja_abierta_existente:
            messages.warning(request, 'La caja ya está abierta. Debes cerrarla antes de abrirla nuevamente.')
            return redirect('pos:caja')
        
        # Ahora abrir/reutilizar la única caja del sistema
        try:
            # Buscar si ya existe la única caja (reutilizar la misma siempre)
            caja_existente = CajaUsuario.objects.filter(
                caja=caja_principal
            ).first()
            
            if caja_existente:
                # Reutilizar la caja existente: abrirla nuevamente
                caja_existente.fecha_cierre = None
                caja_existente.monto_inicial = monto_inicial
                caja_existente.monto_final = None
                caja_existente.fecha_apertura = timezone.now()
                caja_existente.usuario = request.user
                caja_existente.save()
                
                logger.info(f'Caja única reutilizada y abierta: ID={caja_existente.id}, Usuario={request.user.username}, Monto inicial=${monto_inicial:,}')
                messages.success(request, f'Caja abierta exitosamente para el día de hoy')
            else:
                # Si no existe ninguna caja, crear la única caja del sistema
                nueva_caja = CajaUsuario.objects.create(
                    caja=caja_principal,
                    usuario=request.user,
                    monto_inicial=monto_inicial
                )
                
                # Verificar que se creó correctamente
                if nueva_caja and nueva_caja.id:
                    logger.info(f'Caja única creada y abierta: ID={nueva_caja.id}, Usuario={request.user.username}, Monto inicial=${monto_inicial:,}')
                    messages.success(request, f'Caja abierta exitosamente para el día de hoy')
                else:
                    logger.error(f'Error al crear caja: nueva_caja={nueva_caja}')
                    messages.error(request, 'Error al crear la caja. Por favor, intenta nuevamente.')
        except Exception as e:
            logger.error(f'Error al abrir caja: {str(e)}', exc_info=True)
            messages.error(request, f'Error al abrir la caja: {str(e)}')
    
    # Redireccionar explícitamente a la página de caja
    from django.http import HttpResponseRedirect
    from django.urls import reverse
    try:
        caja_url = reverse('pos:caja')
        logger.debug(f'Redirigiendo a: {caja_url}')
        return HttpResponseRedirect(caja_url)
    except Exception as e:
        logger.error(f'Error al redirigir: {str(e)}', exc_info=True)
        # Fallback: redireccionar directamente
        return HttpResponseRedirect('/caja/')


@login_required
@requiere_rol('Administradores', 'Cajeros')
def cerrar_caja_view(request):
    """
    Cerrar la única caja del sistema y todas las registradoras activas.
    
    IMPORTANTE: Esta función NO elimina ningún dato del historial:
    - Las ventas se mantienen intactas
    - Los gastos e ingresos se mantienen intactos
    - Los movimientos de caja se mantienen intactos
    - Solo se actualiza el estado de la caja (fecha_cierre y monto_final)
    
    El historial completo permanece disponible para consulta en cualquier momento.
    """
    if request.method == 'POST':
        # Obtener la única caja del sistema
        caja_principal = Caja.objects.filter(numero=1).first()
        if not caja_principal:
            messages.error(request, 'No existe la Caja Principal')
            return redirect('pos:caja')
        
        # Buscar la única caja del sistema
        caja_abierta = CajaUsuario.objects.filter(
            caja=caja_principal,
            fecha_cierre__isnull=True
        ).first()
        
        if caja_abierta:
            monto_final = int(float(request.POST.get('monto_final', 0)))
            # Nuevo: retiros separados (efectivo y bancos)
            # Compatibilidad: si aún llega dinero_retirar, se toma como retiro de efectivo
            dinero_retirar_efectivo = request.POST.get('dinero_retirar_efectivo', None)
            dinero_retirar_bancos = request.POST.get('dinero_retirar_bancos', None)
            if dinero_retirar_efectivo is None and dinero_retirar_bancos is None:
                dinero_retirar_efectivo = request.POST.get('dinero_retirar', 0)
                dinero_retirar_bancos = 0
            dinero_retirar_efectivo = int(float(dinero_retirar_efectivo or 0))
            dinero_retirar_bancos = int(float(dinero_retirar_bancos or 0))
            
            # Calcular el saldo actual de la caja antes de permitir el retiro
            from datetime import date
            from django.db.models import Sum
            hoy = date.today()
            
            # Obtener todas las ventas de la caja (válidas y anuladas)
            # IMPORTANTE: Usar el mismo filtrado que en caja_view para consistencia
            if caja_abierta.fecha_cierre:
                # Si la caja está cerrada, filtrar ventas entre apertura y cierre
                ventas_caja_todas = Venta.objects.filter(
                    caja=caja_principal,
                    fecha__gte=caja_abierta.fecha_apertura,
                    fecha__lte=caja_abierta.fecha_cierre,
                    completada=True
                )
            else:
                # Si la caja está abierta, filtrar ventas desde la fecha de apertura de la caja
                ventas_caja_todas = Venta.objects.filter(
                    caja=caja_principal,
                    fecha__gte=caja_abierta.fecha_apertura,
                    completada=True
                )
            
            # Separar ventas válidas y anuladas
            ventas_caja = ventas_caja_todas.filter(anulada=False)
            ventas_anuladas_caja = ventas_caja_todas.filter(anulada=True)
            
            # Calcular totales (solo para estadísticas)
            total_ventas = int(ventas_caja.aggregate(total=Sum('total'))['total'] or 0)
            total_anuladas = int(ventas_anuladas_caja.aggregate(total=Sum('total'))['total'] or 0)
            
            # IMPORTANTE: Para el saldo disponible, solo contar ventas en EFECTIVO
            # Las ventas con tarjeta y transferencia no generan dinero físico en la caja
            ventas_efectivo = int(ventas_caja.filter(metodo_pago='efectivo').aggregate(total=Sum('total'))['total'] or 0)
            ventas_anuladas_efectivo = int(ventas_anuladas_caja.filter(metodo_pago='efectivo').aggregate(total=Sum('total'))['total'] or 0)
            ventas_tarjeta = int(ventas_caja.filter(metodo_pago='tarjeta').aggregate(total=Sum('total'))['total'] or 0)
            ventas_transferencia = int(ventas_caja.filter(metodo_pago='transferencia').aggregate(total=Sum('total'))['total'] or 0)
            ventas_anuladas_tarjeta = int(ventas_anuladas_caja.filter(metodo_pago='tarjeta').aggregate(total=Sum('total'))['total'] or 0)
            ventas_anuladas_transferencia = int(ventas_anuladas_caja.filter(metodo_pago='transferencia').aggregate(total=Sum('total'))['total'] or 0)
            
            # Obtener gastos e ingresos de la caja
            # IMPORTANTE: Filtrar por fecha para incluir solo los del período de la caja
            # (desde fecha_apertura hasta fecha_cierre o ahora)
            from .models import GastoCaja
            gastos_todos = GastoCaja.objects.filter(
                caja_usuario=caja_abierta
            )
            
            # Filtrar gastos por el período de la caja para cálculos de totales
            if caja_abierta.fecha_cierre:
                gastos_periodo = gastos_todos.filter(
                    fecha__gte=caja_abierta.fecha_apertura,
                    fecha__lte=caja_abierta.fecha_cierre
                )
            else:
                gastos_periodo = gastos_todos.filter(
                    fecha__gte=caja_abierta.fecha_apertura
                )
            
            # IMPORTANTE: Excluir los retiros de cierre de caja del total de gastos
            # Los retiros son salidas de dinero pero no son gastos operativos
            gastos_sin_retiros = gastos_periodo.filter(
                tipo='gasto'
            ).exclude(
                descripcion__icontains='Retiro de dinero al cerrar caja'
            )
            total_gastos = int(gastos_sin_retiros.aggregate(total=Sum('monto'))['total'] or 0)
            total_ingresos = int(gastos_periodo.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0)
            
            # Obtener monto inicial
            monto_inicial = int(caja_abierta.monto_inicial) if caja_abierta.monto_inicial else 0
            
            # Verificar si hay gastos de devolución en efectivo para calcular correctamente el saldo
            gastos_devolucion_efectivo = gastos_periodo.filter(
                descripcion__icontains='Devolución por anulación',
                tipo='gasto'
            )
            # Filtrar solo los gastos de devolución que corresponden a ventas en efectivo
            gastos_devolucion_efectivo_total = 0
            for gasto in gastos_devolucion_efectivo:
                # Extraer ID de venta de la descripción
                import re
                match = re.search(r'venta #(\d+)', gasto.descripcion, re.IGNORECASE)
                if match:
                    venta_id_devolucion = int(match.group(1))
                    # Verificar si la venta era en efectivo
                    try:
                        venta_devolucion = Venta.objects.get(id=venta_id_devolucion)
                        if venta_devolucion.metodo_pago == 'efectivo':
                            gastos_devolucion_efectivo_total += gasto.monto
                    except Venta.DoesNotExist:
                        pass
            
            # Calcular saldo disponible (solo dinero físico en efectivo, sin incluir el retiro que se va a hacer)
            if gastos_devolucion_efectivo_total > 0:
                # Si hay gastos de devolución en efectivo, significa que el dinero de las ventas anuladas SÍ entró a la caja
                # Por lo tanto, debemos sumarlo al dinero físico (y luego se resta con el gasto de devolución)
                saldo_disponible = monto_inicial + ventas_efectivo + ventas_anuladas_efectivo + total_ingresos - total_gastos
            else:
                # Si NO hay gastos de devolución en efectivo, las ventas anuladas no afectan el dinero físico
                saldo_disponible = monto_inicial + ventas_efectivo + total_ingresos - total_gastos

            # Calcular saldo disponible en bancos (tarjeta + transferencia)
            # Nota: ingresos/gastos operativos no se consideran bancos (se asumen efectivo).
            gastos_devolucion_bancos_total = 0
            for gasto in gastos_devolucion_efectivo:
                import re
                match = re.search(r'venta #(\d+)', gasto.descripcion, re.IGNORECASE)
                if match:
                    venta_id_devolucion = int(match.group(1))
                    try:
                        venta_devolucion = Venta.objects.get(id=venta_id_devolucion)
                        if venta_devolucion.metodo_pago in ('tarjeta', 'transferencia'):
                            gastos_devolucion_bancos_total += gasto.monto
                    except Venta.DoesNotExist:
                        pass

            # Si hay devoluciones por anulación en bancos, asumimos que las ventas anuladas sí ingresaron y luego se devolvió.
            if gastos_devolucion_bancos_total > 0:
                saldo_disponible_bancos = (
                    ventas_tarjeta
                    + ventas_transferencia
                    + ventas_anuladas_tarjeta
                    + ventas_anuladas_transferencia
                )
            else:
                saldo_disponible_bancos = ventas_tarjeta + ventas_transferencia
            
            # Validar retiros: efectivo vs bancos (separados)
            # Permitir retirar $0 (cero) incluso si el saldo es negativo, ya que $0 significa no retirar nada
            if dinero_retirar_efectivo > 0 and dinero_retirar_efectivo > saldo_disponible:
                messages.error(
                    request, 
                    f'No se puede retirar ${dinero_retirar_efectivo:,}. El saldo disponible en EFECTIVO en caja es ${saldo_disponible:,}'
                )
                return redirect('pos:caja')

            if dinero_retirar_bancos > 0 and dinero_retirar_bancos > saldo_disponible_bancos:
                messages.error(
                    request,
                    f'No se puede retirar ${dinero_retirar_bancos:,}. El saldo disponible en BANCOS (tarjeta/transferencia) es ${saldo_disponible_bancos:,}'
                )
                return redirect('pos:caja')
            
            # Si hay dinero a retirar, registrarlo como un gasto antes de cerrar
            # Este gasto se agrega al historial y NO se elimina
            if dinero_retirar_efectivo > 0:
                GastoCaja.objects.create(
                    tipo='gasto',
                    monto=dinero_retirar_efectivo,
                    descripcion=f'Retiro de dinero al cerrar caja (Efectivo) - Usuario: {request.user.get_full_name() or request.user.username}',
                    usuario=request.user,
                    caja_usuario=caja_abierta
                )
            if dinero_retirar_bancos > 0:
                GastoCaja.objects.create(
                    tipo='gasto',
                    monto=dinero_retirar_bancos,
                    descripcion=f'Retiro de dinero al cerrar caja (Bancos) - Usuario: {request.user.get_full_name() or request.user.username}',
                    usuario=request.user,
                    caja_usuario=caja_abierta
                )
            
            # IMPORTANTE: Solo actualizar el estado de la caja (fecha_cierre y monto_final)
            # NO se eliminan ventas, gastos, ingresos ni ningún otro dato del historial
            caja_abierta.fecha_cierre = timezone.now()
            caja_abierta.monto_final = monto_final
            caja_abierta.save()
            
            # Cerrar todas las registradoras activas del usuario
            from .models import RegistradoraActiva
            registradoras_cerradas = RegistradoraActiva.objects.filter(
                usuario=request.user
            ).count()
            
            # Eliminar todas las registradoras activas del usuario
            RegistradoraActiva.objects.filter(usuario=request.user).delete()
            
            # Limpiar la sesión de registradora seleccionada
            if 'registradora_seleccionada' in request.session:
                del request.session['registradora_seleccionada']
                request.session.modified = True
            
            # Mensaje de confirmación
            mensaje = 'Caja cerrada exitosamente'
            if dinero_retirar_efectivo > 0 or dinero_retirar_bancos > 0:
                partes = []
                if dinero_retirar_efectivo > 0:
                    partes.append(f'Efectivo: ${dinero_retirar_efectivo:,}')
                if dinero_retirar_bancos > 0:
                    partes.append(f'Bancos: ${dinero_retirar_bancos:,}')
                mensaje += f'. Se registró retiro como salida ({", ".join(partes)}).'
            if registradoras_cerradas > 0:
                mensaje += f' Se cerraron {registradoras_cerradas} registradora(s) activa(s).'
            
            messages.success(request, mensaje)
        else:
            messages.warning(request, 'No hay una caja abierta')
    
    return redirect('pos:caja')


@login_required
@requiere_rol('Administradores', 'Cajeros')
def registrar_gasto_view(request):
    """Registrar un gasto en la caja única global"""
    if request.method == 'POST':
        # Verificar si hay caja abierta (caja única global)
        # No filtrar por fecha ya que solo puede haber una caja abierta
        caja_abierta = CajaUsuario.objects.filter(
            fecha_cierre__isnull=True
        ).order_by('-fecha_apertura').first()
        
        if not caja_abierta:
            messages.error(request, 'No hay una caja abierta para hoy. Debes abrir una caja antes de registrar gastos.')
            return redirect('pos:caja')
        
        # Asegurar que se use la Caja Principal
        caja_principal = Caja.objects.filter(numero=1).first()
        if not caja_principal:
            caja_principal = Caja.objects.create(
                numero=1,
                nombre='Caja Principal',
                activa=True
            )
        
        # Verificar que la caja abierta pertenezca a la Caja Principal
        if caja_abierta.caja != caja_principal:
            messages.error(request, 'La caja abierta no pertenece a la Caja Principal.')
            return redirect('pos:caja')
        
        try:
            monto = int(float(request.POST.get('monto', 0)))
            descripcion = request.POST.get('descripcion', '').strip()
            
            if monto <= 0:
                messages.error(request, 'El monto debe ser mayor a 0')
                return redirect('pos:caja')
            
            if not descripcion:
                messages.error(request, 'La descripción es requerida')
                return redirect('pos:caja')
            
            from .models import GastoCaja
            GastoCaja.objects.create(
                tipo='gasto',
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                caja_usuario=caja_abierta
            )
            
            messages.success(request, f'Gasto de ${monto:,} registrado exitosamente')
        except ValueError:
            messages.error(request, 'Monto inválido')
        except Exception as e:
            messages.error(request, f'Error al registrar gasto: {str(e)}')
    
    return redirect('pos:caja')


@login_required
@requiere_rol('Administradores', 'Cajeros')
def registrar_ingreso_view(request):
    """Registrar una entrada de dinero en la caja única global"""
    if request.method == 'POST':
        # Verificar si hay caja abierta (caja única global)
        # No filtrar por fecha ya que solo puede haber una caja abierta
        caja_abierta = CajaUsuario.objects.filter(
            fecha_cierre__isnull=True
        ).order_by('-fecha_apertura').first()
        
        if not caja_abierta:
            messages.error(request, 'No hay una caja abierta para hoy. Debes abrir una caja antes de registrar ingresos.')
            return redirect('pos:caja')
        
        # Asegurar que se use la Caja Principal
        caja_principal = Caja.objects.filter(numero=1).first()
        if not caja_principal:
            caja_principal = Caja.objects.create(
                numero=1,
                nombre='Caja Principal',
                activa=True
            )
        
        # Verificar que la caja abierta pertenezca a la Caja Principal
        if caja_abierta.caja != caja_principal:
            messages.error(request, 'La caja abierta no pertenece a la Caja Principal.')
            return redirect('pos:caja')
        
        try:
            monto = int(float(request.POST.get('monto', 0)))
            descripcion = request.POST.get('descripcion', '').strip()
            
            if monto <= 0:
                messages.error(request, 'El monto debe ser mayor a 0')
                return redirect('pos:caja')
            
            if not descripcion:
                messages.error(request, 'La descripción es requerida')
                return redirect('pos:caja')
            
            from .models import GastoCaja
            GastoCaja.objects.create(
                tipo='ingreso',
                monto=monto,
                descripcion=descripcion,
                usuario=request.user,
                caja_usuario=caja_abierta
            )
            
            messages.success(request, f'Entrada de dinero de ${monto:,} registrada exitosamente')
        except ValueError:
            messages.error(request, 'Monto inválido')
        except Exception as e:
            messages.error(request, f'Error al registrar ingreso: {str(e)}')
    
    return redirect('pos:caja')


@login_required
@requiere_rol('Administradores')
def reportes_view(request):
    """Vista de reportes"""
    if not puede_ver_reportes(request.user):
        messages.error(request, 'No tienes permisos para ver reportes')
        return redirect('pos:home')
    
    # Obtener tipo de reporte
    tipo_reporte = request.GET.get('tipo', '')
    
    # Si no hay tipo seleccionado, mostrar página de selección
    if not tipo_reporte:
        return render(request, 'pos/reportes.html', {
            'mostrar_seleccion': True,
        })
    
    # Si es inventario, usar la vista de movimientos de inventario
    if tipo_reporte == 'inventario':
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        from django.db.models import Sum, Q
        from .models import MovimientoStock, Producto
        
        # Filtros
        producto_id = request.GET.get('producto')
        tipo_movimiento = request.GET.get('tipo_mov')
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        
        # Base query para movimientos
        movimientos_qs = MovimientoStock.objects.select_related('producto')
        
        if producto_id:
            movimientos_qs = movimientos_qs.filter(producto_id=producto_id)
        
        if tipo_movimiento:
            movimientos_qs = movimientos_qs.filter(tipo=tipo_movimiento)
        
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                movimientos_qs = movimientos_qs.filter(fecha__date__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                movimientos_qs = movimientos_qs.filter(fecha__date__lte=fecha_hasta_obj)
            except ValueError:
                pass
        
        # Agrupar por código + atributo (cada variante aparece por separado)
        # Usar un diccionario con clave (codigo, atributo) para evitar duplicados
        resumen_dict = {}
        
        # Obtener todos los productos únicos (codigo + atributo) que tienen movimientos
        productos_con_movimientos = movimientos_qs.values('producto_id').distinct()
        
        for item in productos_con_movimientos:
            producto = Producto.objects.get(id=item['producto_id'])
            
            # Clave única: código + atributo (None se convierte a '' para comparación)
            atributo_key = producto.atributo if producto.atributo else ''
            clave = (producto.codigo, atributo_key)
            
            # Si ya procesamos esta combinación código+atributo, saltarla
            if clave in resumen_dict:
                continue
            
            # Obtener todos los productos con este código Y atributo (pueden ser múltiples si hay duplicados)
            productos_mismo_codigo_atributo = Producto.objects.filter(
                codigo=producto.codigo,
                atributo=producto.atributo if producto.atributo else None,
                activo=True
            )
            
            # Obtener todos los movimientos de productos con este código Y atributo
            movimientos_producto = movimientos_qs.filter(
                producto__codigo=producto.codigo,
                producto__atributo=producto.atributo if producto.atributo else None
            )
            
            # Calcular entradas (ingreso) - sumar todos los productos con este código+atributo
            total_entradas = movimientos_producto.filter(tipo='ingreso').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            # Calcular salidas (salida) - sumar todos los productos con este código+atributo
            total_salidas = movimientos_producto.filter(tipo='salida').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            # Calcular ajustes (pueden ser positivos o negativos) - sumar todos
            ajustes = movimientos_producto.filter(tipo='ajuste').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            # Neto = entradas - salidas + ajustes
            neto = total_entradas - total_salidas + ajustes
            
            # Stock actual: sumar el stock de todos los productos con este código+atributo
            stock_actual = sum(p.stock for p in productos_mismo_codigo_atributo)
            
            resumen_dict[clave] = {
                'codigo': producto.codigo,
                'nombre': producto.nombre,
                'atributo': producto.atributo if producto.atributo else '-',
                'total_entradas': int(total_entradas),
                'total_salidas': int(total_salidas),
                'ajustes': int(ajustes),
                'neto': int(neto),
                'stock_actual': stock_actual,
            }
        
        # Convertir el diccionario a lista y ordenar por código y luego por atributo
        resumen_productos = list(resumen_dict.values())
        resumen_productos.sort(key=lambda x: (x['codigo'], x['atributo']))
        
        # Paginación: 50 productos por página
        paginator = Paginator(resumen_productos, 50)
        page = request.GET.get('page', 1)
        
        try:
            resumen_paginated = paginator.page(page)
        except PageNotAnInteger:
            resumen_paginated = paginator.page(1)
        except EmptyPage:
            resumen_paginated = paginator.page(paginator.num_pages)
        
        # Obtener lista de productos para el filtro
        productos = Producto.objects.filter(activo=True).order_by('nombre')
        
        context = {
            'tipo_reporte': 'inventario',
            'resumen_productos': resumen_paginated,
            'productos': productos,
            'producto_id': producto_id,
            'tipo_movimiento': tipo_movimiento,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
        return render(request, 'pos/reportes.html', context)
    
    # Si es caja, continuar con el código existente
    # Obtener fechas del filtro
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    # Ventas del mes actual por defecto
    try:
        if not fecha_desde:
            fecha_desde = timezone.now().replace(day=1).date()
        else:
            fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    except Exception:
        fecha_desde = timezone.now().replace(day=1).date()
    
    try:
        if not fecha_hasta:
            fecha_hasta = timezone.now().date()
        else:
            fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except Exception:
        fecha_hasta = timezone.now().date()

    if fecha_hasta < fecha_desde:
        fecha_desde, fecha_hasta = fecha_hasta, fecha_desde

    # Rango datetime (timezone-aware): por fecha calendario
    from datetime import time, timedelta
    tz = timezone.get_current_timezone()
    inicio_dt = timezone.make_aware(datetime.combine(fecha_desde, time.min), tz)
    fin_dt = timezone.make_aware(datetime.combine(fecha_hasta, time.max), tz)

    # En modo web simplificado: siempre todas las cajas del rango (sin forzar CajaUsuario)
    caja_usuario_id = ''
    caja_usuario_sel = None

    def _display_user(u):
        if not u:
            return ''
        return (u.get_full_name() or u.username)

    def _build_movimientos_caja():
        """
        Reporte estilo "Todos los Movimientos" (Caja):
        Apertura + Ventas + Gastos/Ingresos/Retiros con saldo antes/después.
        """
        items = []

        # Apertura(s)
        aperturas_qs = CajaUsuario.objects.none()
        aperturas_qs = CajaUsuario.objects.filter(
            fecha_apertura__gte=inicio_dt,
            fecha_apertura__lte=fin_dt
        ).select_related('usuario', 'caja')

        for cu in aperturas_qs:
            items.append({
                'fecha': cu.fecha_apertura,
                'tipo': 'Apertura',
                'descripcion': f'Apertura de Caja (CajaUsuario #{cu.id})',
                'delta': int(cu.monto_inicial or 0),
                'monto_abs': int(cu.monto_inicial or 0),
                'metodo_pago': '-',
                'usuario': _display_user(cu.usuario),
                'venta_id': None,
            })

        # Ventas
        for v in Venta.objects.filter(
            fecha__gte=inicio_dt,
            fecha__lte=fin_dt,
            completada=True
        ).select_related('usuario', 'vendedor').order_by('fecha'):
            vendedor = v.vendedor if v.vendedor else v.usuario
            desc = f'Venta #{v.id}'
            if v.registradora_id:
                desc = f'{desc} - Registradora {v.registradora_id}'
            delta = int(v.total or 0)
            tipo = 'Venta'
            if v.anulada:
                tipo = 'Venta Anulada'
                delta = -delta
            items.append({
                'fecha': v.fecha,
                'tipo': tipo,
                'descripcion': desc,
                'delta': delta,
                'monto_abs': int(v.total or 0),
                'metodo_pago': v.get_metodo_pago_display(),
                'usuario': _display_user(vendedor),
                'venta_id': v.id,
            })

        # Gastos / Ingresos (y retiros)
        gastos_qs = GastoCaja.objects.filter(fecha__gte=inicio_dt, fecha__lte=fin_dt).select_related('usuario', 'caja_usuario')

        for g in gastos_qs.order_by('fecha'):
            es_retiro = bool(g.descripcion and 'Retiro de dinero al cerrar caja' in g.descripcion)
            tipo = 'Retiro' if es_retiro else ('Gasto' if g.tipo == 'gasto' else 'Ingreso')
            delta = int(g.monto or 0)
            if g.tipo == 'gasto':
                delta = -delta
            items.append({
                'fecha': g.fecha,
                'tipo': tipo,
                'descripcion': g.descripcion,
                'delta': delta,
                'monto_abs': int(g.monto or 0),
                'metodo_pago': '-',
                'usuario': _display_user(g.usuario),
                'venta_id': None,
            })

        def _orden_tipo(t):
            if t == 'Apertura':
                return 0
            if t.startswith('Venta'):
                return 1
            return 2

        items.sort(key=lambda x: (x['fecha'], _orden_tipo(x['tipo'])))

        saldo = 0
        out = []
        for it in items:
            saldo_antes = saldo
            saldo_despues = saldo + int(it['delta'])
            saldo = saldo_despues
            out.append({
                **it,
                'saldo_antes': saldo_antes,
                'saldo_despues': saldo_despues,
                'fecha_local': timezone.localtime(it['fecha'], tz),
            })
        return out

    # Export (mismo endpoint): CSV / Excel
    export_tipo = request.GET.get('export')
    export_format = (request.GET.get('format') or 'csv').strip().lower()
    if export_tipo in ('ventas', 'movimientos', 'cajas', 'movimientos_caja'):
        from django.http import HttpResponse

        if export_format in ('xlsx', 'excel'):
            from openpyxl import Workbook

            wb = Workbook()

            def _set_headers(ws, headers):
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True)

            if export_tipo == 'ventas':
                ws = wb.active
                ws.title = 'Ventas'
                _set_headers(ws, [
                    'ID', 'Fecha', 'Total', 'Metodo Pago', 'Anulada',
                    'Usuario', 'Vendedor', 'Registradora',
                    'Items Cantidad', 'Items Detalle'
                ])

                qs = Venta.objects.filter(
                    fecha__gte=inicio_dt, fecha__lte=fin_dt, completada=True
                ).select_related('usuario', 'vendedor').prefetch_related('items__producto').order_by('fecha')

                ws_items = wb.create_sheet('Items')
                _set_headers(ws_items, [
                    'VentaID', 'Fecha', 'Producto', 'Cantidad', 'Precio Unitario', 'Subtotal'
                ])

                for v in qs:
                    items = list(v.items.all())
                    items_cant = sum((it.cantidad or 0) for it in items)
                    items_detalle = ' | '.join([
                        f"{it.producto.nombre} x{it.cantidad} (${it.subtotal})" for it in items
                    ])
                    fecha_local = timezone.localtime(v.fecha, tz).strftime('%Y-%m-%d %H:%M:%S')
                    ws.append([
                        v.id,
                        fecha_local,
                        v.total,
                        v.metodo_pago,
                        int(v.anulada),
                        (v.usuario.username if v.usuario else ''),
                        (v.vendedor.username if v.vendedor else ''),
                        (v.registradora_id or ''),
                        items_cant,
                        items_detalle,
                    ])
                    for it in items:
                        ws_items.append([
                            v.id,
                            fecha_local,
                            it.producto.nombre,
                            it.cantidad,
                            it.precio_unitario,
                            it.subtotal,
                        ])

            elif export_tipo == 'movimientos':
                ws = wb.active
                ws.title = 'Movimientos'
                _set_headers(ws, [
                    'Fecha', 'Tipo', 'Descripcion', 'Monto', 'Metodo', 'Usuario', 'CajaUsuarioID'
                ])

                # Ventas
                for v in Venta.objects.filter(
                    fecha__gte=inicio_dt, fecha__lte=fin_dt, completada=True
                ).select_related('usuario').order_by('fecha'):
                    tipo = 'venta_anulada' if v.anulada else 'venta'
                    fecha_local = timezone.localtime(v.fecha, tz).strftime('%Y-%m-%d %H:%M:%S')
                    ws.append([
                        fecha_local,
                        tipo,
                        f'Venta #{v.id}',
                        v.total,
                        v.metodo_pago,
                        (v.usuario.username if v.usuario else ''),
                        '',
                    ])

                # Gastos/Ingresos
                for g in GastoCaja.objects.filter(
                    fecha__gte=inicio_dt, fecha__lte=fin_dt
                ).select_related('usuario', 'caja_usuario').order_by('fecha'):
                    fecha_local = timezone.localtime(g.fecha, tz).strftime('%Y-%m-%d %H:%M:%S')
                    ws.append([
                        fecha_local,
                        g.tipo,
                        g.descripcion,
                        g.monto,
                        '',
                        (g.usuario.username if g.usuario else ''),
                        (g.caja_usuario_id or ''),
                    ])

            elif export_tipo == 'cajas':
                ws = wb.active
                ws.title = 'Cajas'
                _set_headers(ws, [
                    'CajaUsuarioID', 'Caja', 'Apertura', 'Cierre', 'Monto Inicial', 'Monto Final', 'Usuario'
                ])
                for cu in CajaUsuario.objects.filter(
                    fecha_apertura__lte=fin_dt
                ).filter(
                    Q(fecha_cierre__gte=inicio_dt) | Q(fecha_cierre__isnull=True)
                ).select_related('usuario', 'caja').order_by('fecha_apertura'):
                    apertura_local = timezone.localtime(cu.fecha_apertura, tz).strftime('%Y-%m-%d %H:%M:%S')
                    cierre_local = timezone.localtime(cu.fecha_cierre, tz).strftime('%Y-%m-%d %H:%M:%S') if cu.fecha_cierre else ''
                    ws.append([
                        cu.id,
                        (cu.caja.nombre if cu.caja else ''),
                        apertura_local,
                        cierre_local,
                        cu.monto_inicial,
                        (cu.monto_final if cu.monto_final is not None else ''),
                        (cu.usuario.username if cu.usuario else ''),
                    ])
            else:  # movimientos_caja
                ws = wb.active
                ws.title = 'MovimientosCaja'
                _set_headers(ws, [
                    'Fecha/Hora', 'Tipo', 'Descripcion', 'Monto', 'Saldo Antes', 'Saldo Despues', 'Metodo de Pago', 'Usuario'
                ])
                for m in _build_movimientos_caja():
                    ws.append([
                        m['fecha_local'].strftime('%Y-%m-%d %H:%M:%S'),
                        m['tipo'],
                        m['descripcion'],
                        m['delta'],
                        m['saldo_antes'],
                        m['saldo_despues'],
                        m['metodo_pago'] or '-',
                        m['usuario'] or '',
                    ])

            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"reporte_{export_tipo}_{fecha_desde.isoformat()}_a_{fecha_hasta.isoformat()}.xlsx"
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # CSV (por defecto)
        import csv
        filename = f"reporte_{export_tipo}_{fecha_desde.isoformat()}_a_{fecha_hasta.isoformat()}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)

        if export_tipo == 'ventas':
            writer.writerow([
                'ID', 'Fecha', 'Total', 'Metodo Pago', 'Completada', 'Anulada',
                'Usuario', 'Vendedor', 'Registradora',
                'Items Cantidad', 'Items Detalle'
            ])
            for v in Venta.objects.filter(
                fecha__gte=inicio_dt, fecha__lte=fin_dt, completada=True
            ).select_related('usuario', 'vendedor').prefetch_related('items__producto'):
                items = list(v.items.all())
                items_cant = sum((it.cantidad or 0) for it in items)
                items_detalle = ' | '.join([
                    f"{it.producto.nombre} x{it.cantidad} (${it.subtotal})" for it in items
                ])
                fecha_local = timezone.localtime(v.fecha, tz).strftime('%Y-%m-%d %H:%M:%S')
                writer.writerow([
                    v.id,
                    fecha_local,
                    v.total,
                    v.metodo_pago,
                    int(v.completada),
                    int(v.anulada),
                    (v.usuario.username if v.usuario else ''),
                    (v.vendedor.username if v.vendedor else ''),
                    (v.registradora_id or ''),
                    items_cant,
                    items_detalle,
                ])
            return response

        if export_tipo == 'movimientos':
            writer.writerow([
                'Fecha', 'Tipo', 'Descripcion', 'Monto', 'Metodo', 'Usuario', 'CajaUsuarioID'
            ])
            # Ventas
            for v in Venta.objects.filter(
                fecha__gte=inicio_dt, fecha__lte=fin_dt, completada=True
            ).select_related('usuario'):
                tipo = 'venta_anulada' if v.anulada else 'venta'
                fecha_local = timezone.localtime(v.fecha, tz).strftime('%Y-%m-%d %H:%M:%S')
                writer.writerow([
                    fecha_local,
                    tipo,
                    f'Venta #{v.id}',
                    v.total,
                    v.metodo_pago,
                    (v.usuario.username if v.usuario else ''),
                    '',
                ])
            # Gastos/Ingresos
            for g in GastoCaja.objects.filter(fecha__gte=inicio_dt, fecha__lte=fin_dt).select_related('usuario', 'caja_usuario'):
                fecha_local = timezone.localtime(g.fecha, tz).strftime('%Y-%m-%d %H:%M:%S')
                writer.writerow([
                    fecha_local,
                    g.tipo,
                    g.descripcion,
                    g.monto,
                    '',
                    (g.usuario.username if g.usuario else ''),
                    (g.caja_usuario_id or ''),
                ])
            return response

        if export_tipo == 'cajas':
            writer.writerow([
                'CajaUsuarioID', 'Apertura', 'Cierre', 'Monto Inicial', 'Monto Final', 'Usuario'
            ])
            for cu in CajaUsuario.objects.filter(
                fecha_apertura__lte=fin_dt
            ).filter(
                Q(fecha_cierre__gte=inicio_dt) | Q(fecha_cierre__isnull=True)
            ).select_related('usuario', 'caja').order_by('-fecha_apertura'):
                apertura_local = timezone.localtime(cu.fecha_apertura, tz).strftime('%Y-%m-%d %H:%M:%S')
                cierre_local = timezone.localtime(cu.fecha_cierre, tz).strftime('%Y-%m-%d %H:%M:%S') if cu.fecha_cierre else ''
                writer.writerow([
                    cu.id,
                    apertura_local,
                    cierre_local,
                    cu.monto_inicial,
                    (cu.monto_final if cu.monto_final is not None else ''),
                    (cu.usuario.username if cu.usuario else ''),
                ])
            return response

        if export_tipo == 'movimientos_caja':
            writer.writerow([
                'Fecha/Hora', 'Tipo', 'Descripcion', 'Monto', 'Saldo Antes', 'Saldo Despues', 'Metodo de Pago', 'Usuario'
            ])
            for m in _build_movimientos_caja():
                writer.writerow([
                    m['fecha_local'].strftime('%Y-%m-%d %H:%M:%S'),
                    m['tipo'],
                    m['descripcion'],
                    m['delta'],
                    m['saldo_antes'],
                    m['saldo_despues'],
                    m['metodo_pago'] or '-',
                    m['usuario'] or '',
                ])
            return response

    # Ventas (incluye anuladas; se desglosa)
    ventas_qs = Venta.objects.filter(
        fecha__gte=inicio_dt,
        fecha__lte=fin_dt,
        completada=True
    )
    ventas_validas = ventas_qs.filter(anulada=False)
    ventas_anuladas = ventas_qs.filter(anulada=True)

    total_ventas = int(ventas_validas.aggregate(total=Sum('total'))['total'] or 0)
    total_anuladas = int(ventas_anuladas.aggregate(total=Sum('total'))['total'] or 0)
    cantidad_ventas = ventas_validas.count()
    cantidad_anuladas = ventas_anuladas.count()
    promedio_venta = int(total_ventas / cantidad_ventas) if cantidad_ventas > 0 else 0

    # Top productos (solo ventas válidas)
    top_productos = ItemVenta.objects.filter(
        venta__in=ventas_validas
    ).values('producto__nombre').annotate(
        total_vendido=Sum('cantidad'),
        total_valor=Sum('subtotal')
    ).order_by('-total_vendido')[:15]

    # Ventas por método de pago (válidas)
    ventas_por_metodo = ventas_validas.values('metodo_pago').annotate(
        cantidad=Count('id'),
        total=Sum('total')
    ).order_by('-total')

    # Resumen por usuario / vendedor (solo válidas)
    resumen_por_usuario = ventas_validas.values('usuario__username').annotate(
        cantidad=Count('id'),
        total=Sum('total')
    ).order_by('-total')

    resumen_por_vendedor = ventas_validas.values('vendedor__username').annotate(
        cantidad=Count('id'),
        total=Sum('total')
    ).order_by('-total')

    # Totales por método (válidas)
    ventas_efectivo = int(ventas_validas.filter(metodo_pago='efectivo').aggregate(total=Sum('total'))['total'] or 0)
    ventas_tarjeta = int(ventas_validas.filter(metodo_pago='tarjeta').aggregate(total=Sum('total'))['total'] or 0)
    ventas_transferencia = int(ventas_validas.filter(metodo_pago='transferencia').aggregate(total=Sum('total'))['total'] or 0)
    dinero_bancos = ventas_tarjeta + ventas_transferencia

    # Caja: cajas que se solapan con el rango
    cajas_qs = CajaUsuario.objects.filter(
        fecha_apertura__lte=fin_dt
    ).filter(
        Q(fecha_cierre__gte=inicio_dt) | Q(fecha_cierre__isnull=True)
    ).select_related('usuario', 'caja').order_by('-fecha_apertura')

    # Movimientos (gastos/ingresos/retiros) por rango (no depende de una caja específica)
    movimientos_qs = GastoCaja.objects.filter(fecha__gte=inicio_dt, fecha__lte=fin_dt).select_related('usuario', 'caja_usuario').order_by('-fecha')
    total_gastos = int(movimientos_qs.filter(tipo='gasto').exclude(descripcion__icontains='Retiro de dinero al cerrar caja').aggregate(total=Sum('monto'))['total'] or 0)
    total_ingresos = int(movimientos_qs.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or 0)
    total_retiros = int(movimientos_qs.filter(tipo='gasto', descripcion__icontains='Retiro de dinero al cerrar caja').aggregate(total=Sum('monto'))['total'] or 0)

    # Resumen diario (por fecha local)
    from django.db.models.functions import TruncDate
    dia_expr = TruncDate('fecha', tzinfo=tz)

    # Saldo inicial por dia: suma de montos iniciales de cajas abiertas ese dia (segun corte)
    dia_apertura_expr = TruncDate('fecha_apertura', tzinfo=tz)

    saldo_inicial_map = {
        r['dia']: int(r['saldo_inicial'] or 0)
        for r in CajaUsuario.objects.filter(
            fecha_apertura__gte=inicio_dt,
            fecha_apertura__lte=fin_dt
        ).annotate(
            dia=dia_apertura_expr
        ).values('dia').annotate(
            saldo_inicial=Sum('monto_inicial')
        )
    }

    ventas_diarias_qs = ventas_qs.annotate(dia=dia_expr)
    ventas_validas_diarias = ventas_diarias_qs.filter(anulada=False)
    ventas_anuladas_diarias = ventas_diarias_qs.filter(anulada=True)

    ventas_ok_map = {
        r['dia']: r for r in ventas_validas_diarias.values('dia').annotate(
            total_ventas=Sum('total'),
            cantidad_ventas=Count('id'),
            ventas_efectivo=Sum('total', filter=Q(metodo_pago='efectivo')),
            ventas_tarjeta=Sum('total', filter=Q(metodo_pago='tarjeta')),
            ventas_transferencia=Sum('total', filter=Q(metodo_pago='transferencia')),
        )
    }
    ventas_an_map = {
        r['dia']: r for r in ventas_anuladas_diarias.values('dia').annotate(
            total_anuladas=Sum('total'),
            cantidad_anuladas=Count('id'),
        )
    }

    movs_diarias_qs = GastoCaja.objects.filter(
        fecha__gte=inicio_dt, fecha__lte=fin_dt
    ).annotate(dia=dia_expr)

    movs_g_map = {
        r['dia']: r for r in movs_diarias_qs.filter(tipo='gasto').exclude(
            descripcion__icontains='Retiro de dinero al cerrar caja'
        ).values('dia').annotate(total_gastos=Sum('monto'), cantidad_gastos=Count('id'))
    }
    movs_i_map = {
        r['dia']: r for r in movs_diarias_qs.filter(tipo='ingreso').values('dia').annotate(
            total_ingresos=Sum('monto'),
            cantidad_ingresos=Count('id')
        )
    }
    movs_r_map = {
        r['dia']: r for r in movs_diarias_qs.filter(
            tipo='gasto', descripcion__icontains='Retiro de dinero al cerrar caja'
        ).values('dia').annotate(total_retiros=Sum('monto'), cantidad_retiros=Count('id'))
    }

    dias = sorted(set(
        list(ventas_ok_map.keys())
        + list(ventas_an_map.keys())
        + list(movs_g_map.keys())
        + list(movs_i_map.keys())
        + list(movs_r_map.keys())
    ))

    def _n(v):
        return int(v or 0)

    resumen_diario = []
    for dia in dias:
        vo = ventas_ok_map.get(dia, {})
        va = ventas_an_map.get(dia, {})
        mg = movs_g_map.get(dia, {})
        mi = movs_i_map.get(dia, {})
        mr = movs_r_map.get(dia, {})

        saldo_inicial = int(saldo_inicial_map.get(dia, 0) or 0)
        ventas_ef = _n(vo.get('ventas_efectivo'))
        ventas_tj = _n(vo.get('ventas_tarjeta'))
        ventas_tf = _n(vo.get('ventas_transferencia'))
        gastos_sr = _n(mg.get('total_gastos'))
        ingresos = _n(mi.get('total_ingresos'))
        retiros = _n(mr.get('total_retiros'))

        # Neto operativo (sin retiros): efectivo + ingresos - gastos.
        # Los retiros se muestran separados y no afectan este neto.
        neto_operativo = ventas_ef + ingresos - gastos_sr

        resumen_diario.append({
            'dia': dia,
            'saldo_inicial': saldo_inicial,
            'ventas_total': _n(vo.get('total_ventas')),
            'ventas_cantidad': _n(vo.get('cantidad_ventas')),
            'anuladas_total': _n(va.get('total_anuladas')),
            'anuladas_cantidad': _n(va.get('cantidad_anuladas')),
            'ventas_efectivo': ventas_ef,
            'ventas_tarjeta': ventas_tj,
            'ventas_transferencia': ventas_tf,
            'gastos_sin_retiro_total': gastos_sr,
            'gastos_sin_retiro_cantidad': _n(mg.get('cantidad_gastos')),
            'ingresos_total': ingresos,
            'ingresos_cantidad': _n(mi.get('cantidad_ingresos')),
            'retiros_total': retiros,
            'retiros_cantidad': _n(mr.get('cantidad_retiros')),
            'neto_operativo': int(neto_operativo),
        })

    # Paginación (ventas y movimientos)
    from django.core.paginator import Paginator
    ventas_page = request.GET.get('page_ventas', 1)
    movs_page = request.GET.get('page_movs', 1)
    cajas_page = request.GET.get('page_cajas', 1)
    movs_caja_page = request.GET.get('page_movs_caja', 1)

    ventas_paginated = Paginator(
        ventas_qs.select_related('usuario', 'vendedor').prefetch_related('items__producto').order_by('-fecha'), 50
    ).get_page(ventas_page)
    movs_paginated = Paginator(movimientos_qs, 50).get_page(movs_page)
    cajas_paginated = Paginator(cajas_qs, 25).get_page(cajas_page)
    movs_caja_list = _build_movimientos_caja()
    movs_caja_paginated = Paginator(movs_caja_list, 100).get_page(movs_caja_page)

    total_productos = Producto.objects.filter(activo=True).count()

    context = {
        'tipo_reporte': 'caja',
        'modo_simple': True,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'inicio_dt': inicio_dt,
        'fin_dt': fin_dt,
        'usar_corte_dia': False,
        'corte_dia': '',
        'ocultar_cajas_sin_mov': False,

        # Ventas
        'total_ventas': total_ventas,
        'cantidad_ventas': cantidad_ventas,
        'promedio_venta': promedio_venta,
        'total_anuladas': total_anuladas,
        'cantidad_anuladas': cantidad_anuladas,
        'ventas_efectivo': ventas_efectivo,
        'ventas_tarjeta': ventas_tarjeta,
        'ventas_transferencia': ventas_transferencia,
        'dinero_bancos': dinero_bancos,
        'ventas_por_metodo': ventas_por_metodo,
        'resumen_por_usuario': resumen_por_usuario,
        'resumen_por_vendedor': resumen_por_vendedor,
        'top_productos': top_productos,
        'ventas_detalle': ventas_paginated,

        # Caja / movimientos
        'cajas_detalle': cajas_paginated,
        'movimientos_detalle': movs_paginated,
        'movimientos_caja_detalle': movs_caja_paginated,
        'total_gastos': total_gastos,
        'total_ingresos': total_ingresos,
        'total_retiros': total_retiros,
        'resumen_diario': resumen_diario,
        'caja_usuario_id': caja_usuario_id,
        'caja_usuario_sel': caja_usuario_sel,

        'total_productos': total_productos,
    }

    return render(request, 'pos/reportes.html', context)


@login_required
@requiere_rol('Administradores', 'Inventario')
def movimientos_inventario_view(request):
    """Vista de movimientos de inventario (trazabilidad completa)"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Obtener todos los movimientos de stock
    movimientos_list = MovimientoStock.objects.select_related('producto', 'usuario').order_by('-fecha')
    
    # Filtros
    producto_id = request.GET.get('producto')
    tipo_movimiento = request.GET.get('tipo')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if producto_id:
        movimientos_list = movimientos_list.filter(producto_id=producto_id)
    
    if tipo_movimiento:
        movimientos_list = movimientos_list.filter(tipo=tipo_movimiento)
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            movimientos_list = movimientos_list.filter(fecha__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            movimientos_list = movimientos_list.filter(fecha__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # Paginación: 50 movimientos por página
    paginator = Paginator(movimientos_list, 50)
    page = request.GET.get('page', 1)
    
    try:
        movimientos = paginator.page(page)
    except PageNotAnInteger:
        movimientos = paginator.page(1)
    except EmptyPage:
        movimientos = paginator.page(paginator.num_pages)
    
    # Obtener lista de productos para el filtro
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    context = {
        'movimientos': movimientos,
        'productos': productos,
        'producto_id': producto_id,
        'tipo_movimiento': tipo_movimiento,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'pos/movimientos_inventario.html', context)


@login_required
@requiere_rol('Administradores', 'Inventario')
def inventario_view(request):
    """Vista unificada de inventario (Ingreso y Salida de mercancía)"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    ingresos_list = IngresoMercancia.objects.all().order_by('-fecha')
    salidas_list = SalidaMercancia.objects.all().order_by('-fecha')
    
    # Paginación: 20 por página
    paginator_ingresos = Paginator(ingresos_list, 20)
    paginator_salidas = Paginator(salidas_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        ingresos = paginator_ingresos.page(page)
    except PageNotAnInteger:
        ingresos = paginator_ingresos.page(1)
    except EmptyPage:
        ingresos = paginator_ingresos.page(paginator_ingresos.num_pages)
    
    try:
        salidas = paginator_salidas.page(page)
    except PageNotAnInteger:
        salidas = paginator_salidas.page(1)
    except EmptyPage:
        salidas = paginator_salidas.page(paginator_salidas.num_pages)
    
    # Obtener el tipo de pestaña activa desde la URL
    tipo_activo = request.GET.get('tipo', 'ingreso')  # 'ingreso' o 'salida'
    
    context = {
        'ingresos': ingresos,
        'salidas': salidas,
        'tipo_activo': tipo_activo,
    }
    
    return render(request, 'pos/inventario.html', context)


@login_required
@requiere_rol('Administradores', 'Inventario')
def ingreso_mercancia_view(request):
    """Vista de ingreso de mercancía (redirige a inventario)"""
    from django.urls import reverse
    return redirect('{}?tipo=ingreso'.format(reverse('pos:inventario')))


@login_required
@requiere_rol('Administradores', 'Inventario')
def salida_mercancia_view(request):
    """Vista de salida de mercancía (redirige a inventario)"""
    from django.urls import reverse
    return redirect('{}?tipo=salida'.format(reverse('pos:inventario')))


@login_required
@requiere_rol('Administradores')
def marketing_view(request):
    """Vista de ranking de ventas por vendedor"""
    from django.contrib.auth.models import User
    from datetime import timedelta
    
    # Obtener rango de fechas (últimos 30 días por defecto)
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if not fecha_desde:
        fecha_desde = (timezone.now() - timedelta(days=30)).date()
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    
    if not fecha_hasta:
        fecha_hasta = timezone.now().date()
    else:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Obtener TODAS las ventas (incluyendo anuladas) para calcular total bruto
    ventas_todas = Venta.objects.filter(
        fecha__date__gte=fecha_desde,
        fecha__date__lte=fecha_hasta,
        completada=True,
        vendedor__isnull=False
    )
    
    # Obtener ventas NO anuladas para cantidad y promedio
    ventas = ventas_todas.filter(anulada=False)
    
    # Obtener ventas anuladas para restar
    ventas_anuladas = ventas_todas.filter(anulada=True)
    
    # Calcular ranking por vendedor (TODAS LAS VENTAS, incluyendo anuladas)
    # Esto nos da el total bruto de cada vendedor
    ranking_vendedores_bruto = ventas_todas.values('vendedor').annotate(
        total_ventas_bruto=Sum('total'),  # Total bruto (incluye anuladas)
        cantidad_ventas_total=Count('id')
    )
    
    # Calcular estadísticas de anulaciones por vendedor
    anulaciones_por_vendedor = ventas_anuladas.values('vendedor').annotate(
        total_anuladas=Sum('total'),
        cantidad_anuladas=Count('id')
    )
    
    # Calcular estadísticas de ventas válidas por vendedor (para cantidad y promedio)
    ventas_validas_por_vendedor = ventas.values('vendedor').annotate(
        cantidad_ventas=Count('id'),
        promedio_venta=Avg('total')
    )
    
    # Crear diccionarios para fácil acceso
    anulaciones_dict = {item['vendedor']: item for item in anulaciones_por_vendedor}
    ventas_validas_dict = {item['vendedor']: item for item in ventas_validas_por_vendedor}
    
    # Enriquecer con información del usuario y estadísticas de anulaciones
    ranking_completo = []
    for item in ranking_vendedores_bruto:
        try:
            vendedor = User.objects.get(id=item['vendedor'])
            # Obtener estadísticas de anulaciones para este vendedor
            anulaciones_vendedor = anulaciones_dict.get(item['vendedor'], {})
            total_anuladas = anulaciones_vendedor.get('total_anuladas', 0) or 0
            cantidad_anuladas = anulaciones_vendedor.get('cantidad_anuladas', 0) or 0
            
            # Obtener estadísticas de ventas válidas
            ventas_validas_vendedor = ventas_validas_dict.get(item['vendedor'], {})
            cantidad_ventas = ventas_validas_vendedor.get('cantidad_ventas', 0) or 0
            promedio_venta = ventas_validas_vendedor.get('promedio_venta', 0) or 0
            
            # Calcular total neto (ventas brutas - anulaciones)
            total_ventas_bruto = float(item['total_ventas_bruto'] or 0)
            total_anuladas_decimal = float(total_anuladas or 0)
            total_ventas_neto = round(total_ventas_bruto - total_anuladas_decimal, 2)
            
            ranking_completo.append({
                'vendedor': vendedor,
                'nombre_completo': vendedor.get_full_name() or vendedor.username,
                'username': vendedor.username,
                'total_ventas': total_ventas_neto,  # Total neto después de restar anulaciones
                'total_ventas_bruto': round(total_ventas_bruto, 2),  # Total bruto (sin restar anulaciones)
                'cantidad_ventas': cantidad_ventas,
                'promedio_venta': round(float(promedio_venta or 0), 2),
                'total_anuladas': round(total_anuladas_decimal, 2),
                'cantidad_anuladas': cantidad_anuladas,
            })
        except User.DoesNotExist:
            continue
    
    # Estadísticas generales
    # Total bruto (todas las ventas, incluyendo anuladas)
    total_general_bruto = ventas_todas.aggregate(total=Sum('total'))['total'] or 0
    # Cantidad de ventas válidas (no anuladas)
    cantidad_general = ventas.count()
    
    # Estadísticas de anulaciones generales
    total_anuladas_general = ventas_anuladas.aggregate(total=Sum('total'))['total'] or 0
    cantidad_anuladas_general = ventas_anuladas.count()
    
    # Calcular total neto general (ventas brutas - anulaciones)
    total_anuladas_general_decimal = float(total_anuladas_general or 0)
    total_general = round(float(total_general_bruto or 0) - total_anuladas_general_decimal, 2)
    promedio_general = round(total_general / cantidad_general, 2) if cantidad_general > 0 else 0
    
    # Reordenar ranking por total_ventas neto (después de restar anulaciones)
    ranking_completo = sorted(ranking_completo, key=lambda x: x['total_ventas'], reverse=True)
    
    # Actualizar posiciones después de reordenar
    for i, item in enumerate(ranking_completo, 1):
        item['posicion'] = i
    
    # Top vendedor
    top_vendedor = ranking_completo[0] if ranking_completo else None
    
    context = {
        'ranking_vendedores': ranking_completo,
        'total_general': total_general,
        'cantidad_general': cantidad_general,
        'promedio_general': promedio_general,
        'top_vendedor': top_vendedor,
        'total_anuladas_general': round(total_anuladas_general_decimal, 2),
        'cantidad_anuladas_general': cantidad_anuladas_general,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'pos/marketing.html', context)


def formulario_clientes_view(request):
    """Vista de formulario de clientes (pública)"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        tipo_interes = request.POST.get('tipo_interes', 'mayorista')
        empresa = request.POST.get('empresa', '').strip()
        mensaje = request.POST.get('mensaje', '').strip()
        
        # Validación básica
        if not nombre or not email:
            messages.error(request, 'Por favor completa los campos requeridos (Nombre y Email)')
            return render(request, 'pos/formulario_clientes.html')
        
        try:
            ClientePotencial.objects.create(
                nombre=nombre,
                email=email,
                telefono=telefono or None,
                tipo_interes=tipo_interes,
                empresa=empresa or None,
                mensaje=mensaje or None,
                estado='nuevo'
            )
            messages.success(request, 'Cliente registrado exitosamente. Nos pondremos en contacto contigo pronto.')
            return redirect('pos:formulario_clientes')
        except Exception as e:
            messages.error(request, 'Error al registrar el cliente. Por favor intenta nuevamente.')
            return render(request, 'pos/formulario_clientes.html')
    
    return render(request, 'pos/formulario_clientes.html')


@login_required
def clientes_potenciales_view(request):
    """Vista de clientes potenciales"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    clientes_list = ClientePotencial.objects.all().order_by('-fecha_registro')
    
    # Paginación: 20 clientes por página
    paginator = Paginator(clientes_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        clientes = paginator.page(page)
    except PageNotAnInteger:
        clientes = paginator.page(1)
    except EmptyPage:
        clientes = paginator.page(paginator.num_pages)
    
    context = {
        'clientes': clientes,
    }
    
    return render(request, 'pos/clientes_potenciales.html', context)


@login_required
@requiere_rol('Administradores')
def usuarios_view(request):
    """Vista de gestión de usuarios"""
    if not puede_ver_reportes(request.user):
        messages.error(request, 'No tienes permisos para acceder')
        return redirect('pos:home')
    
    from django.contrib.auth.models import User
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    usuarios_list = User.objects.all().order_by('username')
    
    # Paginación: 20 usuarios por página
    paginator = Paginator(usuarios_list, 20)
    page = request.GET.get('page', 1)
    
    try:
        usuarios = paginator.page(page)
    except PageNotAnInteger:
        usuarios = paginator.page(1)
    except EmptyPage:
        usuarios = paginator.page(paginator.num_pages)
    
    context = {
        'usuarios': usuarios,
    }
    
    return render(request, 'pos/usuarios.html', context)


# ============================================
# SISTEMA DE CARRITO CON SESIÓN
# ============================================

def get_carrito(request, tab_id=None):
    """
    Obtener o crear el carrito en la sesión.
    Si se proporciona tab_id, cada pestaña tendrá su propio carrito.
    Si no se proporciona tab_id, usa el carrito por defecto (compatibilidad hacia atrás).
    """
    # Si no hay tab_id, usar el carrito por defecto (compatibilidad)
    if not tab_id:
        if 'carrito' not in request.session:
            request.session['carrito'] = {}
        return request.session['carrito']
    
    # Si hay tab_id, usar carritos múltiples por pestaña
    if 'carritos' not in request.session:
        request.session['carritos'] = {}
    
    if tab_id not in request.session['carritos']:
        request.session['carritos'][tab_id] = {}
    
    return request.session['carritos'][tab_id]


@login_required
def buscar_productos_view(request):
    """Búsqueda de productos (AJAX) - Insensible a tildes"""
    import unicodedata
    
    def normalizar_texto(texto):
        """Normalizar texto removiendo tildes y acentos"""
        if not texto:
            return ''
        # Normalizar a NFD (descomponer caracteres con acentos)
        texto_normalizado = unicodedata.normalize('NFD', texto.lower())
        # Remover marcas diacríticas (tildes, acentos)
        texto_sin_acentos = ''.join(
            char for char in texto_normalizado 
            if unicodedata.category(char) != 'Mn'
        )
        return texto_sin_acentos
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'productos': []})
    
    query_normalizada = normalizar_texto(query)
    
    # Obtener todos los productos activos y filtrar en Python
    # (más flexible para búsqueda insensible a tildes)
    productos_activos = Producto.objects.filter(activo=True)
    
    resultados = []
    for producto in productos_activos:
        nombre_norm = normalizar_texto(producto.nombre)
        codigo_norm = normalizar_texto(producto.codigo)
        codigo_barras_norm = normalizar_texto(producto.codigo_barras or '')
        
        # Buscar en los campos normalizados
        if (query_normalizada in nombre_norm or 
            query_normalizada in codigo_norm or 
            query_normalizada in codigo_barras_norm):
            resultados.append({
                'id': producto.id,
                'nombre': producto.nombre,
                'codigo': producto.codigo,
                'codigo_barras': producto.codigo_barras or '',
                'precio': int(producto.precio),
                'stock': producto.stock,
            })
            
            # Limitar a 10 resultados
            if len(resultados) >= 10:
                break
    
    return JsonResponse({'productos': resultados})


@login_required
def agregar_al_carrito_view(request):
    """Agregar producto al carrito"""
    if request.method == 'POST':
        try:
            producto_id = int(request.POST.get('producto_id'))
            cantidad = int(request.POST.get('cantidad', 1))
            tab_id = request.POST.get('tab_id')  # Obtener tab_id del POST
            
            producto = get_object_or_404(Producto, id=producto_id, activo=True)
            
            if cantidad <= 0:
                return JsonResponse({'success': False, 'error': 'La cantidad debe ser mayor a 0'})
            
            if producto.stock < cantidad:
                return JsonResponse({
                    'success': False,
                    'error': f'Stock insuficiente. Disponible: {producto.stock}'
                })
            
            carrito = get_carrito(request, tab_id)
            producto_key = str(producto_id)
            
            if producto_key in carrito:
                nueva_cantidad = carrito[producto_key]['cantidad'] + cantidad
                if nueva_cantidad > producto.stock:
                    return JsonResponse({
                        'success': False,
                        'error': f'Stock insuficiente. Disponible: {producto.stock}'
                    })
                carrito[producto_key]['cantidad'] = nueva_cantidad
                # Mantener el timestamp original para preservar el orden
            else:
                # Agregar timestamp para mantener el orden de inserción
                from time import time
                carrito[producto_key] = {
                    'producto_id': producto_id,
                    'nombre': producto.nombre,
                    'codigo': producto.codigo,
                    'atributo': producto.atributo or '',
                    'precio': int(producto.precio),
                    'cantidad': cantidad,
                    'stock': producto.stock,
                    'orden': time(),  # Timestamp para mantener el orden
                }
            
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': f'{cantidad} unidad(es) de {producto.nombre} agregada(s)'
            })
            
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Datos inválidos'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def actualizar_cantidad_carrito_view(request, producto_id):
    """Actualizar cantidad de un item en el carrito"""
    if request.method == 'POST':
        try:
            cantidad = int(request.POST.get('cantidad', 1))
            tab_id = request.POST.get('tab_id')
            producto = get_object_or_404(Producto, id=producto_id)
            carrito = get_carrito(request, tab_id)
            producto_key = str(producto_id)
            
            if producto_key not in carrito:
                return JsonResponse({'success': False, 'error': 'Producto no está en el carrito'})
            
            if cantidad <= 0:
                # Eliminar del carrito
                del carrito[producto_key]
            else:
                if cantidad > producto.stock:
                    return JsonResponse({
                        'success': False,
                        'error': f'Stock insuficiente. Disponible: {producto.stock}'
                    })
                carrito[producto_key]['cantidad'] = cantidad
            
            request.session.modified = True
            return JsonResponse({'success': True})
            
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Cantidad inválida'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def actualizar_precio_carrito_view(request, producto_id):
    """Actualizar precio de un item en el carrito"""
    if request.method == 'POST':
        try:
            nuevo_precio = int(float(request.POST.get('precio', 0)))
            tab_id = request.POST.get('tab_id')
            carrito = get_carrito(request, tab_id)
            producto_key = str(producto_id)
            
            if producto_key not in carrito:
                return JsonResponse({'success': False, 'error': 'Producto no está en el carrito'})
            
            if nuevo_precio < 0:
                return JsonResponse({'success': False, 'error': 'El precio no puede ser negativo'})
            
            carrito[producto_key]['precio'] = nuevo_precio
            request.session.modified = True
            
            return JsonResponse({'success': True})
            
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Precio inválido'})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def eliminar_item_carrito_view(request, producto_id):
    """Eliminar item del carrito"""
    if request.method == 'GET':
        tab_id = request.GET.get('tab_id')
        carrito = get_carrito(request, tab_id)
        producto_key = str(producto_id)
        
        if producto_key in carrito:
            del carrito[producto_key]
            request.session.modified = True
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def limpiar_carrito_view(request):
    """Limpiar todo el carrito"""
    if request.method == 'GET':
        tab_id = request.GET.get('tab_id')
        if tab_id:
            if 'carritos' not in request.session:
                request.session['carritos'] = {}
            request.session['carritos'][tab_id] = {}
        else:
            request.session['carrito'] = {}
        request.session.modified = True
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def procesar_venta_completa_view(request):
    """Procesar venta completa desde el carrito de sesión"""
    if request.method == 'POST':
        try:
            # Verificar si hay caja abierta (caja única global)
            # No filtrar por fecha ya que solo puede haber una caja abierta
            caja_abierta = CajaUsuario.objects.filter(
                fecha_cierre__isnull=True
            ).order_by('-fecha_apertura').first()
            
            if not caja_abierta:
                return JsonResponse({
                    'success': False, 
                    'error': 'Debes abrir una caja antes de realizar ventas. Por favor, abre la caja desde el Dashboard.'
                })
            
            tab_id = request.POST.get('tab_id')
            carrito = get_carrito(request, tab_id)
            
            if not carrito:
                return JsonResponse({'success': False, 'error': 'El carrito está vacío'})
            
            metodo_pago = request.POST.get('metodo_pago', 'efectivo')
            monto_recibido = request.POST.get('monto_recibido')
            vendedor_id = request.POST.get('vendedor_id')
            
            # Calcular total (ordenado por orden de inserción)
            items_ordenados = sorted(
                carrito.items(),
                key=lambda x: x[1].get('orden', 0)
            )
            total = sum(
                item_data['precio'] * item_data['cantidad']
                for key, item_data in items_ordenados
            )
            
            # Validar monto recibido si es efectivo
            # IMPORTANTE: Siempre guardamos el total como monto_recibido, no el monto pagado
            # El vuelto se calcula y muestra, pero no se guarda
            monto_recibido_float = None
            if metodo_pago == 'efectivo':
                if monto_recibido:
                    # Limpiar el monto (quitar puntos y comas)
                    import re
                    monto_limpio = re.sub(r'[^0-9]', '', str(monto_recibido))
                    monto_pagado = int(float(monto_limpio)) if monto_limpio else 0
                    if monto_pagado < total:
                        return JsonResponse({
                            'success': False,
                            'error': f'Monto insuficiente. Total: ${total:,.0f}, Recibido: ${monto_pagado:,.0f}'
                        })
                    # Siempre guardar el total como monto_recibido, no el monto pagado
                    monto_recibido_float = int(total)
                else:
                    # Si no se proporciona monto recibido, usar el total como monto recibido (sin vuelto)
                    monto_recibido_float = int(total)
            
            # Obtener vendedor
            vendedor = None
            if vendedor_id:
                from django.contrib.auth.models import User
                try:
                    vendedor = User.objects.get(id=vendedor_id)
                except User.DoesNotExist:
                    pass
            
            # Obtener registradora de la sesión
            registradora_seleccionada = request.session.get('registradora_seleccionada', None)
            registradora_id = None
            if registradora_seleccionada:
                registradora_id = registradora_seleccionada.get('id')
            
            # Crear la venta
            # IMPORTANTE: monto_recibido siempre será igual al total (no se guarda el monto pagado mayor)
            venta = Venta.objects.create(
                usuario=request.user,
                vendedor=vendedor,
                metodo_pago=metodo_pago,
                monto_recibido=monto_recibido_float if monto_recibido_float else None,
                registradora_id=registradora_id,
                completada=True
            )
            
            # Agregar items y actualizar stock (ordenados por orden de inserción)
            items_ordenados = sorted(
                carrito.items(),
                key=lambda x: x[1].get('orden', 0)
            )
            for key, item_data in items_ordenados:
                producto = Producto.objects.get(id=item_data['producto_id'])
                cantidad = item_data['cantidad']
                precio = item_data['precio']
                
                # Verificar stock nuevamente
                if producto.stock < cantidad:
                    venta.delete()
                    return JsonResponse({
                        'success': False,
                        'error': f'Stock insuficiente para {producto.nombre}'
                    })
                
                # Crear item de venta
                ItemVenta.objects.create(
                    venta=venta,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=precio,
                    subtotal=precio * cantidad
                )
                
                # Actualizar stock
                producto.stock -= cantidad
                producto.save()
                
                # Registrar movimiento de stock
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='salida',
                    cantidad=cantidad,
                    stock_anterior=producto.stock + cantidad,
                    stock_nuevo=producto.stock,
                    motivo=f'Venta #{venta.id}',
                    usuario=request.user
                )
            
            # Actualizar total de la venta
            venta.total = total
            
            # Asignar siempre a la Caja Principal (todas las ventas van a la misma caja)
            caja_principal = Caja.objects.filter(numero=1).first()
            if caja_principal:
                venta.caja = caja_principal
            
            venta.save()
            
            # Limpiar carrito de esta pestaña
            if tab_id:
                if 'carritos' not in request.session:
                    request.session['carritos'] = {}
                request.session['carritos'][tab_id] = {}
            else:
                request.session['carrito'] = {}
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'venta_id': venta.id,
                'total': int(total),
                'message': f'Venta #{venta.id} procesada exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


@login_required
def api_usuarios_view(request):
    """API para obtener lista de usuarios (para selector de vendedor)"""
    from django.contrib.auth.models import User
    usuarios = User.objects.filter(is_active=True).order_by('username')
    
    usuarios_data = []
    for usuario in usuarios:
        usuarios_data.append({
            'id': usuario.id,
            'username': usuario.username,
            'nombre_completo': usuario.get_full_name() or usuario.username,
            'email': usuario.email or '',
        })
    
    return JsonResponse({'usuarios': usuarios_data})


# ============================================
# GESTIÓN DE PRODUCTOS
# ============================================

@login_required
@requiere_rol('Administradores', 'Inventario')
def crear_producto_view(request):
    """Vista para crear un nuevo producto"""
    if request.method == 'POST':
        try:
            codigo = request.POST.get('codigo')
            codigo_barras = request.POST.get('codigo_barras') or None
            nombre = request.POST.get('nombre')
            atributo = request.POST.get('atributo') or None
            precio = int(float(request.POST.get('precio', 0)))
            stock = int(request.POST.get('stock', 0))
            activo = request.POST.get('activo') == 'on'
            
            # Verificar si el código ya existe
            if Producto.objects.filter(codigo=codigo).exists():
                messages.error(request, f'El código {codigo} ya existe')
                return redirect('pos:crear_producto')
            
            producto = Producto.objects.create(
                codigo=codigo,
                codigo_barras=codigo_barras,
                nombre=nombre,
                atributo=atributo,
                precio=precio,
                stock=stock,
                activo=activo
            )
            
            # Manejar imagen si se sube
            if 'imagen' in request.FILES:
                producto.imagen = request.FILES['imagen']
                producto.save()
            
            messages.success(request, f'Producto {producto.nombre} creado exitosamente')
            return redirect('pos:productos')
        except Exception as e:
            messages.error(request, f'Error al crear producto: {str(e)}')
    
    return render(request, 'pos/crear_producto.html')


@login_required
@requiere_rol('Administradores', 'Inventario')
def editar_producto_view(request, producto_id):
    """Vista para editar un producto"""
    producto = get_object_or_404(Producto, id=producto_id)
    
    if request.method == 'POST':
        try:
            producto.codigo = request.POST.get('codigo')
            producto.codigo_barras = request.POST.get('codigo_barras') or None
            producto.nombre = request.POST.get('nombre')
            producto.atributo = request.POST.get('atributo') or None
            producto.precio = int(float(request.POST.get('precio', 0)))
            # IMPORTANTE: No actualizar el stock desde aquí
            # El stock solo se modifica mediante movimientos de inventario (ventas, ingresos, salidas)
            # producto.stock = int(request.POST.get('stock', 0))  # COMENTADO: No se actualiza el stock
            producto.activo = request.POST.get('activo') == 'on'
            
            # Manejar imagen si se sube
            if 'imagen' in request.FILES:
                producto.imagen = request.FILES['imagen']
            
            producto.save()
            messages.success(request, f'Producto {producto.nombre} actualizado exitosamente')
            return redirect('pos:productos')
        except Exception as e:
            messages.error(request, f'Error al actualizar producto: {str(e)}')
    
    # Convertir precio a string numérico para que se muestre correctamente en el input
    # Los DecimalField de Django pueden tener problemas con inputs type="number"
    try:
        precio_float = str(int(producto.precio)) if producto.precio else "0"
    except (ValueError, TypeError):
        precio_float = "0"
    
    context = {
        'producto': producto,
        'precio_float': precio_float,
    }
    return render(request, 'pos/editar_producto.html', context)


# ============================================
# GESTIÓN DE INGRESOS DE MERCADERÍA
# ============================================

@login_required
@requiere_rol('Administradores', 'Inventario')
def crear_ingreso_view(request):
    """Vista para crear un nuevo ingreso de mercancía"""
    if request.method == 'POST':
        try:
            proveedor = request.POST.get('proveedor')
            numero_factura = request.POST.get('numero_factura') or None
            observaciones = request.POST.get('observaciones') or None
            
            ingreso = IngresoMercancia.objects.create(
                proveedor=proveedor,
                numero_factura=numero_factura,
                observaciones=observaciones,
                usuario=request.user,
                completado=False
            )
            
            # Procesar items
            items_data = json.loads(request.POST.get('items', '[]'))
            total = 0
            
            for item_data in items_data:
                producto_id = int(item_data['producto_id'])
                cantidad = int(item_data['cantidad'])
                precio_compra = int(float(item_data.get('precio_compra', 0)))
                
                producto = Producto.objects.get(id=producto_id)
                subtotal = cantidad * precio_compra
                
                ItemIngresoMercancia.objects.create(
                    ingreso=ingreso,
                    producto=producto,
                    cantidad=cantidad,
                    precio_compra=precio_compra,
                    subtotal=subtotal
                )
                
                total += subtotal
            
            ingreso.total = total
            ingreso.save()
            
            messages.success(request, f'Ingreso #{ingreso.id} creado exitosamente')
            return redirect('pos:detalle_ingreso', ingreso_id=ingreso.id)
        except Exception as e:
            messages.error(request, f'Error al crear ingreso: {str(e)}')
    
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    context = {'productos': productos}
    return render(request, 'pos/crear_ingreso.html', context)


@login_required
@requiere_rol('Administradores', 'Inventario')
def detalle_ingreso_view(request, ingreso_id):
    """Vista de detalle de ingreso de mercancía"""
    ingreso = get_object_or_404(IngresoMercancia, id=ingreso_id)
    
    if request.method == 'POST':
        # Manejar verificación/desverificación de items (AJAX)
        if 'item_id' in request.POST:
            item_id = int(request.POST.get('item_id'))
            verificar = request.POST.get('verificar') == '1'
            
            try:
                item = ItemIngresoMercancia.objects.get(id=item_id, ingreso=ingreso)
                item.verificado = verificar
                item.save()
                return JsonResponse({'success': True})
            except ItemIngresoMercancia.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Item no encontrado'})
        
        # Manejar completar ingreso (parcial o completo)
        elif 'completar' in request.POST:
            # Buscar items verificados que aún no han sido procesados
            items_a_procesar = ingreso.items.filter(verificado=True, procesado=False)
            
            if items_a_procesar.count() == 0:
                # Verificar si hay items verificados pero ya procesados
                items_verificados_no_procesados = ingreso.items.filter(verificado=True, procesado=False).count()
                if items_verificados_no_procesados == 0:
                    # Verificar si todos están procesados
                    total_items = ingreso.items.count()
                    items_procesados_total = ingreso.items.filter(procesado=True).count()
                    if items_procesados_total == total_items:
                        messages.info(request, 'Todos los items ya han sido procesados')
                    else:
                        messages.error(request, 'Debes verificar al menos un item antes de completar el ingreso')
                return redirect('pos:detalle_ingreso', ingreso_id=ingreso_id)
            
            # Procesar solo los items verificados que no han sido procesados
            items_procesados = 0
            for item in items_a_procesar:
                stock_anterior = item.producto.stock
                item.producto.stock += item.cantidad
                item.producto.save()
                
                # Registrar movimiento de stock
                MovimientoStock.objects.create(
                    producto=item.producto,
                    tipo='ingreso',
                    cantidad=item.cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=item.producto.stock,
                    motivo=f'Ingreso #{ingreso.id} - {ingreso.proveedor}',
                    usuario=request.user
                )
                
                # Marcar item como procesado
                item.procesado = True
                item.save()
                items_procesados += 1
            
            # Verificar si todos los items están procesados para marcar el ingreso como completado
            total_items = ingreso.items.count()
            items_procesados_total = ingreso.items.filter(procesado=True).count()
            items_pendientes = ingreso.items.filter(verificado=False).count()
            
            if items_procesados_total == total_items:
                # Todos los items están procesados
                ingreso.completado = True
                ingreso.save()
                messages.success(request, f'Ingreso completado totalmente. {items_procesados} items procesados en esta operación. Total: {items_procesados_total} items.')
            else:
                # Ingreso parcial - aún hay items pendientes
                items_no_verificados = ingreso.items.filter(verificado=False).count()
                items_verificados_no_procesados = ingreso.items.filter(verificado=True, procesado=False).count()
                
                mensaje = f'Ingreso parcial completado. {items_procesados} items procesados en esta operación.'
                if items_verificados_no_procesados > 0:
                    mensaje += f' {items_verificados_no_procesados} items verificados pendientes de procesar.'
                if items_no_verificados > 0:
                    mensaje += f' {items_no_verificados} items aún no verificados.'
                
                messages.warning(request, mensaje)
    
    context = {'ingreso': ingreso}
    return render(request, 'pos/detalle_ingreso.html', context)


# ============================================
# GESTIÓN DE SALIDAS DE MERCADERÍA
# ============================================

@login_required
@requiere_rol('Administradores', 'Inventario')
def crear_salida_view(request):
    """Vista para crear una nueva salida de mercancía"""
    if request.method == 'POST':
        try:
            tipo = request.POST.get('tipo')
            destino = request.POST.get('destino') or None
            motivo = request.POST.get('motivo')
            
            salida = SalidaMercancia.objects.create(
                tipo=tipo,
                destino=destino,
                motivo=motivo,
                usuario=request.user,
                completado=False
            )
            
            # Procesar items
            items_data = json.loads(request.POST.get('items', '[]'))
            
            for item_data in items_data:
                producto_id = int(item_data['producto_id'])
                cantidad = int(item_data['cantidad'])
                
                producto = Producto.objects.get(id=producto_id)
                
                if producto.stock < cantidad:
                    salida.delete()
                    messages.error(request, f'Stock insuficiente para {producto.nombre}')
                    return redirect('pos:crear_salida')
                
                ItemSalidaMercancia.objects.create(
                    salida=salida,
                    producto=producto,
                    cantidad=cantidad
                )
            
            messages.success(request, f'Salida #{salida.id} creada exitosamente')
            return redirect('pos:detalle_salida', salida_id=salida.id)
        except Exception as e:
            messages.error(request, f'Error al crear salida: {str(e)}')
    
    productos = Producto.objects.filter(activo=True, stock__gt=0).order_by('nombre')
    context = {'productos': productos}
    return render(request, 'pos/crear_salida.html', context)


@login_required
@requiere_rol('Administradores', 'Inventario')
def detalle_salida_view(request, salida_id):
    """Vista de detalle de salida de mercancía"""
    salida = get_object_or_404(SalidaMercancia, id=salida_id)
    
    if request.method == 'POST' and 'completar' in request.POST:
        if not salida.completado:
            # Actualizar stock de productos
            for item in salida.items.all():
                if item.producto.stock < item.cantidad:
                    messages.error(request, f'Stock insuficiente para {item.producto.nombre}')
                    return redirect('pos:detalle_salida', salida_id=salida_id)
                
                item.producto.stock -= item.cantidad
                item.producto.save()
                
                # Registrar movimiento de stock
                MovimientoStock.objects.create(
                    producto=item.producto,
                    tipo='salida',
                    cantidad=item.cantidad,
                    stock_anterior=item.producto.stock + item.cantidad,
                    stock_nuevo=item.producto.stock,
                    motivo=f'Salida #{salida.id} - {salida.get_tipo_display()}',
                    usuario=request.user
                )
            
            salida.completado = True
            salida.save()
            messages.success(request, 'Salida completada y stock actualizado')
        else:
            messages.warning(request, 'Esta salida ya está completada')
    
    context = {'salida': salida}
    return render(request, 'pos/detalle_salida.html', context)


# ============================================
# GESTIÓN DE USUARIOS
# ============================================

@login_required
@requiere_rol('Administradores')
def crear_usuario_view(request):
    """Vista para crear un nuevo usuario"""
    from django.contrib.auth.models import User, Group
    from .models import PerfilUsuario
    
    if request.method == 'POST':
        try:
            username = request.POST.get('username')
            email = request.POST.get('email')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            password = request.POST.get('password')
            is_staff = request.POST.get('is_staff') == 'on'
            is_active = request.POST.get('is_active') == 'on'
            grupo_id = request.POST.get('grupo')
            
            if User.objects.filter(username=username).exists():
                messages.error(request, 'El nombre de usuario ya existe')
                return redirect('pos:crear_usuario')
            
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=is_staff,
                is_active=is_active
            )
            
            # Asignar grupo
            if grupo_id:
                grupo = Group.objects.get(id=grupo_id)
                user.groups.add(grupo)
            
            # Crear perfil
            PerfilUsuario.objects.create(usuario=user)
            
            messages.success(request, f'Usuario {user.username} creado exitosamente')
            return redirect('pos:usuarios')
        except Exception as e:
            messages.error(request, f'Error al crear usuario: {str(e)}')
    
    grupos = Group.objects.all()
    context = {'grupos': grupos}
    return render(request, 'pos/crear_usuario.html', context)


@login_required
@requiere_rol('Administradores')
def editar_usuario_view(request, usuario_id):
    """Vista para editar un usuario"""
    from django.contrib.auth.models import User, Group
    from .models import PerfilUsuario
    
    usuario = get_object_or_404(User, id=usuario_id)
    
    if request.method == 'POST':
        try:
            usuario.email = request.POST.get('email')
            usuario.first_name = request.POST.get('first_name')
            usuario.last_name = request.POST.get('last_name')
            usuario.is_staff = request.POST.get('is_staff') == 'on'
            usuario.is_active = request.POST.get('is_active') == 'on'
            
            # Cambiar contraseña si se proporciona
            password = request.POST.get('password')
            if password:
                usuario.set_password(password)
            
            usuario.save()
            
            # Actualizar grupos
            grupo_id = request.POST.get('grupo')
            usuario.groups.clear()
            if grupo_id:
                grupo = Group.objects.get(id=grupo_id)
                usuario.groups.add(grupo)
            
            messages.success(request, f'Usuario {usuario.username} actualizado exitosamente')
            return redirect('pos:usuarios')
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
    
    grupos = Group.objects.all()
    usuario_grupos = usuario.groups.all()
    context = {
        'usuario': usuario,
        'grupos': grupos,
        'usuario_grupos': usuario_grupos,
    }
    return render(request, 'pos/editar_usuario.html', context)


# ============================================
# GESTIÓN DE CLIENTES POTENCIALES
# ============================================

@login_required
def gestionar_cliente_view(request, cliente_id):
    """Vista para gestionar un cliente potencial"""
    cliente = get_object_or_404(ClientePotencial, id=cliente_id)
    
    if request.method == 'POST':
        try:
            cliente.estado = request.POST.get('estado')
            cliente.notas_internas = request.POST.get('notas_internas') or None
            
            if cliente.estado == 'contactado' and not cliente.fecha_contacto:
                cliente.fecha_contacto = timezone.now()
                cliente.usuario_contacto = request.user
            
            cliente.save()
            messages.success(request, 'Cliente actualizado exitosamente')
            return redirect('pos:clientes_potenciales')
        except Exception as e:
            messages.error(request, f'Error al actualizar cliente: {str(e)}')
    
    context = {'cliente': cliente}
    return render(request, 'pos/gestionar_cliente.html', context)


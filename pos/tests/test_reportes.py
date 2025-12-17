"""
Tests para el módulo de Reportes
"""
import csv
import io
from datetime import datetime
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.http import HttpResponse
from django.test import Client, TestCase, signals
from django.urls import reverse
from django.utils import timezone

from openpyxl import load_workbook

from pos.models import Caja, GastoCaja, ItemVenta, Producto, Venta


# Evitar problemas al copiar contextos instrumentados en tests
signals.template_rendered.receivers = []


class ReportesTestCase(TestCase):
    """Tests de congruencia para la vista de reportes y sus exportaciones"""

    def setUp(self):
        self.user_admin = User.objects.create_user(
            username='admin_test',
            password='testpass123',
            is_staff=True,
        )
        self.user_vendedor = User.objects.create_user(
            username='vendedor_test',
            password='testpass123',
            is_staff=False,
        )

        grupo_admin, _ = Group.objects.get_or_create(name='Administradores')
        self.user_admin.groups.add(grupo_admin)

        self.caja = Caja.objects.create(numero=1, nombre='Caja Principal', activa=True)

        self.prod_a = Producto.objects.create(
            codigo='PROD_A',
            nombre='Producto A',
            precio=1000,
            stock=100,
            activo=True,
        )
        self.prod_b = Producto.objects.create(
            codigo='PROD_B',
            nombre='Producto B',
            precio=2000,
            stock=100,
            activo=True,
        )

        self.client = Client()
        self.client.force_login(self.user_admin)

        # Rango de prueba (mismo día, horas controladas)
        tz = timezone.get_current_timezone()
        self.fecha_desde = datetime(2025, 12, 14, 0, 0, 0, tzinfo=tz).date()
        self.fecha_hasta = datetime(2025, 12, 14, 23, 59, 59, tzinfo=tz).date()

        self.inicio_dt = timezone.make_aware(datetime(2025, 12, 14, 0, 0, 0), tz)
        self.fin_dt = timezone.make_aware(datetime(2025, 12, 14, 23, 59, 59), tz)

        # Ventas dentro del rango
        self.v1 = Venta.objects.create(
            fecha=timezone.make_aware(datetime(2025, 12, 14, 10, 0, 0), tz),
            total=10000,
            completada=True,
            anulada=False,
            metodo_pago='efectivo',
            monto_recibido=10000,
            usuario=self.user_admin,
            vendedor=self.user_vendedor,
            caja=self.caja,
            registradora_id=1,
        )
        ItemVenta.objects.create(
            venta=self.v1, producto=self.prod_a, cantidad=2, precio_unitario=1000, subtotal=2000
        )
        ItemVenta.objects.create(
            venta=self.v1, producto=self.prod_b, cantidad=4, precio_unitario=2000, subtotal=8000
        )

        self.v2 = Venta.objects.create(
            fecha=timezone.make_aware(datetime(2025, 12, 14, 12, 0, 0), tz),
            total=20000,
            completada=True,
            anulada=False,
            metodo_pago='tarjeta',
            monto_recibido=20000,
            usuario=self.user_admin,
            vendedor=self.user_admin,
            caja=self.caja,
            registradora_id=1,
        )
        ItemVenta.objects.create(
            venta=self.v2, producto=self.prod_a, cantidad=20, precio_unitario=1000, subtotal=20000
        )

        self.v3_anulada = Venta.objects.create(
            fecha=timezone.make_aware(datetime(2025, 12, 14, 15, 0, 0), tz),
            total=5000,
            completada=True,
            anulada=True,
            metodo_pago='efectivo',
            monto_recibido=5000,
            usuario=self.user_admin,
            vendedor=self.user_admin,
            caja=self.caja,
            registradora_id=1,
        )

        # Venta fuera del rango (no debe contar)
        Venta.objects.create(
            fecha=timezone.make_aware(datetime(2025, 12, 13, 12, 0, 0), tz),
            total=99999,
            completada=True,
            anulada=False,
            metodo_pago='efectivo',
            monto_recibido=99999,
            usuario=self.user_admin,
            vendedor=self.user_admin,
            caja=self.caja,
            registradora_id=1,
        )

        # Movimientos de caja (gastos/ingresos/retiros) dentro del rango
        GastoCaja.objects.create(
            tipo='gasto',
            monto=3000,
            descripcion='Gasto operativo',
            fecha=timezone.make_aware(datetime(2025, 12, 14, 11, 0, 0), tz),
            usuario=self.user_admin,
        )
        GastoCaja.objects.create(
            tipo='ingreso',
            monto=1000,
            descripcion='Ingreso extra',
            fecha=timezone.make_aware(datetime(2025, 12, 14, 11, 30, 0), tz),
            usuario=self.user_admin,
        )
        # Retiro: debe NO entrar en total_gastos, pero SÍ en total_retiros
        GastoCaja.objects.create(
            tipo='gasto',
            monto=7000,
            descripcion='Retiro de dinero al cerrar caja (Efectivo)',
            fecha=timezone.make_aware(datetime(2025, 12, 14, 16, 0, 0), tz),
            usuario=self.user_admin,
        )

    def test_reportes_contexto_congruente(self):
        """Los totales del contexto deben ser congruentes con los datos del rango."""
        url = reverse('pos:reportes')
        # Nota: en este entorno (Python 3.14 + Django) hay un bug al copiar contextos
        # instrumentados en el test client al renderizar templates. Para poder validar
        # congruencia del contexto de forma determinística, "mockeamos" render().
        def _fake_render(request, template_name, context):
            response = HttpResponse("OK")
            response._reportes_context = context
            return response

        with patch('pos.views.render', side_effect=_fake_render):
            resp = self.client.get(url, {
                'fecha_desde': self.fecha_desde.isoformat(),
                'fecha_hasta': self.fecha_hasta.isoformat(),
            })
        self.assertEqual(resp.status_code, 200)

        ctx = resp._reportes_context

        # Ventas
        self.assertEqual(ctx['total_ventas'], 30000)  # v1 + v2
        self.assertEqual(ctx['cantidad_ventas'], 2)
        self.assertEqual(ctx['total_anuladas'], 5000)  # v3
        self.assertEqual(ctx['cantidad_anuladas'], 1)
        self.assertEqual(ctx['ventas_efectivo'], 10000)  # solo v1 (v3 anulada no cuenta)
        self.assertEqual(ctx['ventas_tarjeta'], 20000)
        self.assertEqual(ctx['ventas_transferencia'], 0)
        self.assertEqual(ctx['dinero_bancos'], 20000)

        # Movimientos (retiro separado)
        self.assertEqual(ctx['total_gastos'], 3000)
        self.assertEqual(ctx['total_ingresos'], 1000)
        self.assertEqual(ctx['total_retiros'], 7000)

        # Resumen diario
        resumen_diario = ctx.get('resumen_diario') or []
        self.assertEqual(len(resumen_diario), 1)
        d0 = resumen_diario[0]
        self.assertEqual(d0['ventas_total'], 30000)
        self.assertEqual(d0['ventas_cantidad'], 2)
        self.assertEqual(d0['anuladas_total'], 5000)
        self.assertEqual(d0['anuladas_cantidad'], 1)
        self.assertEqual(d0['ventas_efectivo'], 10000)
        self.assertEqual(d0['ventas_tarjeta'], 20000)
        self.assertEqual(d0['ventas_transferencia'], 0)
        self.assertEqual(d0['gastos_sin_retiro_total'], 3000)
        self.assertEqual(d0['ingresos_total'], 1000)
        self.assertEqual(d0['retiros_total'], 7000)
        self.assertEqual(d0['neto_operativo'], 10000 + 1000 - 3000)
        self.assertEqual(d0['neto_despues_retiros'], 10000 + 1000 - 3000 - 7000)

        # Resúmenes
        resumen_usuarios = list(ctx['resumen_por_usuario'])
        self.assertTrue(any(r['usuario__username'] == 'admin_test' and int(r['total']) == 30000 for r in resumen_usuarios))

        resumen_vendedores = list(ctx['resumen_por_vendedor'])
        self.assertTrue(any(r['vendedor__username'] == 'vendedor_test' and int(r['total']) == 10000 for r in resumen_vendedores))
        self.assertTrue(any(r['vendedor__username'] == 'admin_test' and int(r['total']) == 20000 for r in resumen_vendedores))

    def test_export_csv_ventas_incluye_items(self):
        """El CSV de ventas debe incluir columnas de items y contenido esperado."""
        url = reverse('pos:reportes')
        resp = self.client.get(url, {
            'fecha_desde': self.fecha_desde.isoformat(),
            'fecha_hasta': self.fecha_hasta.isoformat(),
            'export': 'ventas',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn('text/csv', resp.get('Content-Type', ''))

        content = resp.content.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        self.assertGreaterEqual(len(rows), 2)
        header = rows[0]
        self.assertIn('Items Cantidad', header)
        self.assertIn('Items Detalle', header)

        # Debe existir una fila de v1 con items_cantidad = 6 (2+4)
        row_v1 = next((r for r in rows[1:] if r and r[0] == str(self.v1.id)), None)
        self.assertIsNotNone(row_v1)
        self.assertEqual(int(row_v1[9]), 6)
        self.assertIn('Producto A x2', row_v1[10])
        self.assertIn('Producto B x4', row_v1[10])

    def test_export_xlsx_ventas_con_hoja_items(self):
        """El XLSX de ventas debe incluir hojas 'Ventas' y 'Items' con headers correctos."""
        url = reverse('pos:reportes')
        resp = self.client.get(url, {
            'fecha_desde': self.fecha_desde.isoformat(),
            'fecha_hasta': self.fecha_hasta.isoformat(),
            'export': 'ventas',
            'format': 'xlsx',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resp.get('Content-Type', ''),
        )

        wb = load_workbook(filename=io.BytesIO(resp.content))
        self.assertIn('Ventas', wb.sheetnames)
        self.assertIn('Items', wb.sheetnames)

        ws_ventas = wb['Ventas']
        headers = [cell.value for cell in ws_ventas[1]]
        self.assertIn('Items Cantidad', headers)
        self.assertIn('Items Detalle', headers)

        ws_items = wb['Items']
        headers_items = [cell.value for cell in ws_items[1]]
        self.assertEqual(headers_items[:3], ['VentaID', 'Fecha', 'Producto'])


